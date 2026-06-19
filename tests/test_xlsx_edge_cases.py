"""Edge-case and regression tests for the xlsx tool.

Covers number-format defaults, the currency integer variant, thousands-separator
parsing, circular-reference detection without recalc, recalc skip-reason
surfacing, unresolved-reference warnings, sheet-name collision warnings,
the formula-reference pipeline (cross-sheet ranges, current-row vs
table-relative refs), formula references inside typed columns, and
comma-safe parsing of the types directive.
"""

from __future__ import annotations

import io
import logging
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from xlsx_tools.base_xlsx_tool import markdown_to_excel  # noqa: E402
from xlsx_tools.helpers import (  # noqa: E402
    DEFAULT_NUMBER_FORMAT,
    DEFAULT_NUMBER_FORMAT_DECIMALS,
    _currency_base_format,
    _strip_thousands_separators,
    _default_number_format_for,
)

OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


def _intercept_upload(markdown: str, **kwargs) -> bytes:
    """Run markdown_to_excel with upload patched out; return raw xlsx bytes."""
    captured = {}

    def fake_upload(file_obj, suffix, **kw):
        captured["data"] = file_obj.read()
        return "fake://test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown, **kwargs)
    return captured["data"]


# ── Number-format defaults (dash/parens in financial mode, and large
#    numbers preserve decimals) ─────────────────────────────────────────────


class TestNumberFormatDefaults:
    """Plain numeric cells preserve decimals instead of force-rounding."""

    def test_large_whole_number_uses_integer_format(self):
        data = _intercept_upload(
            "| Revenue |\n|---|\n| 1000000 |\n"
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1000000
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT  # "#,##0"

    def test_large_non_whole_number_preserves_decimals(self):
        """A large non-whole value keeps its decimals (1500.75 must not
        display as 1,501)."""
        data = _intercept_upload(
            "| Price |\n|---|\n| 1500.75 |\n"
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1500.75
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT_DECIMALS  # "#,##0.00"

    def test_small_whole_number_no_force_format(self):
        """Small whole numbers still get the integer default."""
        data = _intercept_upload("| Count |\n|---|\n| 5 |\n")
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 5
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT


class TestNumberFormatDefaultsHelper:
    """_default_number_format_for picks the right format by magnitude."""

    def test_whole_number_gets_integer_format(self):
        assert _default_number_format_for(1000.0) == DEFAULT_NUMBER_FORMAT

    def test_decimal_number_gets_decimal_format(self):
        assert _default_number_format_for(1000.5) == DEFAULT_NUMBER_FORMAT_DECIMALS

    def test_plain_percent_default_format(self):
        """Without financial mode, plain percent stays at 0.0%."""
        data = _intercept_upload("| Growth |\n|---|\n| 10% |\n")
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == "0.0%"


# ── Zero-decimal (integer) currency variant ──────────────────────────────────


class TestCurrencyIntegerVariant:
    def test_currency_integer_variant_picks_zero_decimal(self):
        assert _currency_base_format("$", "integer") == "$#,##0"
        assert _currency_base_format("$", "int") == "$#,##0"
        assert _currency_base_format("$", "whole") == "$#,##0"

    def test_currency_default_stays_two_decimals(self):
        assert _currency_base_format("$", None) == "$#,##0.00"
        assert _currency_base_format("$", "dash") == "$#,##0.00"

    def test_currency_integer_with_dash_variant(self):
        """integer + dash combines: zero-decimal base WITH the dash sections
        (zeros render as '-', negatives in parens)."""
        from xlsx_tools.helpers import _apply_format_variant
        base = _currency_base_format("$", "integer:dash")
        assert base == "$#,##0"
        assert _apply_format_variant(base, "integer:dash") == "$#,##0;($#,##0);-"

    def test_currency_integer_euro(self):
        assert _currency_base_format("€", "integer") == "#,##0 €"

    def test_currency_integer_unknown_symbol_fallback(self):
        assert _currency_base_format("₿", "integer") == '#,##0 "₿"'

    def test_currency_integer_end_to_end(self):
        markdown = """<!-- types: currency:$:integer -->
| Revenue ($mm) |
|---|
| $1500 |
"""
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1500.0
        assert ws["A2"].number_format == "$#,##0"


# ── Thousands separators in default numeric path ─────────────────────────────


class TestThousandsSeparatorParsing:
    def test_strip_thousands_english(self):
        assert _strip_thousands_separators("1,234.56") == "1234.56"

    def test_strip_thousands_european(self):
        assert _strip_thousands_separators("1.234,56") == "1234.56"

    def test_strip_thousands_bare(self):
        assert _strip_thousands_separators("1,234") == "1234"
        assert _strip_thousands_separators("1,5") == "1.5"  # decimal comma

    def test_strip_thousands_no_separators(self):
        assert _strip_thousands_separators("1234") == "1234"

    def test_plain_cell_with_commas_parses_as_number(self):
        """'1,234' parses as a number rather than staying a string."""
        data = _intercept_upload(
            "| Population |\n|---|\n| 1,234 |\n"
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1234.0
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT

    def test_plain_cell_with_commas_and_decimals(self):
        data = _intercept_upload(
            "| GDP |\n|---|\n| 1,234.56 |\n"
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1234.56
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT_DECIMALS


# ── Circular detection runs regardless of recalc setting ─────────────────────


class TestCircularDetectionWithoutRecalc:
    def test_circular_ref_detected_when_recalc_disabled(self):
        """Cycle detection runs even when XLSX_RECALC_ENABLED=false. Verify the
        standalone detection path logs the cycle."""
        from unittest.mock import MagicMock
        from config import Config

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        fake_cfg = MagicMock(spec=Config)
        fake_cfg.xlsx_recalc_enabled = False
        fake_cfg.xlsx_recalc_strict = False
        fake_cfg.xlsx_default_font = None
        fake_cfg.xlsx_recalc_timeout_seconds = 30

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch("xlsx_tools.base_xlsx_tool.get_config", return_value=fake_cfg):
            mock_upload.return_value = "fake://circ.xlsx"
            with self._capture_logs("xlsx_tools.base_xlsx_tool") as logs:
                result = markdown_to_excel(markdown)
        assert result == "fake://circ.xlsx"
        joined = "\n".join(logs)
        assert "circular" in joined.lower() or "#CIRC" in joined

    def test_circular_ref_not_fatal_by_default(self):
        """Cycles are logged but NOT fatal when XLSX_RECALC_STRICT is off (default)."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ.xlsx"
            # Must NOT raise (strict is off by default).
            result = markdown_to_excel(markdown)
        assert result == "fake://circ.xlsx"

    def test_circular_ref_fatal_when_strict_enabled(self):
        """With XLSX_RECALC_STRICT=true, cycles fail the call."""
        from unittest.mock import MagicMock
        from config import Config

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        fake_cfg = MagicMock(spec=Config)
        fake_cfg.xlsx_recalc_enabled = True
        fake_cfg.xlsx_recalc_strict = True
        fake_cfg.xlsx_default_font = None
        fake_cfg.xlsx_recalc_timeout_seconds = 30

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch("xlsx_tools.base_xlsx_tool.get_config", return_value=fake_cfg):
            mock_upload.return_value = "fake://circ.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown)
        assert "#CIRC" in str(exc_info.value) or "circular" in str(exc_info.value).lower()

    @staticmethod
    def _capture_logs(logger_name: str):
        """Context manager capturing log records from a named logger."""
        from contextlib import contextmanager

        records: list[str] = []

        @contextmanager
        def _cm():
            lg = logging.getLogger(logger_name)
            handler = logging.Handler()
            handler.emit = lambda record: records.append(
                f"{record.levelname}: {record.getMessage()}"
            )
            lg.addHandler(handler)
            try:
                yield records
            finally:
                lg.removeHandler(handler)

        return _cm()


# ── Recalc skip reason surfaced ──────────────────────────────────────────────


class TestRecalcSkipReasonSurfaced:
    def test_skip_reason_surfaced_when_strict_enabled(self):
        """When XLSX_RECALC_STRICT=true but the engine can't complete, the
        skip reason should surface in the raised error (so the model can
        rewrite the offending formula)."""
        from unittest.mock import MagicMock
        from config import Config
        from xlsx_tools.formula_engine import RecalcResult

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| 1 | =A2*2 |\n"
        )
        fake_result = RecalcResult(
            recalc_performed=False,
            skip_reason="recalculation engine error: NotImplementedError: XLOOKUP",
        )
        fake_cfg = MagicMock(spec=Config)
        fake_cfg.xlsx_recalc_enabled = True
        fake_cfg.xlsx_recalc_strict = True
        fake_cfg.xlsx_default_font = None
        fake_cfg.xlsx_recalc_timeout_seconds = 30

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch("xlsx_tools.base_xlsx_tool.get_config", return_value=fake_cfg), \
             patch(
                 "xlsx_tools.formula_engine.recalculate_workbook",
                 return_value=fake_result,
             ):
            mock_upload.return_value = "fake://skip.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown)
        msg = str(exc_info.value)
        assert "XLOOKUP" in msg or "could not be completed" in msg.lower()

    def test_skip_reason_not_fatal_when_strict_disabled(self):
        """With XLSX_RECALC_STRICT=false (default), skip reason is logged but file delivered."""
        from xlsx_tools.formula_engine import RecalcResult

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| 1 | =A2*2 |\n"
        )
        fake_result = RecalcResult(
            recalc_performed=False,
            skip_reason="recalculation engine error: oops",
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch(
                 "xlsx_tools.formula_engine.recalculate_workbook",
                 return_value=fake_result,
             ):
            mock_upload.return_value = "fake://skip-delivered.xlsx"
            # strict=False (default) → must not raise.
            result = markdown_to_excel(markdown)
        assert result == "fake://skip-delivered.xlsx"


# ── Unresolved reference warnings ────────────────────────────────────────────


class TestUnresolvedReferenceWarnings:
    def test_unknown_table_ref_warns(self, caplog):
        """A T99 reference (when only 1 table exists) should warn."""
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            adjust_formula_references("=T99.B[0]", current_excel_row=5)
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "T99" in joined

    def test_unknown_cross_sheet_warns(self, caplog):
        """A cross-sheet ref to a sheet that doesn't exist should warn."""
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            adjust_formula_references(
                "=Nope!T1.B[0]",
                current_excel_row=5,
                all_sheet_table_positions={"Real": {"T1": 1}},
            )
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "Nope" in joined
        assert "Real" in joined  # known sheets listed

    def test_known_table_ref_does_not_warn(self, caplog):
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            result = adjust_formula_references(
                "=T1.B[0]",
                current_excel_row=5,
                table_positions={"T1": 1},
            )
        warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert not warnings
        assert result == "=B2"

    def test_known_cross_sheet_does_not_warn(self, caplog):
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            result = adjust_formula_references(
                "=Real!T1.B[0]",
                current_excel_row=5,
                all_sheet_table_positions={"Real": {"T1": 1}},
            )
        warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert not warnings


# ── Sheet-name collision detection ───────────────────────────────────────────


class TestSheetNameCollisionWarning:
    def test_duplicate_sheet_name_warns(self, caplog):
        """Two identical sheet names should warn (openpyxl auto-renames)."""
        markdown = (
            "## Sheet: Model\n\n"
            "| A |\n|---|\n| 1 |\n\n"
            "## Sheet: Model\n\n"
            "| B |\n|---|\n| 2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://collision.xlsx"
            with caplog.at_level(logging.WARNING, logger="xlsx_tools.base_xlsx_tool"):
                markdown_to_excel(markdown)
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "collide" in joined.lower() or "Model" in joined

    def test_distinct_sheet_names_do_not_warn(self, caplog):
        markdown = (
            "## Sheet: Alpha\n\n"
            "| A |\n|---|\n| 1 |\n\n"
            "## Sheet: Beta\n\n"
            "| B |\n|---|\n| 2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://distinct.xlsx"
            with caplog.at_level(logging.WARNING, logger="xlsx_tools.base_xlsx_tool"):
                markdown_to_excel(markdown)
        collisions = [
            r for r in caplog.records
            if "collide" in r.getMessage().lower()
        ]
        assert not collisions


# ── Formula-reference pipeline ───────────────────────────────────────────────


class TestCrossSheetFunctionRangePrefix:
    """=SheetName!T1.SUM(B[0]:B[2]) must emit =SUM(Sheet!B2:B4), not the
    invalid =SUM(Sheet!B2:Sheet!B4) (which yields #VALUE!)."""

    def test_cross_sheet_sum_single_prefix(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.SUM(B[0]:B[2])", 2, {"T1": 1}, all_pos
        )
        assert result == "=SUM(Data!B2:B4)"

    def test_cross_sheet_average_single_prefix(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.AVERAGE(B[0]:B[2])", 2, {"T1": 1}, all_pos
        )
        assert result == "=AVERAGE(Data!B2:B4)"

    def test_cross_sheet_sum_end_to_end_recalcs(self):
        """End-to-end: the formula must recalc without #VALUE!."""
        markdown = (
            "## Sheet: Data\n\n"
            "| V |\n|---|\n| 10 |\n| 20 |\n| 30 |\n\n"
            "## Sheet: Summary\n\n"
            "| Total |\n|---|\n| =Data!T1.SUM(A[0]:A[2]) |\n"
        )
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        assert wb["Summary"]["A2"].value == 60

    def test_cross_sheet_range_still_correct(self):
        """The sibling cross-sheet range pattern stays correct — guard
        against a regression in the function pattern."""
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.B[0]:T1.B[2]", 2, {"T1": 1}, all_pos
        )
        assert result == "=Data!B2:B4"


class TestCurrentRowRelativeRefs:
    """B[0]/A[0] is current-row-relative (not table-relative); it must
    resolve to the row the formula lives on, for every data row."""

    def test_unit_current_row_offset_zero(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for cur in [2, 3, 4, 5]:
            result = adjust_formula_references("=B[0]", cur, table_positions=tp)
            assert result == f"=B{cur}", f"row {cur}: {result}"

    def test_unit_current_row_negative_offset(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        # B[-1] on row N → B(N-1); the most common running-total pattern.
        for cur in [2, 3, 4, 5]:
            result = adjust_formula_references("=B[-1]", cur, table_positions=tp)
            assert result == f"=B{cur - 1}", f"row {cur}: {result}"

    def test_unit_current_row_positive_offset(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for cur in [2, 3, 4]:
            result = adjust_formula_references("=B[1]", cur, table_positions=tp)
            assert result == f"=B{cur + 1}", f"row {cur}: {result}"

    def test_unit_range_pattern_uses_current_row(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        # =SUM(B[0]:E[0]) at row 4 → =SUM(B4:E4)
        result = adjust_formula_references("=SUM(B[0]:E[0])", 4, table_positions=tp)
        assert result == "=SUM(B4:E4)"
        # Range spanning rows: =SUM(B[-1]:B[1]) at row 3 → =SUM(B2:B4)
        result = adjust_formula_references("=SUM(B[-1]:B[1])", 3, table_positions=tp)
        assert result == "=SUM(B2:B4)"

    def test_table_relative_T1_still_correct(self):
        """The T1.B[n] form is table-relative (offset from first data row)
        and must be unaffected by the current-row resolution."""
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for n, expected_row in [(0, 2), (1, 3), (2, 4)]:
            result = adjust_formula_references(f"=T1.B[{n}]", 99, table_positions=tp)
            assert result == f"=B{expected_row}"

    def test_end_to_end_running_total(self):
        """The classic running-total pattern must produce correct cached values."""
        markdown = (
            "| Delta | RunningTotal |\n"
            "|---|---|\n"
            "| 10 | =A[0] |\n"
            "| 20 | =B[-1]+A[0] |\n"
            "| 30 | =B[-1]+A[0] |\n"
        )
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 10
        assert ws["B3"].value == 30
        assert ws["B4"].value == 60

    def test_end_to_end_per_row_multiply(self):
        """=A[0]*2 on each row must hit each row's A, not always row 2."""
        markdown = (
            "| Base | TwiceBase |\n"
            "|---|---|\n"
            "| 5 | =A[0]*2 |\n"
            "| 10 | =A[0]*2 |\n"
            "| 15 | =A[0]*2 |\n"
        )
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 10
        assert ws["B3"].value == 20
        assert ws["B4"].value == 30




# ── Formula references inside typed columns ──────────────────────────────────


class TestFormulaRefsInTypedColumns:
    """A formula using the relative/table-reference syntax (B[0], T1.B[0],
    Sheet!T1.B[0]) inside a `types`-directive column must have its
    references resolved, so the typed-column path matches the non-typed
    path (an unresolved B[-1] literal is #NAME? in Excel)."""

    def test_relative_ref_resolved_in_number_column(self):
        markdown = (
            "<!-- types: text, number -->\n"
            "| Item | Qty |\n|---|---|\n"
            "| A | 10 |\n| B | 20 |\n| Total | =SUM(B[-2]:B[-1]) |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        # Reference must be resolved, not left as the literal B[-2]:B[-1].
        assert wbf.active["B4"].value == "=SUM(B2:B3)"
        assert wbv.active["B4"].value == 30

    def test_table_ref_resolved_in_typed_column(self):
        markdown = (
            "<!-- types: text, number -->\n"
            "| Item | Qty |\n|---|---|\n"
            "| A | 10 |\n| Total | =T1.B[0] |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        # T1.B[0] → first data row of T1 = B2.
        assert wbf.active["B3"].value == "=B2"

    def test_cross_sheet_relative_ref_resolved_in_typed_column(self):
        markdown = (
            "## Sheet: Data\n\n"
            "| V |\n|---|\n| 10 |\n| 20 |\n\n"
            "## Sheet: Sum\n\n"
            "<!-- types: currency:$ -->\n"
            "| Total |\n|---|\n| =Data!T1.A[0] |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbf["Sum"]["A2"].value == "=Data!A2"
        assert wbv["Sum"]["A2"].value == 10
        # The column's currency format is applied to the formula result.
        assert wbf["Sum"]["A2"].number_format == "$#,##0.00"

    def test_absolute_ref_in_typed_column_still_works(self):
        """Guard against regression — absolute refs already worked."""
        markdown = (
            "<!-- types: text, number -->\n"
            "| Item | Qty |\n|---|---|\n"
            "| A | 10 |\n| Total | =SUM(B2:B2) |\n"
        )
        data = _intercept_upload(markdown)
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbv.active["B3"].value == 10

    def test_typed_relative_formula_does_not_abort_recalc(self):
        """With recalc enabled, a typed-column relative formula must resolve
        correctly and not fail the call (an unresolved '=SUM(B[-1])' would
        otherwise raise in the engine)."""
        markdown = (
            "<!-- types: text, currency:$ -->\n"
            "| Item | Amount |\n|---|---|\n"
            "| A | $100 |\n| B | $200 |\n| Total | =SUM(B[-2]:B[-1]) |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://typed.xlsx"
            # Must not raise.
            result = markdown_to_excel(markdown)
        assert result == "fake://typed.xlsx"

    def test_relative_ref_resolved_in_percent_column(self):
        """A formula in a `percent` column: refs resolved, percent format
        applied to the result."""
        markdown = (
            "<!-- types: text, percent -->\n"
            "| Metric | Rate |\n|---|---|\n"
            "| Y1 | 10% |\n| Y2 | 30% |\n| Avg | =AVERAGE(B[-2]:B[-1]) |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbf.active["B4"].value == "=AVERAGE(B2:B3)"
        # 10% and 30% stored as 0.1 / 0.3 → average 0.2.
        assert abs(wbv.active["B4"].value - 0.2) < 1e-9
        assert wbf.active["B4"].number_format == "0.0%"

    def test_relative_ref_resolved_in_multiple_column(self):
        """A formula in a `multiple` column: refs resolved, 0.0x format."""
        markdown = (
            "<!-- types: text, multiple -->\n"
            "| Comp | EV/EBITDA |\n|---|---|\n"
            "| A | 12.0x |\n| B | 10.0x |\n| Mean | =AVERAGE(B[-2]:B[-1]) |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbf.active["B4"].value == "=AVERAGE(B2:B3)"
        assert wbv.active["B4"].value == 11
        assert wbf.active["B4"].number_format == '0.0"x"'

    def test_relative_ref_resolved_in_bool_column(self):
        """A formula in a `bool` column must still resolve refs (bool type
        has no number format, so none is forced on the result)."""
        markdown = (
            "<!-- types: text, bool -->\n"
            "| Check | Flag |\n|---|---|\n"
            "| A | yes |\n| B | =NOT(B[-1]) |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbf.active["B3"].value == "=NOT(B2)"
        # B2 is TRUE → NOT(TRUE) = FALSE.
        assert wbv.active["B3"].value is False

    def test_relative_ref_resolved_in_date_column(self):
        """A formula in a `date` column must resolve refs; date format may
        be applied to the (serial) result."""
        markdown = (
            "<!-- types: text, date -->\n"
            "| Event | When |\n|---|---|\n"
            "| Start | 2024-01-01 |\n| NextDay | =B[-1]+1 |\n"
        )
        data = _intercept_upload(markdown)
        wbf = load_workbook(io.BytesIO(data))
        wbv = load_workbook(io.BytesIO(data), data_only=True)
        assert wbf.active["B3"].value == "=B2+1"
        # 2024-01-01 + 1 day → 2024-01-02 (read back as datetime via format).
        assert wbv.active["B3"].value is not None


# ── Comma-safe parsing of the types directive ────────────────────────────────


class TestTypesDirectiveCommaInFormat:
    """A comma inside a number/currency format (e.g. number:#,##0) must not
    be treated as a column separator — that would truncate the format and
    shift every later column by one, silently corrupting data."""

    def test_literal_comma_format_not_split(self):
        from xlsx_tools.helpers import _parse_types_directive
        assert _parse_types_directive("number:#,##0, text") == [
            "number:#,##0",
            "text",
        ]

    def test_multi_column_comma_format(self):
        from xlsx_tools.helpers import _parse_types_directive
        assert _parse_types_directive(
            "text, currency:$, number:#,##0.00, percent"
        ) == ["text", "currency:$", "number:#,##0.00", "percent"]

    def test_multi_section_comma_format(self):
        from xlsx_tools.helpers import _parse_types_directive
        assert _parse_types_directive("number:#,##0;(#,##0);-, text") == [
            "number:#,##0;(#,##0);-",
            "text",
        ]

    def test_variant_specs_unaffected(self):
        """Comma-free variant specs must keep working unchanged."""
        from xlsx_tools.helpers import _parse_types_directive
        assert _parse_types_directive(
            "text, currency:$:integer:dash, multiple, percent:integer"
        ) == ["text", "currency:$:integer:dash", "multiple", "percent:integer"]

    def test_comma_format_end_to_end_preserves_format_and_columns(self):
        markdown = (
            "<!-- types: number:#,##0, text -->\n"
            "| Amount | Code |\n|---|---|\n"
            "| 1234 | 007 |\n"
        )
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        # Format must survive intact (must not be truncated to '#').
        assert ws["A2"].number_format == "#,##0"
        assert ws["A2"].value == 1234
        # The `text` column must still be text — leading zero preserved
        # (a column shift would turn '007' into the number 7).
        assert ws["B2"].value == "007"


# ── Currency precision + section-style variant combination ───────────────────


class TestCurrencyIntegerSectionVariant:
    """currency:$:integer:dash (and :parens) must apply BOTH the zero-decimal
    precision AND the dash/parens sections — the section token must not be
    dropped when combined with the integer precision keyword."""

    def test_integer_dash_combines(self):
        from xlsx_tools.helpers import _currency_base_format, _apply_format_variant
        base = _currency_base_format("$", "integer:dash")
        assert _apply_format_variant(base, "integer:dash") == "$#,##0;($#,##0);-"

    def test_integer_parens_combines(self):
        from xlsx_tools.helpers import _currency_base_format, _apply_format_variant
        base = _currency_base_format("$", "integer:parens")
        assert _apply_format_variant(base, "integer:parens") == "$#,##0;($#,##0)"

    def test_plain_dash_still_two_decimals(self):
        from xlsx_tools.helpers import _currency_base_format, _apply_format_variant
        base = _currency_base_format("$", "dash")
        assert _apply_format_variant(base, "dash") == "$#,##0.00;($#,##0.00);-"

    def test_integer_dash_end_to_end(self):
        markdown = (
            "<!-- types: currency:$:integer:dash -->\n"
            "| Revenue ($mm) |\n|---|\n| $1500 |\n"
        )
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1500.0
        assert ws["A2"].number_format == "$#,##0;($#,##0);-"


# ── Typed column format defaults ─────────────────────────────────────────────


class TestTypedColumnFormats:
    """Typed columns produce the expected formats without any financial flag."""

    def test_typed_number_uses_plain_format(self):
        """Without dash/parens variant, typed `number` columns keep the plain format."""
        markdown = "<!-- types: number -->\n| X |\n|---|\n| 100000 |\n"
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == "#,##0"

    def test_typed_number_with_explicit_dash_variant(self):
        markdown = "<!-- types: number:dash -->\n| X |\n|---|\n| 100000 |\n"
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == "#,##0;(#,##0);-"

    def test_typed_number_literal_format_respected(self):
        markdown = "<!-- types: number:#,##0.000 -->\n| X |\n|---|\n| 1.5 |\n"
        data = _intercept_upload(markdown)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == "#,##0.000"

