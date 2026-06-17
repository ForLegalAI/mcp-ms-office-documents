"""Tests for financial-modeling features of the xlsx tool.

Covers:
- ``apply_financial_styling`` — CFA color coding (blue inputs, black
  local formulas, green cross-sheet refs, yellow sourced cells).
- ``apply_default_font`` — font family propagation, monospace preservation.
- ``parse_sources_directive`` — source-citation parsing including ranges.
- Number-format variants: ``number:dash``, ``currency:$:parens``,
  ``percent:dash`` etc.
- ``years-as-text`` convention when ``financial_modeling=True``.
- Cell comments for source citations.
- End-to-end via ``markdown_to_excel`` with the new parameters.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from xlsx_tools.helpers import (  # noqa: E402
    FINANCIAL_INPUT_COLOR,
    FINANCIAL_FORMULA_COLOR,
    FINANCIAL_CROSS_SHEET_COLOR,
    FINANCIAL_ASSUMPTION_FILL,
    NUMBER_FORMAT_VARIANTS,
    PERCENT_FORMAT_VARIANTS,
    _apply_format_variant,
    _apply_percent_format_variant,
    _expand_coord_range,
    _is_year_string,
    apply_default_font,
    apply_financial_styling,
    attach_source_comment,
    parse_sources_directive,
)
from xlsx_tools.base_xlsx_tool import markdown_to_excel  # noqa: E402

OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


# ── Color coding ─────────────────────────────────────────────────────────────


class TestFinancialColorCoding:
    """apply_financial_styling picks the right color per cell kind."""

    def _make_cell(self, value=None) -> Cell:
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        if value is not None:
            cell.value = value
        return cell

    def test_literal_value_gets_blue_input_color(self):
        cell = self._make_cell(1000)
        apply_financial_styling(cell, "1000")
        # openpyxl stores color as RGB string; the leading "FF" is alpha.
        rgb = cell.font.color.rgb if cell.font.color else None
        assert rgb is not None
        assert rgb.endswith(FINANCIAL_INPUT_COLOR)

    def test_local_formula_gets_black_color(self):
        cell = self._make_cell("=SUM(B2:B5)")
        apply_financial_styling(cell, "=SUM(B2:B5)")
        rgb = cell.font.color.rgb
        assert rgb.endswith(FINANCIAL_FORMULA_COLOR)

    def test_cross_sheet_formula_gets_green_color(self):
        cell = self._make_cell("=Inputs!B2")
        apply_financial_styling(cell, "=Inputs!B2")
        rgb = cell.font.color.rgb
        assert rgb.endswith(FINANCIAL_CROSS_SHEET_COLOR)

    def test_sourced_cell_gets_yellow_fill(self):
        cell = self._make_cell(1000)
        apply_financial_styling(cell, "1000", source_cells={"A1"})
        fill = cell.fill
        assert fill.fill_type == "solid"
        assert fill.fgColor.rgb.endswith(FINANCIAL_ASSUMPTION_FILL)

    def test_non_sourced_cell_no_fill_change(self):
        cell = self._make_cell(1000)
        # Apply with empty source set — fill should remain default (none).
        apply_financial_styling(cell, "1000", source_cells=set())
        assert cell.fill.fill_type is None


# ── Default font ─────────────────────────────────────────────────────────────


class TestDefaultFont:
    """apply_default_font propagates family but preserves monospace."""

    def test_font_family_applied(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = "hello"
        apply_default_font(cell, "Arial")
        assert cell.font.name == "Arial"

    def test_monospace_font_preserved(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        # Pre-set a monospace font (as inline code formatting would do).
        from openpyxl.styles import Font
        cell.font = Font(name="Courier New")
        apply_default_font(cell, "Arial")
        assert cell.font.name == "Courier New"

    def test_none_font_no_change(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        original_name = cell.font.name
        apply_default_font(cell, None)
        assert cell.font.name == original_name


# ── Source-citation directive parser ─────────────────────────────────────────


class TestSourcesDirective:
    """parse_sources_directive handles per-cell and range forms."""

    def test_per_cell_entries(self):
        result = parse_sources_directive(
            "B2=Source: Annual Report 2024, B5=Source: Internal forecast"
        )
        assert result == {
            "B2": "Source: Annual Report 2024",
            "B5": "Source: Internal forecast",
        }

    def test_range_form_expands(self):
        result = parse_sources_directive("B2:B4=Same source applies")
        assert result == {
            "B2": "Same source applies",
            "B3": "Same source applies",
            "B4": "Same source applies",
        }

    def test_empty_value(self):
        assert parse_sources_directive("") == {}

    def test_malformed_entries_ignored(self):
        result = parse_sources_directive("B2=valid, garbage, =nokey, B5=")
        assert result == {"B2": "valid"}

    def test_range_expander_helper(self):
        assert _expand_coord_range("B2:B5") == ["B2", "B3", "B4", "B5"]
        assert _expand_coord_range("B5:B2") == ["B2", "B3", "B4", "B5"]  # reversed ok

    def test_attach_source_comment(self):
        wb = Workbook()
        cell = wb.active["A1"]
        attach_source_comment(cell, "Source: 10-K filing 2024")
        assert isinstance(cell.comment, Comment)
        assert "10-K filing 2024" in cell.comment.text

    def test_attach_source_comment_empty_does_nothing(self):
        wb = Workbook()
        cell = wb.active["A1"]
        attach_source_comment(cell, "")
        assert cell.comment is None


# ── Number-format variants ───────────────────────────────────────────────────


class TestFormatVariants:
    """Variants (dash, parens) produce the right Excel format strings."""

    def test_number_dash_format(self):
        assert NUMBER_FORMAT_VARIANTS["dash"] == "#,##0;(#,##0);-"

    def test_number_parens_format(self):
        assert NUMBER_FORMAT_VARIANTS["parens"] == "#,##0;(#,##0)"

    def test_currency_with_variant(self):
        result = _apply_format_variant("$#,##0.00", "dash")
        assert result == "$#,##0.00;($#,##0.00);-"

    def test_currency_with_parens_variant(self):
        result = _apply_format_variant("€#,##0.00", "parens")
        assert result == "€#,##0.00;(€#,##0.00)"

    def test_percent_dash_variant(self):
        assert _apply_percent_format_variant("dash") == "0.0%;(0.0%);-"

    def test_percent_no_variant_returns_legacy_format(self):
        """Preserving backward compat: bare 'percent' directive → 0%."""
        assert _apply_percent_format_variant(None) == "0%"


# ── Years-as-text ────────────────────────────────────────────────────────────


class TestYearsAsText:
    """_is_year_string detects 4-digit years only."""

    def test_four_digit_year(self):
        assert _is_year_string("2024") is True
        assert _is_year_string("1999") is True

    def test_two_digit_year_rejected(self):
        assert _is_year_string("24") is False

    def test_non_year_text_rejected(self):
        assert _is_year_string("Revenue") is False
        assert _is_year_string("2024a") is False


# ── End-to-end via markdown_to_excel ─────────────────────────────────────────


def _intercept_upload(markdown: str, **kwargs) -> bytes:
    captured = {}

    def fake_upload(file_obj, suffix, **kw):
        captured["data"] = file_obj.read()
        return "fake://test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown, **kwargs)
    return captured["data"]


class TestEndToEndFinancial:
    """markdown_to_excel applies financial features when requested."""

    def test_financial_modeling_colors_inputs_and_formulas(self):
        markdown = """| Input | Formula |
|-------|---------|
| 100   | =A2*2   |
"""
        data = _intercept_upload(markdown, financial_modeling=True, recalc=False)
        (OUTPUT_DIR / "financial_colors.xlsx").write_bytes(data)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # A2 is the literal input "100" → blue.
        input_rgb = ws["A2"].font.color.rgb
        assert input_rgb.endswith(FINANCIAL_INPUT_COLOR)
        # B2 is a local formula → black.
        formula_rgb = ws["B2"].font.color.rgb
        assert formula_rgb.endswith(FINANCIAL_FORMULA_COLOR)

    def test_financial_modeling_treats_years_as_text(self):
        markdown = """| Year | Revenue |
|------|---------|
| 2024 | 1000    |
| 2025 | 1100    |
"""
        data = _intercept_upload(markdown, financial_modeling=True, recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # With financial_modeling, the year cell stays text.
        assert ws["A2"].value == "2024"
        assert isinstance(ws["A2"].value, str)

    def test_financial_modeling_off_converts_years_to_numbers(self):
        """Without financial_modeling, 2024 becomes a number (default behaviour)."""
        markdown = """| Year | Revenue |
|------|---------|
| 2024 | 1000    |
"""
        data = _intercept_upload(markdown, financial_modeling=False, recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Without financial modeling, default conversion applies.
        # openpyxl may store as int or float depending on parsing.
        assert ws["A2"].value == 2024 or ws["A2"].value == 2024.0

    def test_default_font_applied_to_all_cells(self):
        markdown = """| Col |
|-----|
| abc |
"""
        data = _intercept_upload(markdown, default_font="Times New Roman", recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Header cell and data cell both get the chosen family.
        assert ws["A1"].font.name == "Times New Roman"
        assert ws["A2"].font.name == "Times New Roman"

    def test_source_directive_attaches_comment(self):
        markdown = """<!-- sources: B2=Source: Annual Report 2024 -->
| Item | Value |
|------|-------|
| A    | 100   |
"""
        data = _intercept_upload(markdown, recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        cell = ws["B2"]
        assert cell.comment is not None
        assert "Annual Report 2024" in cell.comment.text

    def test_source_directive_yellow_fill_with_financial(self):
        markdown = """<!-- sources: B2=Source: Annual Report -->
| Item | Value |
|------|-------|
| A    | 100   |
"""
        data = _intercept_upload(markdown, financial_modeling=True, recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        fill = ws["B2"].fill
        assert fill.fill_type == "solid"
        assert fill.fgColor.rgb.endswith(FINANCIAL_ASSUMPTION_FILL)

    def test_number_dash_variant_in_types_directive(self):
        markdown = """<!-- types: number:dash -->
| Value |
|-------|
| 0     |
| -5    |
"""
        data = _intercept_upload(markdown, recalc=False)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Both data cells should have the dash variant number format.
        assert ws["A2"].number_format == "#,##0;(#,##0);-"
        assert ws["A3"].number_format == "#,##0;(#,##0);-"

    def test_currency_with_parens_variant(self):
        markdown = """<!-- types: currency:$:parens -->
| Amount |
|--------|
| $100   |
| ($50)  |
"""
        data = _intercept_upload(markdown, recalc=False)
        (OUTPUT_DIR / "currency_parens.xlsx").write_bytes(data)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].number_format == "$#,##0.00;($#,##0.00)"
        assert ws["A3"].number_format == "$#,##0.00;($#,##0.00)"
        # Values parsed as numbers.
        assert ws["A2"].value == 100.0

    def test_combined_financial_features(self):
        """A realistic financial-model markdown uses colors, sources, and dash variant together."""
        markdown = """<!-- types: text, number:dash -->
<!-- sources: B2=Source: 10-K 2024, B3=Source: 10-K 2024 -->
| Year | Revenue |
|------|---------|
| 2024 | 1000000 |
| 2025 | 1100000 |
| Sum  | =B2+B3  |
"""
        data = _intercept_upload(
            markdown, financial_modeling=True, recalc=True,
        )
        (OUTPUT_DIR / "financial_combined.xlsx").write_bytes(data)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # Years stay as text.
        assert ws["A2"].value == "2024"
        assert ws["A3"].value == "2025"

        # Source cells get yellow background.
        assert ws["B2"].fill.fill_type == "solid"
        assert ws["B3"].fill.fill_type == "solid"

        # Source comments present.
        assert ws["B2"].comment is not None
        assert "10-K 2024" in ws["B2"].comment.text

        # Number format applied.
        assert ws["B2"].number_format == "#,##0;(#,##0);-"

        # Sum formula recalc'd to a cached value (financial_modeling doesn't disable recalc).
        wb_cached = load_workbook(io.BytesIO(data), data_only=True)
        assert wb_cached.active["B4"].value == 2100000


# ── Round 2 additions ────────────────────────────────────────────────────────


class TestMultiplesFormat:
    """A1: valuation multiples render as '12.5x' via the multiple type."""

    def test_multiples_format_variants_dict(self):
        from xlsx_tools.helpers import MULTIPLES_FORMAT_VARIANTS
        assert MULTIPLES_FORMAT_VARIANTS["default"] == '0.0"x"'
        assert MULTIPLES_FORMAT_VARIANTS["dash"] == '0.0"x";(0.0"x");-'
        assert MULTIPLES_FORMAT_VARIANTS["parens"] == '0.0"x";(0.0"x")'

    def test_apply_multiples_format_variant(self):
        from xlsx_tools.helpers import _apply_multiples_format_variant
        assert _apply_multiples_format_variant(None) == '0.0"x"'
        assert _apply_multiples_format_variant("dash") == '0.0"x";(0.0"x");-'
        assert _apply_multiples_format_variant("default") == '0.0"x"'
        # Unknown variant falls back to default.
        assert _apply_multiples_format_variant("bogus") == '0.0"x"'

    def test_multiple_type_parses_value(self):
        """The multiple type stores the raw number (no scaling)."""
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook

        for raw, expected in [("12.5", 12.5), ("12.5x", 12.5), ("8", 8.0)]:
            wb = Workbook()
            cell = wb.active["A1"]
            _apply_column_type(cell, raw, "multiple")
            assert cell.value == expected
            assert cell.number_format == '0.0"x"'

    def test_multiple_type_with_dash_variant(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook

        wb = Workbook()
        cell = wb.active["A1"]
        _apply_column_type(cell, "12.5", "multiple:dash")
        assert cell.value == 12.5
        assert cell.number_format == '0.0"x";(0.0"x");-'

    def test_multiple_type_unparseable_keeps_text(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook

        wb = Workbook()
        cell = wb.active["A1"]
        _apply_column_type(cell, "not a number", "multiple")
        assert cell.value == "not a number"

    def test_multiple_type_e2e_via_markdown(self):
        """End-to-end: types directive with 'multiple' produces '12.5x' display."""
        markdown = """<!-- types: multiple -->
| EV/EBITDA |
|-----------|
| 12.5      |
"""
        data = _intercept_upload(markdown, recalc=False)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].value == 12.5
        assert ws["A2"].number_format == '0.0"x"'


class TestExternalReferenceColor:
    """B5: external-workbook references get red font in financial mode."""

    def _style(self, formula):
        from xlsx_tools.helpers import apply_financial_styling
        from openpyxl import Workbook
        wb = Workbook()
        cell = wb.active["A1"]
        cell.value = formula
        apply_financial_styling(cell, formula)
        return cell.font.color.rgb

    def test_external_reference_unquoted_gets_red(self):
        from xlsx_tools.helpers import FINANCIAL_EXTERNAL_COLOR
        rgb = self._style("=[Forecast.xlsx]Sheet1!A1")
        assert rgb.endswith(FINANCIAL_EXTERNAL_COLOR)

    def test_external_reference_quoted_gets_red(self):
        from xlsx_tools.helpers import FINANCIAL_EXTERNAL_COLOR
        rgb = self._style("='[Forecast.xlsx]Sheet1'!A1")
        assert rgb.endswith(FINANCIAL_EXTERNAL_COLOR)

    def test_local_formula_still_black(self):
        from xlsx_tools.helpers import FINANCIAL_FORMULA_COLOR
        rgb = self._style("=SUM(B2:B5)")
        assert rgb.endswith(FINANCIAL_FORMULA_COLOR)

    def test_cross_sheet_still_green(self):
        from xlsx_tools.helpers import FINANCIAL_CROSS_SHEET_COLOR
        rgb = self._style("=Inputs!B2")
        assert rgb.endswith(FINANCIAL_CROSS_SHEET_COLOR)

    def test_literal_still_blue(self):
        from xlsx_tools.helpers import FINANCIAL_INPUT_COLOR
        rgb = self._style("1000")
        assert rgb.endswith(FINANCIAL_INPUT_COLOR)
