"""Tests for the gap-fix rounds: number-format defaults, currency integer
variant, thousands-separator parsing, unconditional circular detection,
recalc skip-reason surfacing, unresolved-reference warnings, sheet-name
collision warnings (Round 7), and three reference-pipeline bug fixes
(Round 8).

Each test class maps 1:1 to a gap identified in the review pass.
"""

from __future__ import annotations

import io
import logging
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from xlsx_tools.base_xlsx_tool import markdown_to_excel  # noqa: E402
from xlsx_tools.helpers import (  # noqa: E402
    DEFAULT_NUMBER_FORMAT,
    DEFAULT_NUMBER_FORMAT_DECIMALS,
    FINANCIAL_DEFAULT_NUMBER,
    FINANCIAL_DEFAULT_NUMBER_DECIMALS,
    FINANCIAL_DEFAULT_PERCENT,
    _currency_base_format,
    _strip_thousands_separators,
    _default_number_format_for,
)

OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


def _intercept_upload(markdown: str, **kwargs) -> bytes:
    """Run markdown_to_excel with upload patched out; return raw xlsx bytes."""
    captured = {}

    def fake_upload(file_obj, suffix, **kw):
        captured["data"] = file_obj.read()
        return "fake://test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown, **kwargs)
    return captured["data"]


# ── Gap 1+2: number-format defaults (dash/parens in financial mode,
#    and large numbers preserve decimals) ──────────────────────────────────────


class TestNumberFormatDefaults:
    """Gap 2: plain numeric cells preserve decimals instead of force-rounding."""

    def test_large_whole_number_uses_integer_format(self):
        data = _intercept_upload(
            "| Revenue |\n|---|\n| 1000000 |\n", recalc=False
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1000000
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT  # "#,##0"

    def test_large_non_whole_number_preserves_decimals(self):
        """Was a bug: 1500.75 was force-rounded to display 1,501."""
        data = _intercept_upload(
            "| Price |\n|---|\n| 1500.75 |\n", recalc=False
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1500.75
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT_DECIMALS  # "#,##0.00"

    def test_small_whole_number_no_force_format(self):
        """Small whole numbers still get the integer default (unchanged)."""
        data = _intercept_upload("| Count |\n|---|\n| 5 |\n", recalc=False)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 5
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT


class TestFinancialDefaultFormats:
    """Gap 1: financial_modeling applies dash/parens as the default format."""

    def test_financial_mode_whole_number_uses_dash_variant(self):
        # Use a 5-digit value so it isn't treated as a 4-digit year label.
        data = _intercept_upload(
            "| Revenue |\n|---|\n| 100000 |\n",
            financial_modeling=True, recalc=False,
        )
        ws = load_workbook(io.BytesIO(data)).active
        # zeros render as "-", negatives in parens
        assert ws["A2"].number_format == FINANCIAL_DEFAULT_NUMBER

    def test_financial_mode_non_whole_uses_dash_variant_with_decimals(self):
        data = _intercept_upload(
            "| Price |\n|---|\n| 1500.75 |\n",
            financial_modeling=True, recalc=False,
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == FINANCIAL_DEFAULT_NUMBER_DECIMALS

    def test_financial_mode_percent_uses_dash_variant(self):
        data = _intercept_upload(
            "| Growth |\n|---|\n| 10% |\n",
            financial_modeling=True, recalc=False,
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == FINANCIAL_DEFAULT_PERCENT

    def test_financial_mode_formula_gets_dash_format(self):
        """Formula cells in financial mode get the dash default so a result
        of 0 renders as '-' and a negative as '(...)'."""
        data = _intercept_upload(
            "| A | Diff |\n|---|---|\n| 100 | =A2-A2 |\n",
            financial_modeling=True, recalc=False,
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["B2"].number_format == FINANCIAL_DEFAULT_NUMBER

    def test_non_financial_mode_unchanged(self):
        """Without financial_modeling, plain percent stays at 0.0%."""
        data = _intercept_upload(
            "| Growth |\n|---|\n| 10% |\n", recalc=False
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].number_format == "0.0%"

    def test_default_number_format_for_helper(self):
        assert _default_number_format_for(1000.0, False) == DEFAULT_NUMBER_FORMAT
        assert _default_number_format_for(1000.5, False) == DEFAULT_NUMBER_FORMAT_DECIMALS
        assert _default_number_format_for(1000.0, True) == FINANCIAL_DEFAULT_NUMBER
        assert _default_number_format_for(1000.5, True) == FINANCIAL_DEFAULT_NUMBER_DECIMALS


# ── Gap 5: zero-decimal (integer) currency variant ───────────────────────────


class TestCurrencyIntegerVariant:
    def test_currency_integer_variant_picks_zero_decimal(self):
        assert _currency_base_format("$", "integer") == "$#,##0"
        assert _currency_base_format("$", "int") == "$#,##0"
        assert _currency_base_format("$", "whole") == "$#,##0"

    def test_currency_default_stays_two_decimals(self):
        assert _currency_base_format("$", None) == "$#,##0.00"
        assert _currency_base_format("$", "dash") == "$#,##0.00"

    def test_currency_integer_with_dash_variant(self):
        """integer + dash combines: zero-decimal base, dash applied."""
        from xlsx_tools.helpers import _apply_format_variant
        base = _currency_base_format("$", "integer:dash")
        assert base == "$#,##0"
        assert _apply_format_variant(base, "integer:dash") == "$#,##0"

    def test_currency_integer_euro(self):
        assert _currency_base_format("€", "integer") == "#,##0 €"

    def test_currency_integer_unknown_symbol_fallback(self):
        assert _currency_base_format("₿", "integer") == '#,##0 "₿"'

    def test_currency_integer_end_to_end(self):
        markdown = """<!-- types: currency:$:integer -->
| Revenue ($mm) |
|---|
| $1500 |
"""
        data = _intercept_upload(markdown, recalc=False)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1500.0
        assert ws["A2"].number_format == "$#,##0"


# ── Gap 6: thousands separators in default numeric path ─────────────────────


class TestThousandsSeparatorParsing:
    def test_strip_thousands_english(self):
        assert _strip_thousands_separators("1,234.56") == "1234.56"

    def test_strip_thousands_european(self):
        assert _strip_thousands_separators("1.234,56") == "1234.56"

    def test_strip_thousands_bare(self):
        assert _strip_thousands_separators("1,234") == "1234"
        assert _strip_thousands_separators("1,5") == "1.5"  # decimal comma

    def test_strip_thousands_no_separators(self):
        assert _strip_thousands_separators("1234") == "1234"

    def test_plain_cell_with_commas_parses_as_number(self):
        """Was a bug: '1,234' stayed a string because float() raised."""
        data = _intercept_upload(
            "| Population |\n|---|\n| 1,234 |\n", recalc=False
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1234.0
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT

    def test_plain_cell_with_commas_and_decimals(self):
        data = _intercept_upload(
            "| GDP |\n|---|\n| 1,234.56 |\n", recalc=False
        )
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == 1234.56
        assert ws["A2"].number_format == DEFAULT_NUMBER_FORMAT_DECIMALS


# ── Gap 3: circular detection runs even when recalc=False ────────────────────


class TestCircularDetectionWithoutRecalc:
    def test_circular_ref_detected_when_recalc_false(self):
        """The tool description promises cycle detection runs without the
        recalc engine. Verify the standalone detection path executes when
        recalc=False by checking the warning log."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ.xlsx"
            with self._capture_logs("xlsx_tools.base_xlsx_tool") as logs:
                # recalc=False explicitly → cycle logged but NOT fatal.
                result = markdown_to_excel(markdown, recalc=False)
        assert result == "fake://circ.xlsx"
        joined = "\n".join(logs)
        assert "circular" in joined.lower() or "#CIRC" in joined

    def test_circular_ref_not_fatal_when_recalc_false(self):
        """User opted out of recalc → don't surprise-fail on cycles."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ.xlsx"
            # Must NOT raise.
            result = markdown_to_excel(markdown, recalc=False)
        assert result == "fake://circ.xlsx"

    def test_circular_ref_fatal_when_recalc_true(self):
        """Explicit recalc=True still fails on cycles (zero-errors policy)."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown, recalc=True)
        assert "#CIRC" in str(exc_info.value) or "circular" in str(exc_info.value).lower()

    @staticmethod
    def _capture_logs(logger_name: str):
        """Context manager capturing log records from a named logger."""
        from contextlib import contextmanager

        records: list[str] = []

        @contextmanager
        def _cm():
            lg = logging.getLogger(logger_name)
            handler = logging.Handler()
            handler.emit = lambda record: records.append(
                f"{record.levelname}: {record.getMessage()}"
            )
            lg.addHandler(handler)
            try:
                yield records
            finally:
                lg.removeHandler(handler)

        return _cm()


# ── Gap 7: recalc skip reason surfaced ───────────────────────────────────────


class TestRecalcSkipReasonSurfaced:
    def test_skip_reason_surfaced_when_recalc_explicit(self):
        """When recalc=True is explicit but the engine can't complete, the
        skip reason should surface in the raised error (so the model can
        rewrite the offending formula)."""
        from xlsx_tools.formula_engine import RecalcResult

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| 1 | =A2*2 |\n"
        )
        # Force the engine to report a skip (e.g. unsupported function).
        fake_result = RecalcResult(
            recalc_performed=False,
            skip_reason="recalculation engine error: NotImplementedError: XLOOKUP",
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch(
                 "xlsx_tools.formula_engine.recalculate_workbook",
                 return_value=fake_result,
             ):
            mock_upload.return_value = "fake://skip.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown, recalc=True)
        msg = str(exc_info.value)
        assert "XLOOKUP" in msg or "could not be completed" in msg.lower()

    def test_skip_reason_not_fatal_when_recalc_default(self):
        """Defaulted recalc (None) → skip reason logged but file delivered."""
        from xlsx_tools.formula_engine import RecalcResult

        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| 1 | =A2*2 |\n"
        )
        fake_result = RecalcResult(
            recalc_performed=False,
            skip_reason="recalculation engine error: oops",
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload, \
             patch(
                 "xlsx_tools.formula_engine.recalculate_workbook",
                 return_value=fake_result,
             ):
            mock_upload.return_value = "fake://skip-delivered.xlsx"
            # recalc=None (default) → must not raise.
            result = markdown_to_excel(markdown)
        assert result == "fake://skip-delivered.xlsx"


# ── Gap 4: unresolved reference warnings ─────────────────────────────────────


class TestUnresolvedReferenceWarnings:
    def test_unknown_table_ref_warns(self, caplog):
        """A T99 reference (when only 1 table exists) should warn."""
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            adjust_formula_references("=T99.B[0]", current_excel_row=5)
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "T99" in joined

    def test_unknown_cross_sheet_warns(self, caplog):
        """A cross-sheet ref to a sheet that doesn't exist should warn."""
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            adjust_formula_references(
                "=Nope!T1.B[0]",
                current_excel_row=5,
                all_sheet_table_positions={"Real": {"T1": 1}},
            )
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "Nope" in joined
        assert "Real" in joined  # known sheets listed

    def test_known_table_ref_does_not_warn(self, caplog):
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            result = adjust_formula_references(
                "=T1.B[0]",
                current_excel_row=5,
                table_positions={"T1": 1},
            )
        warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert not warnings
        assert result == "=B2"

    def test_known_cross_sheet_does_not_warn(self, caplog):
        from xlsx_tools.helpers import adjust_formula_references

        with caplog.at_level(logging.WARNING, logger="xlsx_tools.helpers"):
            result = adjust_formula_references(
                "=Real!T1.B[0]",
                current_excel_row=5,
                all_sheet_table_positions={"Real": {"T1": 1}},
            )
        warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
        assert not warnings


# ── Gap 8: sheet-name collision detection ────────────────────────────────────


class TestSheetNameCollisionWarning:
    def test_duplicate_sheet_name_warns(self, caplog):
        """Two identical sheet names should warn (openpyxl auto-renames)."""
        markdown = (
            "## Sheet: Model\n\n"
            "| A |\n|---|\n| 1 |\n\n"
            "## Sheet: Model\n\n"
            "| B |\n|---|\n| 2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://collision.xlsx"
            with caplog.at_level(logging.WARNING, logger="xlsx_tools.base_xlsx_tool"):
                markdown_to_excel(markdown, recalc=False)
        joined = "\n".join(r.getMessage() for r in caplog.records)
        assert "collide" in joined.lower() or "Model" in joined

    def test_distinct_sheet_names_do_not_warn(self, caplog):
        markdown = (
            "## Sheet: Alpha\n\n"
            "| A |\n|---|\n| 1 |\n\n"
            "## Sheet: Beta\n\n"
            "| B |\n|---|\n| 2 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://distinct.xlsx"
            with caplog.at_level(logging.WARNING, logger="xlsx_tools.base_xlsx_tool"):
                markdown_to_excel(markdown, recalc=False)
        collisions = [
            r for r in caplog.records
            if "collide" in r.getMessage().lower()
        ]
        assert not collisions


# ════════════════════════════════════════════════════════════════════════════
# ROUND 8 — three real bugs reproduced and fixed.
# ════════════════════════════════════════════════════════════════════════════


class TestCrossSheetFunctionRangePrefix:
    """Bug 1: =SheetName!T1.SUM(B[0]:B[2]) must emit =SUM(Sheet!B2:B4),
    not the invalid =SUM(Sheet!B2:Sheet!B4) (which yields #VALUE!)."""

    def test_cross_sheet_sum_single_prefix(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.SUM(B[0]:B[2])", 2, {"T1": 1}, all_pos
        )
        assert result == "=SUM(Data!B2:B4)"

    def test_cross_sheet_average_single_prefix(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.AVERAGE(B[0]:B[2])", 2, {"T1": 1}, all_pos
        )
        assert result == "=AVERAGE(Data!B2:B4)"

    def test_cross_sheet_sum_end_to_end_recalcs(self):
        """End-to-end: the corrected formula must recalc without #VALUE!."""
        markdown = (
            "## Sheet: Data\n\n"
            "| V |\n|---|\n| 10 |\n| 20 |\n| 30 |\n\n"
            "## Sheet: Summary\n\n"
            "| Total |\n|---|\n| =Data!T1.SUM(A[0]:A[2]) |\n"
        )
        data = _intercept_upload(markdown, recalc=True)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        assert wb["Summary"]["A2"].value == 60

    def test_cross_sheet_range_still_correct(self):
        """The sibling cs_range pattern was already correct — guard against
        a regression when fixing the function pattern."""
        from xlsx_tools.helpers import adjust_formula_references
        all_pos = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=Data!T1.B[0]:T1.B[2]", 2, {"T1": 1}, all_pos
        )
        assert result == "=Data!B2:B4"


class TestCurrentRowRelativeRefs:
    """Bug 2: B[0]/A[0] is documented as CURRENT-row-relative but resolved
    table-relative, silently corrupting every row past the first data row."""

    def test_unit_current_row_offset_zero(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for cur in [2, 3, 4, 5]:
            result = adjust_formula_references("=B[0]", cur, table_positions=tp)
            assert result == f"=B{cur}", f"row {cur}: {result}"

    def test_unit_current_row_negative_offset(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        # B[-1] on row N → B(N-1); the most common running-total pattern.
        for cur in [2, 3, 4, 5]:
            result = adjust_formula_references("=B[-1]", cur, table_positions=tp)
            assert result == f"=B{cur - 1}", f"row {cur}: {result}"

    def test_unit_current_row_positive_offset(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for cur in [2, 3, 4]:
            result = adjust_formula_references("=B[1]", cur, table_positions=tp)
            assert result == f"=B{cur + 1}", f"row {cur}: {result}"

    def test_unit_range_pattern_uses_current_row(self):
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        # =SUM(B[0]:E[0]) at row 4 → =SUM(B4:E4)
        result = adjust_formula_references("=SUM(B[0]:E[0])", 4, table_positions=tp)
        assert result == "=SUM(B4:E4)"
        # Range spanning rows: =SUM(B[-1]:B[1]) at row 3 → =SUM(B2:B4)
        result = adjust_formula_references("=SUM(B[-1]:B[1])", 3, table_positions=tp)
        assert result == "=SUM(B2:B4)"

    def test_table_relative_T1_still_correct(self):
        """The T1.B[n] form is table-relative (offset from first data row)
        and must be unaffected by the current-row fix."""
        from xlsx_tools.helpers import adjust_formula_references
        tp = {"T1": 1}
        for n, expected_row in [(0, 2), (1, 3), (2, 4)]:
            result = adjust_formula_references(f"=T1.B[{n}]", 99, table_positions=tp)
            assert result == f"=B{expected_row}"

    def test_end_to_end_running_total(self):
        """The classic running-total pattern must produce correct cached values."""
        markdown = (
            "| Delta | RunningTotal |\n"
            "|---|---|\n"
            "| 10 | =A[0] |\n"
            "| 20 | =B[-1]+A[0] |\n"
            "| 30 | =B[-1]+A[0] |\n"
        )
        data = _intercept_upload(markdown, recalc=True)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 10
        assert ws["B3"].value == 30
        assert ws["B4"].value == 60

    def test_end_to_end_per_row_multiply(self):
        """=A[0]*2 on each row must hit each row's A, not always row 2."""
        markdown = (
            "| Base | TwiceBase |\n"
            "|---|---|\n"
            "| 5 | =A[0]*2 |\n"
            "| 10 | =A[0]*2 |\n"
            "| 15 | =A[0]*2 |\n"
        )
        data = _intercept_upload(markdown, recalc=True)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 10
        assert ws["B3"].value == 20
        assert ws["B4"].value == 30


class TestYearAsStringRestriction:
    """Bug 3: financial mode forced ALL 4-digit numbers to text, not just
    plausible years. Revenue of 1500 became text '1500'."""

    def test_year_string_in_range(self):
        from xlsx_tools.helpers import _is_year_string
        for y in ["2024", "2025", "2099", "2100", "1900", "2050"]:
            assert _is_year_string(y), f"{y} should be a year"

    def test_year_string_out_of_range(self):
        from xlsx_tools.helpers import _is_year_string
        for n in ["1500", "1899", "2101", "5000", "9999", "1234"]:
            assert not _is_year_string(n), f"{n} should NOT be a year"

    def test_non_four_digit_not_year(self):
        from xlsx_tools.helpers import _is_year_string
        for v in ["2050E", "2024A", "abc", "12345", "123"]:
            assert not _is_year_string(v)

    def test_revenue_4_digit_stays_numeric_in_financial_mode(self):
        """A 4-digit revenue value must remain a number, not become text."""
        markdown = "| Revenue |\n|---|\n| 1500 |\n| 2500 |\n"
        data = _intercept_upload(markdown, financial_modeling=True, recalc=False)
        ws = load_workbook(io.BytesIO(data)).active
        assert isinstance(ws["A2"].value, (int, float))
        assert ws["A2"].value == 1500
        assert ws["A2"].number_format == "#,##0;(#,##0);-"
        assert isinstance(ws["A3"].value, (int, float))

    def test_year_column_stays_text_in_financial_mode(self):
        """Genuine years (2024, 2025) must still be text labels."""
        markdown = "| Year |\n|---|\n| 2024 |\n| 2025 |\n"
        data = _intercept_upload(markdown, financial_modeling=True, recalc=False)
        ws = load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == "2024"
        assert isinstance(ws["A2"].value, str)
        assert ws["A3"].value == "2025"

    def test_sum_over_4_digit_revenue_correct_in_financial_mode(self):
        """End-to-end: SUM over 4-digit revenue cells must total correctly,
        even when the file is later opened in real Excel (which does NOT
        coerce text-numerics in SUM the way the recalc engine does)."""
        markdown = (
            "| Revenue |\n|---|\n| 1500 |\n| 2500 |\n| 3500 |\n| =SUM(A2:A4) |\n"
        )
        data = _intercept_upload(markdown, financial_modeling=True, recalc=True)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # All inputs must be numeric (not text) so Excel-native SUM works.
        assert isinstance(ws["A2"].value, (int, float))
        assert isinstance(ws["A3"].value, (int, float))
        assert isinstance(ws["A4"].value, (int, float))
        assert ws["A5"].value == 7500
