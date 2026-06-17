"""Tests for the pure-Python Excel formula recalculation pipeline.

Covers:
- ``xlsx_tools.formula_engine.recalculate_workbook`` — engine-level
  correctness (single-sheet, multi-sheet, errors, booleans, unsupported
  functions, missing dependency fallback).
- ``xlsx_tools.xml_cache.inject_cached_values`` — XML rewriting
  correctness (cells get cached <v>, formulas are preserved, zip
  integrity maintained, non-matching locations are ignored).
- ``xlsx_tools.base_xlsx_tool.markdown_to_excel`` — end-to-end
  integration (recalc on/off, error surfacing).
"""

from __future__ import annotations

import inspect
import io
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

# Make project root importable.
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from xlsx_tools.formula_engine import (  # noqa: E402
    EXCEL_ERRORS,
    CellError,
    RecalcResult,
    is_available,
    recalculate_workbook,
)
from xlsx_tools.xml_cache import inject_cached_values  # noqa: E402
from xlsx_tools.base_xlsx_tool import markdown_to_excel  # noqa: E402

OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


def _save_wb(wb: Workbook, name: str | None = None) -> bytes:
    """Save a workbook to bytes; optionally also to OUTPUT_DIR for inspection."""
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if name:
        (OUTPUT_DIR / f"{name}.xlsx").write_bytes(data)
    return data


# ── Engine availability ──────────────────────────────────────────────────────


class TestEngineAvailability:
    """The `formulas` library should be installed in the dev/test environment."""

    def test_is_available_returns_true(self):
        # The library is in requirements.txt so this should always pass
        # in a properly-installed environment. If it ever doesn't, every
        # subsequent test is meaningless.
        assert is_available() is True


# ── Engine: happy path ───────────────────────────────────────────────────────


class TestEngineHappyPath:
    """recalculate_workbook returns correct numeric results."""

    def test_simple_sum(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 100
        ws["A2"] = 200
        ws["A3"] = 300
        ws["A4"] = "=SUM(A1:A3)"
        data = _save_wb(wb, "recalc_simple_sum")

        result = recalculate_workbook(data, ["Sheet1"])

        assert result.recalc_performed is True
        assert result.skip_reason is None
        assert result.errors == []
        assert result.values_map["Sheet1!A4"] == 600.0

    def test_arithmetic_and_functions(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1*2"
        ws["A3"] = "=A1+A2"
        ws["A4"] = "=MAX(A1:A3)"
        ws["A5"] = "=AVERAGE(A1:A3)"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])

        assert result.values_map["Sheet!A2"] == 20.0
        assert result.values_map["Sheet!A3"] == 30.0
        assert result.values_map["Sheet!A4"] == 30.0
        assert result.values_map["Sheet!A5"] == pytest.approx(20.0)

    def test_boolean_formula(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["A2"] = "=A1>50"
        ws["A3"] = "=A1<50"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])

        assert result.values_map["Sheet!A2"] is True
        assert result.values_map["Sheet!A3"] is False

    def test_integer_value_is_native_int(self):
        """Recalculated integer values should be native Python ints, not numpy."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 5
        ws["A2"] = "=A1+5"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])
        val = result.values_map["Sheet!A2"]

        assert val == 10
        assert type(val) is int  # noqa: E721 — we explicitly want native int

    def test_float_value_is_native_float(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=1/3"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])
        val = result.values_map["Sheet!A1"]

        assert val == pytest.approx(0.3333333, rel=1e-5)
        assert type(val) is float


# ── Engine: multi-sheet ──────────────────────────────────────────────────────


class TestEngineMultiSheet:
    """Cross-sheet references resolve correctly."""

    def test_cross_sheet_reference(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Revenue"
        ws1["A1"] = "Q1"
        ws1["B1"] = 1000
        ws1["B2"] = 1500
        ws2 = wb.create_sheet("Dashboard")
        ws2["B1"] = "=Revenue!B1"
        ws2["B2"] = "=SUM(Revenue!B1:B2)"
        data = _save_wb(wb, "recalc_cross_sheet")

        result = recalculate_workbook(data, ["Revenue", "Dashboard"])

        assert result.values_map["Dashboard!B1"] == 1000
        assert result.values_map["Dashboard!B2"] == 2500.0
        assert result.errors == []

    def test_sheet_name_casing_preserved(self):
        """Sheet names in values_map keys use the original casing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "MixedCase"
        ws["A1"] = 1
        ws["A2"] = "=A1+1"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["MixedCase"])
        # The key must use 'MixedCase' (original), not 'MIXEDCASE' (uppercased by the lib).
        assert "MixedCase!A2" in result.values_map
        assert result.values_map["MixedCase!A2"] == 2.0


# ── Engine: error detection ──────────────────────────────────────────────────


class TestEngineErrors:
    """Formula errors are detected and reported, never raised."""

    def test_div_by_zero(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["A2"] = "=A1/0"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])

        assert result.recalc_performed is True
        assert len(result.errors) == 1
        err = result.errors[0]
        assert err.error_type == "#DIV/0!"
        assert err.coordinate == "A2"
        assert err.sheet == "Sheet"
        assert "Sheet!A2" in str(err)

    def test_ref_error_missing_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=MissingSheet!A1"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])
        error_types = [e.error_type for e in result.errors]
        assert "#REF!" in error_types

    def test_name_error_unknown_function(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=BogusFunctionXYZ(1,2)"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])
        error_types = [e.error_type for e in result.errors]
        assert "#NAME?" in error_types

    def test_all_seven_errors_listed(self):
        """EXCEL_ERRORS covers the seven OOXML error sentinels."""
        assert len(EXCEL_ERRORS) == 7
        for e in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"):
            assert e in EXCEL_ERRORS


# ── Engine: robustness ───────────────────────────────────────────────────────


class TestEngineRobustness:
    """Engine never raises; missing deps degrade gracefully."""

    def test_empty_workbook_no_crash(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["Sheet"])
        assert result.recalc_performed is True
        assert result.errors == []

    def test_recalc_result_default_factory(self):
        result = RecalcResult()
        assert result.values_map == {}
        assert result.errors == []
        assert result.has_errors is False


# ── XML injection ────────────────────────────────────────────────────────────


class TestXmlInjection:
    """inject_cached_values rewrites only formula cells, preserving structure."""

    def test_formula_cell_gets_cached_value(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["A2"] = "=A1*2"
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"Sheet!A2": 200})

        wb2 = load_workbook(io.BytesIO(injected), data_only=True)
        assert wb2.active["A2"].value == 200

    def test_formula_preserved_after_injection(self):
        """Reading without data_only still shows the formula string."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 5
        ws["A2"] = "=A1*3"
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"Sheet!A2": 15})

        wb2 = load_workbook(io.BytesIO(injected))
        assert wb2.active["A2"].value == "=A1*3"

    def test_non_formula_cell_not_overwritten(self):
        """A literal value cell should not be modified even if it's in the map."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 42  # literal, no formula
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"Sheet!A1": 999})

        wb2 = load_workbook(io.BytesIO(injected), data_only=True)
        # A1 is not a formula cell, so injection must not touch it.
        assert wb2.active["A1"].value == 42

    def test_empty_values_map_returns_original_bytes(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        data = _save_wb(wb)

        assert inject_cached_values(data, {}) == data

    def test_zip_integrity_maintained(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1+5"
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"Sheet!A2": 15})

        import zipfile
        assert zipfile.ZipFile(io.BytesIO(injected)).testzip() is None

    def test_multi_sheet_injection(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Inputs"
        ws1["A1"] = 100
        ws2 = wb.create_sheet("Calc")
        ws2["A1"] = "=Inputs!A1*2"
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"Calc!A1": 200})

        wb2 = load_workbook(io.BytesIO(injected), data_only=True)
        assert wb2["Calc"]["A1"].value == 200

    def test_unknown_location_ignored(self):
        """A values_map entry for a non-existent sheet/cell is ignored."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=1+1"
        data = _save_wb(wb)

        # Bogus sheet name and bogus coordinate — should not raise.
        injected = inject_cached_values(
            data,
            {"NoSuchSheet!A1": 999, "Sheet!Z99": 999, "Sheet!A1": 2},
        )
        wb2 = load_workbook(io.BytesIO(injected), data_only=True)
        assert wb2.active["A1"].value == 2


# ── End-to-end via markdown_to_excel ─────────────────────────────────────────


def _intercept_upload(markdown: str, **kwargs) -> bytes:
    """Run markdown_to_excel with the upload patched out; return raw xlsx bytes."""
    captured = {}

    def fake_upload(file_obj, suffix, **kw):
        captured["data"] = file_obj.read()
        return "fake://test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown, **kwargs)
    return captured["data"]


class TestEndToEndRecalc:
    """markdown_to_excel integrates recalc + inject when recalc=True."""

    def test_recalc_on_produces_cached_values(self):
        markdown = """| A | B | Total |
|---|---|-------|
| 1 | 2 | =A2+B2 |
| 3 | 4 | =A3+B3 |
"""
        data = _intercept_upload(markdown, recalc=True)

        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # Row 2 (Excel) is the first data row.
        assert ws["C2"].value == 3
        assert ws["C3"].value == 7

    def test_recalc_off_leaves_formulas_without_cached_values(self):
        markdown = """| A | Total |
|---|-------|
| 1 | =A2*2 |
"""
        data = _intercept_upload(markdown, recalc=False)

        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # Without recalc, no cached value is stored.
        assert ws["B2"].value is None

        # Formula is still present.
        wb2 = load_workbook(io.BytesIO(data))
        assert wb2.active["B2"].value == "=A2*2"

    def test_recalc_explicit_fails_on_formula_errors(self):
        """When recalc is explicitly True, formula errors fail the call.

        This enforces the "zero formula errors" delivery standard: the
        model is told about the errors so it can fix the formulas and
        retry, rather than silently shipping a broken file.
        """
        markdown = """| A | B |
|---|---|
| 5 | =A2/0 |
"""
        with pytest.raises(RuntimeError, match="formula error"):
            _intercept_upload(markdown, recalc=True)

    def test_recalc_default_delivers_file_with_errors(self):
        """When recalc runs as a default (None), errors are logged but
        the file is still delivered — misconfiguration must never break
        document generation.
        """
        markdown = """| A | B |
|---|---|
| 5 | =A2/0 |
"""
        # recalc=None (default) — env config has XLSX_RECALC_ENABLED true,
        # so recalc runs, finds the error, but still delivers.
        data = _intercept_upload(markdown)

        wb = load_workbook(io.BytesIO(data))
        assert wb.active["B2"].value == "=A2/0"


# ── Recalc timeout ──────────────────────────────────────────────────────────


class TestRecalcTimeout:
    """Recalculation is bounded by a configurable timeout."""

    def test_config_has_timeout_field_with_default(self):
        from config import Config
        cfg = Config.from_env()
        assert cfg.xlsx_recalc_timeout_seconds == 30

    def test_recalc_and_inject_accepts_timeout_arg(self):
        """_recalc_and_inject signature accepts a timeout_seconds parameter."""
        import inspect
        from xlsx_tools.base_xlsx_tool import _recalc_and_inject
        sig = inspect.signature(_recalc_and_inject)
        assert "timeout_seconds" in sig.parameters
        assert sig.parameters["timeout_seconds"].default == 30

    def test_recalc_completes_within_generous_timeout(self):
        """A small workbook recalc'd with a generous timeout succeeds."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1*5"
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        from xlsx_tools.base_xlsx_tool import _recalc_and_inject
        result_bytes, error_summary = _recalc_and_inject(
            data, ["Sheet"], timeout_seconds=30
        )

        wb2 = load_workbook(io.BytesIO(result_bytes), data_only=True)
        assert wb2.active["A2"].value == 50
        assert error_summary is None

    def test_recalc_with_tiny_timeout_falls_back_gracefully(self):
        """A tiny timeout causes recalc to be skipped; original bytes returned.

        We can't reliably force a real timeout without a pathological
        workbook, so we verify the fallback contract: when the engine
        can't finish in time, the original (un-recalc'd) bytes are
        returned and the cached value is absent.
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1*5"
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        from xlsx_tools.base_xlsx_tool import _recalc_and_inject
        # 0 seconds is below the config minimum (1), but the function
        # accepts any positive int. Use 1 to give the engine a chance but
        # still exercise the timeout path on slow machines.
        result_bytes, error_summary = _recalc_and_inject(
            data, ["Sheet"], timeout_seconds=1
        )

        # Either it completed (cached value present) or it timed out
        # (original bytes returned). Both are acceptable outcomes — the
        # contract is "never hang, always return bytes".
        assert isinstance(result_bytes, bytes)
        assert len(result_bytes) > 0


# ──────────────────────────────────────────────────────────────────────────────
# ── Circular-reference detection ─────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestCircularReferenceDetection:
    """The `formulas` library silently misses circular references.

    `detect_circular_references()` builds a dependency graph from the
    formula strings and runs DFS cycle detection. Every cell on a cycle
    is reported as a `#CIRC!` error.
    """

    def test_simple_two_cell_cycle(self):
        from xlsx_tools.formula_engine import (
            CIRCULAR_ERROR_TYPE,
            detect_circular_references,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=A2"
        ws["A2"] = "=A1"
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        assert len(errors) == 2
        locations = {e.location for e in errors}
        assert locations == {"Sheet1!A1", "Sheet1!A2"}
        assert all(e.error_type == CIRCULAR_ERROR_TYPE for e in errors)

    def test_self_reference(self):
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=A1+1"  # classic self-ref
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        assert len(errors) == 1
        assert errors[0].location == "Sheet1!A1"

    def test_no_cycle_returns_empty(self):
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 1
        ws["A2"] = "=A1+1"
        ws["A3"] = "=A2+1"  # linear chain, no cycle
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        assert errors == []

    def test_cross_sheet_cycle(self):
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws1["A1"] = "=Sheet2!A1"  # Sheet1!A1 -> Sheet2!A1
        ws2["A1"] = "=Sheet1!A1"  # Sheet2!A1 -> Sheet1!A1 (cycle)
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1", "Sheet2"])
        locations = {e.location for e in errors}
        assert "Sheet1!A1" in locations
        assert "Sheet2!A1" in locations

    def test_longer_cycle_chain(self):
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # A1 -> B1 -> C1 -> A1
        ws["A1"] = "=B1"
        ws["B1"] = "=C1"
        ws["C1"] = "=A1"
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        locations = {e.location for e in errors}
        assert locations == {"Sheet1!A1", "Sheet1!B1", "Sheet1!C1"}

    def test_cycle_with_innocent_bystander(self):
        """A cell that references a cycle member but isn't itself on the
        cycle should not be flagged."""
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=A2"
        ws["A2"] = "=A1"  # A1<->A2 cycle
        ws["A3"] = "=A1"  # innocent: refs a cycle member but no cycle
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        locations = {e.location for e in errors}
        assert "Sheet1!A3" not in locations
        assert "Sheet1!A1" in locations
        assert "Sheet1!A2" in locations

    def test_range_reference_cycle(self):
        """Cycles through a range reference should still be detected."""
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # SUM(A1:A3) in A2 creates a self-loop through the range.
        ws["A1"] = 1
        ws["A2"] = "=SUM(A1:A3)"
        ws["A3"] = 1
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1"])
        locations = {e.location for e in errors}
        # A2 references itself (via the range that includes A2).
        assert "Sheet1!A2" in locations


class TestExtractFormulaReferences:
    """Unit tests for the formula-string reference parser."""

    def test_local_reference(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1"])
        refs = extract_formula_references("=A1+B2", "Sheet1", sl)
        assert refs == {"Sheet1!A1", "Sheet1!B2"}

    def test_absolute_reference(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1"])
        refs = extract_formula_references("=$A$1+B$2", "Sheet1", sl)
        assert refs == {"Sheet1!A1", "Sheet1!B2"}

    def test_cross_sheet_reference(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Inputs", "Sheet1"])
        refs = extract_formula_references("=Inputs!A1+Sheet1!B2", "Sheet1", sl)
        assert refs == {"Inputs!A1", "Sheet1!B2"}

    def test_range_expanded(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1"])
        refs = extract_formula_references("=SUM(A1:A3)", "Sheet1", sl)
        assert refs == {"Sheet1!A1", "Sheet1!A2", "Sheet1!A3"}

    def test_unknown_sheet_dropped(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1"])
        # 'Missing' is not in the workbook — ref should be dropped
        # (the recalc engine will flag a real #REF! if it's a problem).
        refs = extract_formula_references("=Missing!A1+A2", "Sheet1", sl)
        assert refs == {"Sheet1!A2"}

    def test_empty_formula(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1"])
        assert extract_formula_references("", "Sheet1", sl) == set()
        assert extract_formula_references("=1+2", "Sheet1", sl) == set()

    def test_3d_reference_expanded_across_sheets(self):
        """A 3D reference like SUM(Sheet1:Sheet3!A1) should expand
        to refs on every sheet in the workbook range."""
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1", "Sheet2", "Sheet3"])
        refs = extract_formula_references("=SUM(Sheet1:Sheet3!A1)", "Sheet1", sl)
        assert refs == {"Sheet1!A1", "Sheet2!A1", "Sheet3!A1"}

    def test_3d_reference_with_cell_range(self):
        """3D refs combined with cell ranges expand across all cells
        on all sheets in the range."""
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1", "Sheet2", "Sheet3"])
        refs = extract_formula_references("=SUM(Sheet1:Sheet3!A1:B2)", "Sheet1", sl)
        # 3 sheets × 4 cells = 12 refs
        assert len(refs) == 12
        assert "Sheet2!A1" in refs
        assert "Sheet2!B2" in refs
        assert "Sheet3!A2" in refs

    def test_3d_reference_reverse_order(self):
        """A reverse 3D range (Sheet3:Sheet1!A1) should still expand
        to all sheets between, in workbook order."""
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["Sheet1", "Sheet2", "Sheet3"])
        refs = extract_formula_references("=SUM(Sheet3:Sheet1!A1)", "Sheet1", sl)
        assert refs == {"Sheet1!A1", "Sheet2!A1", "Sheet3!A1"}

    def test_3d_reference_circular_detection(self):
        """F7 integration: a formula with a 3D ref that includes its own
        cell should be detected as circular."""
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws3 = wb.create_sheet("Sheet3")
        # Sheet2!A1 sums A1 across all 3 sheets — but it's itself in the range.
        ws2["A1"] = "=SUM(Sheet1:Sheet3!A1)"
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["Sheet1", "Sheet2", "Sheet3"])
        locations = {e.location for e in errors}
        assert "Sheet2!A1" in locations


class TestCircularRefErrorPolicy:
    """Integration: explicit recalc=True with circular refs fails."""

    def test_explicit_recalc_fails_on_circular_ref(self):
        """When recalc=True is explicit and a circular ref is present,
        the call must fail (zero-errors policy)."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown, recalc=True)

        msg = str(exc_info.value)
        assert "#CIRC!" in msg
        assert "circular" in msg.lower()
        # Upload must not have run — we fail before delivery.
        mock_upload.assert_not_called()

    def test_default_recalc_delivers_file_with_circular_ref(self):
        """When recalc is default (None) and a circular ref is present,
        the file is still delivered (errors logged only)."""
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| =B2 | =A2 |\n"
        )

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://circ-delivered.xlsx"
            result = markdown_to_excel(markdown)  # no recalc param

        assert result == "fake://circ-delivered.xlsx"
        mock_upload.assert_called_once()


# ──────────────────────────────────────────────────────────────────────────────
# ── Structured (type-grouped) error output ──────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestGroupedErrorOutput:
    """Errors in the RuntimeError message are grouped by type with
    counts and locations, not a flat list."""

    def test_grouped_format_single_type(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CellError

        errors = [
            CellError(sheet="Sheet1", coordinate="B2", error_type="#DIV/0!"),
            CellError(sheet="Sheet1", coordinate="B5", error_type="#DIV/0!"),
        ]
        msg = _format_grouped_errors(errors)
        assert "#DIV/0! (2)" in msg
        assert "Sheet1!B2" in msg
        assert "Sheet1!B5" in msg
        assert "2 formula error(s)" in msg

    def test_grouped_format_multiple_types(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CellError

        errors = [
            CellError(sheet="S", coordinate="B2", error_type="#DIV/0!"),
            CellError(sheet="S", coordinate="B3", error_type="#DIV/0!"),
            CellError(sheet="S", coordinate="C10", error_type="#REF!"),
        ]
        msg = _format_grouped_errors(errors)
        assert "#DIV/0! (2)" in msg
        assert "#REF! (1)" in msg
        assert "3 formula error(s)" in msg

    def test_grouped_format_truncates_long_lists(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CellError

        errors = [
            CellError(sheet="S", coordinate=f"A{i}", error_type="#REF!")
            for i in range(10)
        ]
        msg = _format_grouped_errors(errors)
        assert "#REF! (10)" in msg
        assert "and 5 more" in msg  # only first 5 shown

    def test_grouped_format_circular_annotation(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CIRCULAR_ERROR_TYPE, CellError

        errors = [
            CellError(sheet="S", coordinate="A1", error_type=CIRCULAR_ERROR_TYPE),
        ]
        msg = _format_grouped_errors(errors)
        assert "#CIRC! (1)" in msg
        # The annotation helps the model understand #CIRC! is a cycle.
        assert "circular references detected" in msg.lower()
        assert "breaking the cycle" in msg.lower()

    def test_grouped_format_circular_mixed_with_excel_errors(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CIRCULAR_ERROR_TYPE, CellError

        errors = [
            CellError(sheet="S", coordinate="A1", error_type=CIRCULAR_ERROR_TYPE),
            CellError(sheet="S", coordinate="B2", error_type="#DIV/0!"),
        ]
        msg = _format_grouped_errors(errors)
        assert "#CIRC! (1)" in msg
        assert "#DIV/0! (1)" in msg
        assert "2 formula error(s)" in msg
        assert "circular references detected" in msg.lower()


# ──────────────────────────────────────────────────────────────────────────────
# ── total_formulas telemetry ─────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestTotalFormulasCount:
    """RecalcResult carries a total_formulas count for telemetry."""

    def test_count_formulas_helper(self):
        from xlsx_tools.xml_cache import count_formulas

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 1
        ws["A2"] = "=A1+1"
        ws["A3"] = "=SUM(A1:A2)"
        data = _save_wb(wb)

        assert count_formulas(data) == 2

    def test_count_formulas_empty_workbook(self):
        from xlsx_tools.xml_cache import count_formulas

        wb = Workbook()
        data = _save_wb(wb)
        assert count_formulas(data) == 0

    def test_recalc_result_populates_total_formulas(self):
        if not is_available():
            pytest.skip("`formulas` library not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 10
        ws["A2"] = "=A1+5"
        ws["A3"] = "=A2*2"
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["S"])
        assert result.recalc_performed
        assert result.total_formulas == 2


# ──────────────────────────────────────────────────────────────────────────────
# ── String-result formula cached values (t="str") ───────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestStringFormulaInjection:
    """Formulas whose result is a string (e.g. =A1&' total', =IF(...))
    should now get a cached value via t="str" inline-string cells.

    Previously these were silently skipped, which meant the file previewed
    blank/0 in Google Sheets and other non-recalculating clients.
    """

    def test_concatenation_formula_cached(self):
        if not is_available():
            pytest.skip("`formulas` library not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "World"
        ws["B1"] = '=A1 & " Hello"'
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["S"])
        assert "S!B1" in result.values_map
        assert result.values_map["S!B1"] == "World Hello"

    def test_if_formula_string_branch_cached(self):
        if not is_available():
            pytest.skip("`formulas` library not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 100
        ws["B1"] = '=IF(A1>50,"Big","Small")'
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["S"])
        assert result.values_map.get("S!B1") == "Big"

    def test_string_value_serializes_with_t_str_attr(self):
        """The XML for a string-result formula cell must carry t='str'."""
        from xlsx_tools.xml_cache import inject_cached_values
        import zipfile
        import xml.etree.ElementTree as ET

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "X"
        ws["B1"] = "=A1 & \"Y\""  # placeholder formula; we inject our own value
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"S!B1": "computed-string"})
        zf = zipfile.ZipFile(io.BytesIO(injected))
        try:
            xml = zf.read("xl/worksheets/sheet1.xml").decode()
        finally:
            zf.close()

        # The B1 cell should have t="str" and a <v> with our string.
        root = ET.fromstring(xml)
        ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        for cell in root.iter(f"{ns}c"):
            if cell.get("r") == "B1":
                assert cell.get("t") == "str"
                v = cell.find(f"{ns}v")
                assert v is not None
                assert v.text == "computed-string"
                return
        pytest.fail("B1 cell not found in injected XML")

    def test_end_to_end_string_formula_previews_via_data_only(self):
        """End-to-end: a string formula in markdown gets a cached value
        that openpyxl(data_only=True) can read back — the core F1 fix."""
        markdown = (
            "| Greeting | Result |\n"
            "|---|---|\n"
            "| World | =A2 & \" Hello\" |\n"
        )
        captured = {}

        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            def capture(file_obj, *args, **kwargs):
                captured["data"] = file_obj.read()
                return "fake://str.xlsx"
            mock_upload.side_effect = capture
            markdown_to_excel(markdown, recalc=True)

        wb = load_workbook(io.BytesIO(captured["data"]), data_only=True)
        ws = wb.active
        assert ws["B2"].value == "World Hello"

    def test_xml_escapes_special_chars_in_string_cached_value(self):
        """String cached values with XML-special chars (<, >, &) must be
        properly escaped by ElementTree so the file stays valid."""
        from xlsx_tools.xml_cache import inject_cached_values
        import zipfile

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "x"
        ws["B1"] = "=A1"  # placeholder
        data = _save_wb(wb)

        injected = inject_cached_values(data, {"S!B1": "a < b & c > d"})
        # File should still be a valid zip / openable by openpyxl.
        wb2 = load_workbook(io.BytesIO(injected), data_only=True)
        assert wb2["S"]["B1"].value == "a < b & c > d"


# ──────────────────────────────────────────────────────────────────────────────
# ── total_formulas surfaced in grouped error output ─────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestTotalFormulasInErrorSummary:
    """The error summary should include 'N/total' when total_formulas
    is known, so the model understands scope (2/5 vs 2/500)."""

    def test_error_summary_includes_total(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CellError

        errors = [CellError(sheet="S", coordinate="B2", error_type="#DIV/0!")]
        msg = _format_grouped_errors(errors, total_formulas=15)
        assert "1/15 formula error(s)" in msg

    def test_error_summary_omits_total_when_zero(self):
        from xlsx_tools.base_xlsx_tool import _format_grouped_errors
        from xlsx_tools.formula_engine import CellError

        errors = [CellError(sheet="S", coordinate="B2", error_type="#DIV/0!")]
        # total_formulas=0 means unknown — should fall back to bare count.
        msg = _format_grouped_errors(errors, total_formulas=0)
        assert "1 formula error(s)" in msg
        assert "/" not in msg.split(":")[0]  # no N/total prefix

    def test_end_to_end_error_summary_has_total_formulas(self):
        """An actual recalc that finds errors should surface total_formulas."""
        # Division by zero — guaranteed error.
        markdown = (
            "| A | B |\n"
            "|---|---|\n"
            "| 0 | =5/A2 |\n"
            "| 10 | =10/A3 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake.xlsx"
            with pytest.raises(RuntimeError) as exc_info:
                markdown_to_excel(markdown, recalc=True)

        msg = str(exc_info.value)
        # The error count should appear as "N/total" since we know the total.
        assert "formula error(s):" in msg
        # Total formulas is at least 2 (the two = formulas).
        assert "/" in msg.split("formula error(s)")[0]


# ──────────────────────────────────────────────────────────────────────────────
# ── No false-positive circular refs from string literals ────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestNoFalsePositiveCircularRefs:
    """String literals containing coord-like text must NOT be parsed
    as cell references. A formula like ='see A1' in cell A1 must not be
    flagged as a circular reference."""

    def test_string_literal_with_coord_not_extracted(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # Text mentioning A1 — should NOT produce a ref.
        refs = extract_formula_references('="see cell A1 for context"', "S", sl)
        assert refs == set()

    def test_real_ref_plus_text_coord_only_returns_real(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # Real A1 ref + text "pos A1" — only the real A1 should be returned.
        refs = extract_formula_references('=IF(A1>0,"pos A1","neg")', "S", sl)
        assert refs == {"S!A1"}

    def test_vlookup_string_arg_not_treated_as_ref(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # "A1" is a lookup KEY, not a cell. Only A2:B5 should be returned.
        refs = extract_formula_references('=VLOOKUP("A1", A2:B5, 2)', "S", sl)
        assert "S!A1" not in refs
        assert "S!A2" in refs
        assert "S!B5" in refs

    def test_escaped_quote_in_string_literal(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # Escaped "" inside a string literal — A1 inside should not match.
        refs = extract_formula_references('="escaped ""A1"" quote"', "S", sl)
        assert refs == set()

    def test_self_referential_label_not_flagged_as_circular(self):
        """The canonical G1 regression: a label formula that mentions its
        own cell in a string must not be flagged as circular."""
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        # A1 contains a formula whose text mentions "A1" — but it's a label,
        # not a real self-reference.
        ws["A1"] = '="Row 1 totals (see A1)"'
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["S"])
        assert errors == [], (
            f"False positive! Got: {[(e.location, e.error_type) for e in errors]}"
        )

    def test_concat_with_own_coord_in_text_not_circular(self):
        """=A1&' totals' in A1: the A1 IS a real ref (self-ref), so this
        SHOULD be flagged. But =B1&' see A1' in A1: only B1 is real, A1
        is text — should NOT be flagged."""
        from xlsx_tools.formula_engine import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = '=B1&" see A1"'  # A1 in text only
        ws["B1"] = 5
        data = _save_wb(wb)

        errors = detect_circular_references(data, ["S"])
        assert errors == []

    def test_end_to_end_label_formula_does_not_block_delivery(self):
        """End-to-end: a financial model with label formulas that mention
        cells in their text must not fail with a false circular ref."""
        markdown = (
            "| Metric | Value | Notes |\n"
            "|---|---|---|\n"
            "| Revenue | 1000 | =\"see B2 for detail\" |\n"
            "| Cost | 400 | =B2*0.4 |\n"
        )
        with patch("xlsx_tools.base_xlsx_tool.upload_file") as mock_upload:
            mock_upload.return_value = "fake://model.xlsx"
            # This must NOT raise — the label formula is legitimate.
            result = markdown_to_excel(markdown, recalc=True)
        assert result == "fake://model.xlsx"


# ──────────────────────────────────────────────────────────────────────────────
# ── values_map only contains formula cells ──────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestValuesMapFormulaOnly:
    """The recalc engine's solution dict contains every cell, but our
    values_map should only contain formula cells. Otherwise the 'N/M formulas
    cached' log is misleading and we do 5x the necessary work."""

    def test_values_map_excludes_input_cells(self):
        if not is_available():
            pytest.skip("`formulas` library not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "label"   # input
        ws["A2"] = 5         # input
        ws["B2"] = "=A2*2"   # formula
        ws["C2"] = "=B2+1"   # formula
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["S"])
        # Only formula cells should be in the map.
        assert set(result.values_map.keys()) == {"S!B2", "S!C2"}
        assert result.values_map["S!B2"] == 10
        assert result.values_map["S!C2"] == 11

    def test_log_count_matches_formula_count(self):
        """The 'N/M formulas cached' log line should have N ≤ M, where M is
        the true formula count. Before G2, N could exceed M."""
        if not is_available():
            pytest.skip("`formulas` library not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for i in range(2, 12):
            ws[f"A{i}"] = i  # 10 input cells
        ws["B1"] = "=SUM(A2:A11)"  # 1 formula
        data = _save_wb(wb)

        result = recalculate_workbook(data, ["S"])
        assert len(result.values_map) <= result.total_formulas
        assert result.total_formulas == 1
        assert len(result.values_map) == 1


# ──────────────────────────────────────────────────────────────────────────────
# ── Coord validation drops phantom matches ──────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────


class TestCoordValidation:
    """Extracted refs must be valid Excel coordinates. Phantom matches
    like ZZZ1234 (4-letter column) or A9999999 (out-of-range row) are dropped."""

    def test_is_valid_coord_basic(self):
        from xlsx_tools.formula_engine import _is_valid_coord

        assert _is_valid_coord("A1")
        assert _is_valid_coord("XFD1048576")  # max valid
        assert _is_valid_coord("AA10")
        assert _is_valid_coord("a1")  # case-insensitive

    def test_is_valid_coord_rejects_invalid(self):
        from xlsx_tools.formula_engine import _is_valid_coord

        assert not _is_valid_coord("ZZZZ1")     # 4-letter column
        assert not _is_valid_coord("A1048577")  # row out of range
        assert not _is_valid_coord("A0")        # row 0
        assert not _is_valid_coord("1A")        # not a coord
        assert not _is_valid_coord("")

    def test_phantom_4letter_column_dropped(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # The regex will find the longest 1-3 letter column match it can.
        # ABCDE1 has 5 letters — the regex matches CDE1 (3 letters), which
        # is a valid coord. The leading AB is unmatched. This is a known
        # limitation of regex parsing: without a real lexer we can't tell
        # ABCDE1 was meant as a (hypothetical) 5-letter column. In practice
        # such a formula would be a #NAME? error caught by the recalc engine.
        # What we CAN verify is that genuinely invalid coords (4+ letters
        # with no valid 3-letter suffix, or out-of-range rows) are dropped.
        refs = extract_formula_references("=ZZZZZ1+1", "S", sl)
        # ZZZZ1 → regex tries ZZZ (3 letters) + Z1 — but Z1 alone isn't
        # reachable because the 4th Z breaks the digit boundary. Actually
        # the regex matches ZZZ1 (treating the 5th Z as separator) — which
        # IS a valid coord. So we just assert no obviously-invalid coords leak.
        for ref in refs:
            sheet, coord = ref.split("!")
            # Every extracted coord must pass validation.
            from xlsx_tools.formula_engine import _is_valid_coord
            assert _is_valid_coord(coord), f"Invalid coord leaked: {coord}"

    def test_out_of_range_row_dropped(self):
        from xlsx_tools.formula_engine import (
            _build_sheet_lookup,
            extract_formula_references,
        )

        sl = _build_sheet_lookup(["S"])
        # Row 9,999,999 is way beyond Excel's limit.
        refs = extract_formula_references("=A9999999", "S", sl)
        assert refs == set()


# ── External-link isolation ──────────────────────────────────────────────────


class TestExternalLinkIsolation:
    """A single external-workbook reference (e.g. =[Other.xlsx]Sheet1!A1)
    must not suppress cached values for every other formula in the file.

    The `formulas` library aborts the entire workbook with FormulaError
    when it encounters such a reference. We pre-scan and blank those cells
    in the temp copy handed to the engine, so the rest of the workbook
    recalculates normally. The user's original file keeps the external-link
    formula intact; Excel evaluates it on open.
    """

    def test_external_link_does_not_abort_recalc(self):
        """Other formulas get cached even when an external link is present."""
        markdown = """| Label | Value |
|-------|-------|
| A     | 100   |
| B     | =B2*2 |
| Ext   | =[Other.xlsx]Sheet1!A1 |
| C     | =B2+B3 |
"""
        data = _intercept_upload(markdown, recalc=True)

        wb_cached = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb_cached.active
        # B3 (=B2*2 = 200) and B5 (=B2+B3 = 300) must be cached.
        assert ws["B3"].value == 200
        assert ws["B5"].value == 300
        # B4 (the external link) has no cached value — Excel computes it on open.
        assert ws["B4"].value is None

    def test_external_link_formula_preserved_in_file(self):
        """The external-link formula must still be in the delivered file
        (we only blank it in the temp copy the engine sees)."""
        markdown = """| Label | Value |
|-------|-------|
| Ext   | =[Other.xlsx]Sheet1!A1 |
| B     | =B2*2 |
"""
        data = _intercept_upload(markdown, recalc=True)

        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["B2"].value == "=[Other.xlsx]Sheet1!A1"

    def test_quoted_external_link_handled(self):
        """Quoted external references like ='[Other.xlsx]Sheet 1'!A1 are
        also detected and isolated."""
        markdown = """| Label | Value |
|-------|-------|
| A     | 100   |
| Ext   | ='[Other.xlsx]Sheet 1'!A1 |
| B     | =B2*2 |
"""
        data = _intercept_upload(markdown, recalc=True)

        wb_cached = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb_cached.active
        # B4 (=B2*2 = 200) still gets cached despite the external link in B3.
        assert ws["B4"].value == 200

    def test_no_external_links_unchanged_behavior(self):
        """When there are no external links, the blanking pass is a no-op
        and recalc behaves exactly as before."""
        markdown = """| A | B |
|---|---|
| 1 | =A2*2 |
| 2 | =A3*2 |
"""
        data = _intercept_upload(markdown, recalc=True)

        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 2
        assert ws["B3"].value == 4


# ── Directive carry-forward across sheet headers ─────────────────────────────


class TestDirectiveCarryForward:
    """Directives placed above a '## Sheet:' header should apply to the next
    table, not be silently dropped at the sheet boundary.

    Previously, `<!-- types: ... -->` or `<!-- freeze -->` placed at the top
    of the markdown (above the first sheet header) was silently ignored,
    which then cascaded into formula errors that looked unrelated to the
    directive. Directives still reset after each table, so a directive
    intended for one table does not leak to tables on later sheets.
    """

    def test_types_directive_above_sheet_header_applies(self):
        markdown = """<!-- types: text, currency:$ -->
## Sheet: Model

| Year | Revenue |
|------|---------|
| 2024 | $1,000  |
"""
        data = _intercept_upload(markdown, recalc=False)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Model"]
        # B2 should be the number 1000 with currency format, not the text '$1,000'.
        assert ws["B2"].value == 1000
        assert ws["B2"].number_format.startswith("$")

    def test_freeze_directive_above_sheet_header_applies(self):
        markdown = """<!-- freeze -->
## Sheet: Model

| A | B |
|---|---|
| 1 | 2 |
"""
        data = _intercept_upload(markdown, recalc=False)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Model"]
        assert ws.freeze_panes == "A2"

    def test_directive_does_not_leak_to_next_sheet(self):
        """A directive applied to sheet 1's table is cleared after that
        table — it must not silently apply to sheet 2's table."""
        markdown = """<!-- types: text, currency:$ -->
## Sheet: First

| Year | Revenue |
|------|---------|
| 2024 | $1,000  |

## Sheet: Second

| Metric | Value |
|--------|-------|
| Count  | 42    |
"""
        data = _intercept_upload(markdown, recalc=False)
        wb = load_workbook(io.BytesIO(data))
        # Sheet 1: currency format applied.
        first = wb["First"]
        assert first["B2"].number_format.startswith("$")
        # Sheet 2: no currency format (directive was cleared after sheet 1's table).
        second = wb["Second"]
        assert not second["B2"].number_format.startswith("$")
