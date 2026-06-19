"""Tests for general-purpose xlsx styling features.

Covers:
- ``apply_default_font`` — font family propagation, monospace preservation.
- ``parse_sources_directive`` — source-citation parsing including ranges.
- Number-format variants: ``number:dash``, ``currency:$:parens``,
  ``percent:dash``, ``multiple`` type, etc.
- Cell comments for source citations.
- End-to-end via ``markdown_to_excel`` (general styling, no financial flag).
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from xlsx_tools.helpers import (  # noqa: E402
    NUMBER_FORMAT_VARIANTS,
    PERCENT_FORMAT_VARIANTS,
    MULTIPLES_FORMAT_VARIANTS,
    _apply_format_variant,
    _apply_percent_format_variant,
    _apply_multiples_format_variant,
    _apply_column_type,
    _expand_coord_range,
    apply_default_font,
    attach_source_comment,
    parse_sources_directive,
)
from xlsx_tools.base_xlsx_tool import markdown_to_excel  # noqa: E402

OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


def _intercept_upload(markdown: str, **kwargs) -> bytes:
    captured = {}

    def fake_upload(file_obj, suffix, **kw):
        captured["data"] = file_obj.read()
        return "fake://test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown, **kwargs)
    return captured["data"]


# ── Default font ─────────────────────────────────────────────────────────────


class TestDefaultFont:
    """apply_default_font propagates family but preserves monospace."""

    def test_font_family_applied(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = "hello"
        apply_default_font(cell, "Arial")
        assert cell.font.name == "Arial"

    def test_monospace_font_preserved(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        from openpyxl.styles import Font
        cell.font = Font(name="Courier New")
        apply_default_font(cell, "Arial")
        assert cell.font.name == "Courier New"

    def test_none_font_no_change(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        original_name = cell.font.name
        apply_default_font(cell, None)
        assert cell.font.name == original_name


# ── Source-citation directive parser ─────────────────────────────────────────


class TestSourcesDirective:
    """parse_sources_directive handles per-cell and range forms."""

    def test_per_cell_entries(self):
        result = parse_sources_directive(
            "B2=Source: Annual Report 2024, B5=Source: Internal forecast"
        )
        assert result == {
            "B2": "Source: Annual Report 2024",
            "B5": "Source: Internal forecast",
        }

    def test_range_form_expands(self):
        result = parse_sources_directive("B2:B4=Same source applies")
        assert result == {
            "B2": "Same source applies",
            "B3": "Same source applies",
            "B4": "Same source applies",
        }

    def test_empty_value(self):
        assert parse_sources_directive("") == {}

    def test_malformed_entries_ignored(self):
        result = parse_sources_directive("B2=valid, garbage, =nokey, B5=")
        assert result == {"B2": "valid"}

    def test_range_expander_helper(self):
        assert _expand_coord_range("B2:B5") == ["B2", "B3", "B4", "B5"]
        assert _expand_coord_range("B5:B2") == ["B2", "B3", "B4", "B5"]

    def test_single_cell_overrides_range_when_listed_first(self):
        result = parse_sources_directive(
            "B2=Specific source, B2:B3=Range source"
        )
        assert result["B2"] == "Specific source"
        assert result["B3"] == "Range source"

    def test_single_cell_overrides_range_when_listed_last(self):
        result = parse_sources_directive(
            "B2:B3=Range source, B2=Specific source"
        )
        assert result["B2"] == "Specific source"
        assert result["B3"] == "Range source"

    def test_multiple_singles_override_range(self):
        result = parse_sources_directive(
            "B2:B4=Range, B2=First, B4=Third"
        )
        assert result["B2"] == "First"
        assert result["B3"] == "Range"
        assert result["B4"] == "Third"

    def test_attach_source_comment(self):
        wb = Workbook()
        cell = wb.active["A1"]
        attach_source_comment(cell, "Source: 10-K filing 2024")
        assert isinstance(cell.comment, Comment)
        assert "10-K filing 2024" in cell.comment.text

    def test_attach_source_comment_empty_does_nothing(self):
        wb = Workbook()
        cell = wb.active["A1"]
        attach_source_comment(cell, "")
        assert cell.comment is None


# ── Number-format variants ───────────────────────────────────────────────────


class TestFormatVariants:
    """Variants (dash, parens) produce the right Excel format strings."""

    def test_number_dash_format(self):
        assert NUMBER_FORMAT_VARIANTS["dash"] == "#,##0;(#,##0);-"

    def test_number_parens_format(self):
        assert NUMBER_FORMAT_VARIANTS["parens"] == "#,##0;(#,##0)"

    def test_currency_with_variant(self):
        result = _apply_format_variant("$#,##0.00", "dash")
        assert result == "$#,##0.00;($#,##0.00);-"

    def test_currency_with_parens_variant(self):
        result = _apply_format_variant("€#,##0.00", "parens")
        assert result == "€#,##0.00;(€#,##0.00)"

    def test_percent_dash_variant(self):
        assert _apply_percent_format_variant("dash") == "0.0%;(0.0%);-"

    def test_percent_no_variant_returns_default_format(self):
        assert _apply_percent_format_variant(None) == "0.0%"

    def test_percent_integer_variant_returns_no_decimal(self):
        assert _apply_percent_format_variant("integer") == "0%"


# ── Multiples format ─────────────────────────────────────────────────────────


class TestMultiplesFormat:
    """Valuation multiples render as '12.5x' via the multiple type."""

    def test_multiples_format_variants_dict(self):
        assert MULTIPLES_FORMAT_VARIANTS["default"] == '0.0"x"'
        assert MULTIPLES_FORMAT_VARIANTS["dash"] == '0.0"x";(0.0"x");-'
        assert MULTIPLES_FORMAT_VARIANTS["parens"] == '0.0"x";(0.0"x")'

    def test_apply_multiples_format_variant(self):
        assert _apply_multiples_format_variant(None) == '0.0"x"'
        assert _apply_multiples_format_variant("dash") == '0.0"x";(0.0"x");-'
        assert _apply_multiples_format_variant("default") == '0.0"x"'
        assert _apply_multiples_format_variant("bogus") == '0.0"x"'

    def test_multiple_type_parses_value(self):
        for raw, expected in [("12.5", 12.5), ("12.5x", 12.5), ("8", 8.0)]:
            wb = Workbook()
            cell = wb.active["A1"]
            _apply_column_type(cell, raw, "multiple")
            assert cell.value == expected
            assert cell.number_format == '0.0"x"'

    def test_multiple_type_with_dash_variant(self):
        wb = Workbook()
        cell = wb.active["A1"]
        _apply_column_type(cell, "12.5", "multiple:dash")
        assert cell.value == 12.5
        assert cell.number_format == '0.0"x";(0.0"x");-'

    def test_multiple_type_unparseable_keeps_text(self):
        wb = Workbook()
        cell = wb.active["A1"]
        _apply_column_type(cell, "not a number", "multiple")
        assert cell.value == "not a number"

    def test_multiple_type_e2e_via_markdown(self):
        markdown = """<!-- types: multiple -->
| EV/EBITDA |
|-----------|
| 12.5      |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].value == 12.5
        assert ws["A2"].number_format == '0.0"x"'

    def test_number_multiple_alias_parses_value(self):
        for alias in ("number:multiple", "number:multiples"):
            wb = Workbook()
            cell = wb.active["A1"]
            _apply_column_type(cell, "12.5x", alias)
            assert cell.value == 12.5, f"alias {alias!r} did not parse as number"
            assert cell.number_format == '0.0"x"', f"alias {alias!r} did not get multiples format"

    def test_number_multiple_alias_e2e_via_markdown(self):
        markdown = """<!-- types: text, number:multiple -->
| Label | Multiple |
|-------|----------|
| A     | 12.5x    |
| B     | =B2*2    |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        assert ws["B2"].value == 12.5
        assert ws["B3"].value == 25


# ── End-to-end general styling ───────────────────────────────────────────────


class TestEndToEndStyling:
    """markdown_to_excel applies general-purpose styling correctly."""

    def test_default_font_applied_to_all_cells(self):
        markdown = """| Col |
|-----|
| abc |
"""
        data = _intercept_upload(markdown, default_font="Times New Roman")
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A1"].font.name == "Times New Roman"
        assert ws["A2"].font.name == "Times New Roman"

    def test_source_directive_attaches_comment(self):
        markdown = """<!-- sources: B2=Source: Annual Report 2024 -->
| Item | Value |
|------|-------|
| A    | 100   |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        cell = ws["B2"]
        assert cell.comment is not None
        assert "Annual Report 2024" in cell.comment.text

    def test_number_dash_variant_in_types_directive(self):
        markdown = """<!-- types: number:dash -->
| Value |
|-------|
| 0     |
| -5    |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].number_format == "#,##0;(#,##0);-"
        assert ws["A3"].number_format == "#,##0;(#,##0);-"

    def test_currency_with_parens_variant(self):
        markdown = """<!-- types: currency:$:parens -->
| Amount |
|--------|
| $100   |
| ($50)  |
"""
        data = _intercept_upload(markdown)
        (OUTPUT_DIR / "currency_parens.xlsx").write_bytes(data)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].number_format == "$#,##0.00;($#,##0.00)"
        assert ws["A3"].number_format == "$#,##0.00;($#,##0.00)"
        assert ws["A2"].value == 100.0

    def test_formula_cells_get_light_blue_fill(self):
        """Formula cells always receive the light-blue formula_fill."""
        markdown = """| A | B |
|---|---|
| 1 | =A2*2 |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        fill = ws["B2"].fill
        assert fill.fill_type == "solid"
        assert fill.fgColor.rgb.endswith("E7F3FF")

    def test_source_directive_attaches_comment_no_financial_mode(self):
        """Source comments work without any financial-mode flag."""
        markdown = """<!-- sources: A2=Source: Internal model -->
| Metric |
|--------|
| 42     |
"""
        data = _intercept_upload(markdown)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A2"].comment is not None
        assert "Internal model" in ws["A2"].comment.text
