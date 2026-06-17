"""Pure-Python Excel formula evaluation via the `formulas` library.

This module isolates the third-party `formulas` library behind a small,
typed interface so that:

- callers (``xlsx_tools.base_xlsx_tool``) never import ``formulas`` directly,
- the import is lazy (only paid when recalculation is actually requested),
- the library's quirks (filename-scoped keys, numpy-array values, tqdm
  progress bars on stderr) are absorbed here rather than leaking out.

Nothing in this module shells out to an external binary. The `formulas`
library is pure Python and EUPL-1.1+ licensed, which is compatible with
the MIT license of this server.
"""

from __future__ import annotations

import logging
import os
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Union

logger = logging.getLogger(__name__)


# The seven Excel error sentinels as defined by the OOXML spec. Any cell
# whose evaluated value contains one of these strings is a formula error
# that the caller almost certainly wants to know about before the file
# ships.
EXCEL_ERRORS: tuple[str, ...] = (
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)


@dataclass
class CellError:
    """A single formula error detected during recalculation."""

    sheet: str          # Original (case-correct) sheet name as stored in the workbook.
    coordinate: str     # Excel coordinate, e.g. "B5".
    error_type: str     # One of EXCEL_ERRORS.

    @property
    def location(self) -> str:
        """``Sheet!Cell`` reference suitable for showing to the caller."""
        # Quote sheet names that contain spaces or special characters.
        if re.search(r"[^A-Za-z0-9_]", self.sheet):
            return f"'{self.sheet}'!{self.coordinate}"
        return f"{self.sheet}!{self.coordinate}"

    def __str__(self) -> str:
        return f"{self.location}: {self.error_type}"


@dataclass
class RecalcResult:
    """Outcome of attempting to recalculate a workbook.

    Attributes:
        recalc_performed: True if the engine ran at all. False when the
            engine could not be initialised (e.g. dependency missing) or
            raised before producing a solution.
        values_map: Mapping of ``"Sheet!Cell"`` -> computed Python scalar.
            Only cells that contained a formula AND produced a non-error
            value are included. Numeric values are unwrapped from numpy
            types to native Python int/float/bool. String results are
            included and serialized with ``t="str"`` (inline string, no
            shared-strings-table entry required).
        errors: Formula errors detected. Empty if none.
        skip_reason: Human-readable reason when recalc was skipped
            (engine unavailable or engine raised). None otherwise.
    """

    recalc_performed: bool = False
    values_map: dict[str, Any] = field(default_factory=dict)
    errors: list[CellError] = field(default_factory=list)
    skip_reason: str | None = None
    total_formulas: int = 0

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


# ── Internal helpers ─────────────────────────────────────────────────────────


# Match the `[filename]SHEET` prefix the `formulas` library prepends to
# every solution key, e.g. ``'[test.xlsx]REVENUE'!B2``. We strip the
# filename and capture the (uppercased) sheet name and the cell ref.
_KEY_RE = re.compile(
    r"""
    ^'?\[               # optional opening quote + literal [
        [^\]]+          #   filename (anything but ']')
    \]                  # literal ]
    ([^']+)             # group 1: sheet name (uppercased by the library)
    '?                  # optional closing quote
    !                   # bang
    ([A-Za-z]+\$?\d+)   # group 2: cell coordinate (e.g. B2, $A$1)
    $
    """,
    re.VERBOSE,
)


def _build_sheet_lookup(sheet_names: list[str]) -> dict[str, str]:
    """Map uppercased sheet name -> original sheet name as stored in the workbook.

    The `formulas` library uppercases sheet names in its solution keys.
    We need to recover the original casing so that the locations we
    report back to the user (and the keys in ``values_map``) match what
    openpyxl wrote.
    """
    lookup: dict[str, str] = {}
    for name in sheet_names:
        lookup[name.upper()] = name
    return lookup


def _unwrap_scalar(value: Any) -> Any:
    """Pull a native Python scalar out of a `formulas` solution value.

    The library wraps single-cell results in a numpy 2D array of shape
    ``(1, 1)`` (or occasionally a ``Ranges`` object whose ``.value`` is
    such an array). We extract ``arr[0][0]`` and convert numpy scalar
    types to native Python types so that downstream XML serialisation
    and openpyxl round-trips behave predictably.

    Returns ``None`` if the value cannot be unwrapped to a scalar — the
    caller treats ``None`` as "do not inject a cached value".
    """
    # `Ranges` and similar objects expose the underlying array via `.value`.
    arr = getattr(value, "value", value)

    # Numpy arrays and array-likes.
    try:
        # Shape (1, 1) or (1,) — pull the single element.
        if hasattr(arr, "shape") and hasattr(arr, "__getitem__"):
            shape = arr.shape
            if shape == (1, 1):
                arr = arr[0][0]
            elif shape == (1,):
                arr = arr[0]
            elif shape == ():
                arr = arr.item()
            else:
                # Multi-cell result (e.g. array formula) — too complex to
                # inject as a single cached value; skip.
                return None
    except Exception:
        return None

    # Convert numpy scalar types to native Python types.
    # bool first because numpy bool_ is also detected as integer.
    try:
        import numpy as np  # local import; numpy is a transitive dep of `formulas`
    except ImportError:
        np = None

    if np is not None and isinstance(arr, np.generic):
        arr = arr.item()

    # The `formulas` library always returns floats for arithmetic results
    # (e.g. ``=A1+5`` with A1=5 yields ``10.0``, not ``10``). Convert
    # whole-number floats back to int — the XML is cleaner (``<v>10</v>``
    # vs ``<v>10.0</v>``) and Excel stores them equivalently. Skip this
    # for values that exceed the safe-integer range to avoid precision loss.
    if isinstance(arr, float) and arr.is_integer() and abs(arr) < 2**53:
        arr = int(arr)

    return arr


def _is_error_string(value: Any) -> str | None:
    """Return the matching Excel error sentinel if ``value`` is one, else None."""
    if not isinstance(value, str):
        return None
    for err in EXCEL_ERRORS:
        if value == err or err in value:
            return err
    return None


def _suppress_tqdm() -> None:
    """Tell tqdm (used by the `formulas` library) to be quiet during recalculation.

    Without this the library writes a progress bar to stderr on every
    workbook calculation, which pollutes server logs.
    """
    os.environ.setdefault("TQDM_DISABLE", "1")


def _collect_formula_cells(
    xlsx_bytes_or_path: bytes | str,
    sheet_lookup: dict[str, str],
) -> set[str] | None:
    """Return the set of ``"Sheet!Cell"`` keys for cells that contain a formula.

    Used to filter the recalc engine's solution dict (which includes every
    cell, not just formulas) down to just the formula cells we actually
    want to cache. Returns ``None`` if the workbook can't be parsed — the
    caller treats None as "filter disabled" and caches everything (the
    injection layer filters by presence of ``<f>`` anyway, so this is
    just an optimisation, not a correctness requirement).
    """
    try:
        from openpyxl import load_workbook
        import io as _io

        if isinstance(xlsx_bytes_or_path, (bytes, bytearray)):
            wb = load_workbook(_io.BytesIO(xlsx_bytes_or_path))
        else:
            wb = load_workbook(str(xlsx_bytes_or_path))
    except Exception as e:
        logger.debug("Formula-cell collection skipped (workbook parse failed): %s", e)
        return None

    formula_keys: set[str] = set()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_keys.add(_format_location(sheet_name, cell.coordinate))
    finally:
        wb.close()

    return formula_keys


# A formula references an external workbook when it contains a sheet
# reference whose workbook name is wrapped in square brackets — e.g.
# `=[Budget.xlsx]Sheet1!A1` or `='[C:\Reports\FY24.xlsx]P&L'!B2`. The
# `formulas` library can't resolve these (the other workbook isn't
# available in-process) and aborts the whole workbook with FormulaError.
# Match the bracketed-bookbook pattern after the leading '='.
_EXTERNAL_LINK_RE = re.compile(r"=\s*'?\[", re.IGNORECASE)


def _blank_external_link_formulas(xlsx_path: str) -> list[str]:
    """Blank out external-link formulas in ``xlsx_path`` in place.

    Opens the temp copy of the workbook that is handed to the recalc
    engine and clears the value of any cell whose formula references an
    external workbook (e.g. ``=[Other.xlsx]Sheet1!A1``). Returns the
    list of affected ``"Sheet!Cell"`` keys for logging/telemetry.

    This keeps a single unsupported cell from suppressing cached values
    for every other formula in the file. The user's original file is
    never modified — only the temp copy the engine loads from — and
    Excel will recalc the external links natively when the file opens.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        logger.debug("External-link scan skipped (workbook parse failed): %s", e)
        return []

    skipped: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if (
                        isinstance(val, str)
                        and val.startswith("=")
                        and _EXTERNAL_LINK_RE.match(val)
                    ):
                        skipped.append(_format_location(sheet_name, cell.coordinate))
                        cell.value = None
        if skipped:
            wb.save(xlsx_path)
    except Exception as e:
        logger.debug("External-link blanking failed: %s", e)
    finally:
        wb.close()

    return skipped


# ── Public API ───────────────────────────────────────────────────────────────


def is_available() -> bool:
    """Return True if the `formulas` library can be imported.

    Used by callers to decide whether to attempt recalculation at all
    without paying the import cost on the hot path.
    """
    try:
        import formulas  # noqa: F401
        return True
    except ImportError:
        return False


def recalculate_workbook(
    xlsx_bytes_or_path: Union[bytes, str, Path],
    sheet_names: list[str],
) -> RecalcResult:
    """Evaluate every formula in an XLSX workbook in-process.

    Args:
        xlsx_bytes_or_path: Either the raw xlsx bytes produced by
            ``openpyxl.Workbook.save`` or a path to an xlsx file on disk.
            Passing a path avoids a temp-file write and is preferred when
            the caller already has one.
        sheet_names: Sheet names as stored in the workbook (original
            casing). Used to translate the uppercased sheet names in the
            library's solution keys back to the caller's casing.

    Returns:
        A :class:`RecalcResult`. The ``values_map`` keys use the
        ``"Sheet!Cell"`` form (sheet name quoted if it contains spaces),
        matching how Excel addresses cells in cross-sheet formulas.

    The function never raises for engine-level failures (unsupported
    functions, parse errors, missing dependencies). In those cases it
    returns a result with ``recalc_performed=False`` and a
    ``skip_reason``. Formula-level errors (``#REF!`` etc.) are reported
    via ``result.errors`` rather than raised.
    """
    if not is_available():
        return RecalcResult(
            recalc_performed=False,
            skip_reason="`formulas` library not installed",
        )

    sheet_lookup = _build_sheet_lookup(sheet_names)
    result = RecalcResult(recalc_performed=True)

    # Materialise the workbook on disk for the library to load. `formulas`
    # only accepts a file path, not a stream.
    tmp_path: str | None = None
    try:
        if isinstance(xlsx_bytes_or_path, (bytes, bytearray)):
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as tmp:
                tmp.write(xlsx_bytes_or_path)
                tmp_path = tmp.name
            load_path = tmp_path
        else:
            load_path = str(xlsx_bytes_or_path)

        _suppress_tqdm()

        # Lazy import so a missing/optional dependency doesn't slow down
        # callers that aren't using recalculation.
        import formulas

        # Pre-pass: collect the set of cells that actually contain formulas.
        # The `formulas` library's solution dict includes EVERY cell (inputs
        # and formulas alike), so without this filter we'd build a values_map
        # 5-10x larger than needed and log a misleading "N/M formulas cached"
        # count. We only want to cache values for cells whose original content
        # was a formula (= prefix).
        formula_cells = _collect_formula_cells(xlsx_bytes_or_path, sheet_lookup)

        # Isolate external-link formulas. The `formulas` library aborts the
        # ENTIRE workbook with FormulaError when it encounters an external-
        # workbook reference like `=[Other.xlsx]Sheet1!A1`. A single such
        # cell anywhere would otherwise suppress cached values for every
        # formula in the file. We blank those cells in the temp copy handed
        # to the engine (the user's file is untouched; Excel recalcs external
        # links on open anyway). External-link cells are recorded as skipped.
        skipped_external = _blank_external_link_formulas(load_path)
        if skipped_external:
            logger.info(
                "External-link formulas not recalculated in-process (%d cell(s)); "
                "Excel will evaluate them on open: %s",
                len(skipped_external),
                ", ".join(skipped_external[:10]),
            )

        logger.debug("Loading workbook into formulas engine: %s", load_path)
        model = formulas.ExcelModel().loads(load_path).finish()
        solution = model.calculate()

        for raw_key, value in solution.items():
            match = _KEY_RE.match(raw_key)
            if not match:
                continue

            upper_sheet, coordinate = match.group(1), match.group(2)
            original_sheet = sheet_lookup.get(upper_sheet, upper_sheet)

            # Only cache values for cells that actually contained a formula.
            # The engine returns values for input cells too; injecting those
            # would be wasted work (xml_cache filters them anyway) and
            # skews the "N/M formulas cached" log line.
            location_key_check = _format_location(original_sheet, coordinate)
            if formula_cells is not None and location_key_check not in formula_cells:
                continue

            scalar = _unwrap_scalar(value)
            if scalar is None:
                continue

            # Error sentinel? Record and skip injection.
            err = _is_error_string(scalar)
            if err is not None:
                result.errors.append(
                    CellError(
                        sheet=original_sheet,
                        coordinate=coordinate,
                        error_type=err,
                    )
                )
                continue

            location_key = _format_location(original_sheet, coordinate)
            result.values_map[location_key] = scalar

    except Exception as e:
        # Anything from the engine (unsupported feature, parse failure,
        # out-of-memory, ...) is treated as "skip recalc, ship without
        # cached values". Don't fail the whole document-generation call.
        logger.warning(
            "Excel formula recalculation skipped due to engine error: %s",
            e,
            exc_info=True,
        )
        return RecalcResult(
            recalc_performed=False,
            skip_reason=f"recalculation engine error: {e}",
        )
    finally:
        if tmp_path is not None:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    # Count total formulas for telemetry / result reporting. Best-effort.
    try:
        from .xml_cache import count_formulas
        if isinstance(xlsx_bytes_or_path, (bytes, bytearray)):
            result.total_formulas = count_formulas(bytes(xlsx_bytes_or_path))
    except Exception:
        pass

    logger.info(
        "Excel recalc complete: %d/%d formulas cached, %d formula errors",
        len(result.values_map),
        result.total_formulas,
        len(result.errors),
    )
    return result


def _format_location(sheet: str, coordinate: str) -> str:
    """Build the ``"Sheet!Cell"`` key used in ``RecalcResult.values_map``."""
    if re.search(r"[^A-Za-z0-9_]", sheet):
        return f"'{sheet}'!{coordinate}"
    return f"{sheet}!{coordinate}"


# ── Circular-reference detection ─────────────────────────────────────────────
#
# The `formulas` library silently resolves circular references to nothing
# (0 cached values, 0 errors) rather than flagging them. Excel itself
# shows a warning dialog and a 0 result. To catch this class of bug at
# generation time — financial models must contain no unintended circular
# references — we build a dependency graph from the formula
# strings and run a DFS cycle detection. Any cell on a cycle is reported
# as a synthetic "#CIRC!" error (not an OOXML sentinel; our own marker)
# so it surfaces through the existing error-reporting policy.

# Synthetic error type for circular references. Deliberately distinct from
# the seven OOXML sentinels in EXCEL_ERRORS so callers can tell that this
# was detected by our graph analysis, not by the recalc engine.
CIRCULAR_ERROR_TYPE = "#CIRC!"


# Parse cell references out of a formula string. Handles:
#   - Cross-sheet:  Sheet1!A1, 'Sheet 1'!A1, [Book.xlsx]Sheet1!A1
#   - Local:        A1, $A$1
#   - Ranges:       A1:B5 (expanded to individual cells), Sheet!A1:B5
# We deliberately keep this regex-based rather than using a full formula
# parser — we only need the references, not the AST, and false positives
# (e.g. a literal string "A1" inside a formula) are tolerable for
# cycle detection because they'd only produce a spurious edge, not a
# missed cycle.
_REF_TOKENS_RE = re.compile(
    r"""
    (?:                                 # optional sheet/book prefix
        (?:\[ [^\]]+ \])?               #   optional [Workbook.xlsx]
        '?                              #   optional opening quote
        (?: ([A-Za-z_][\w\s]*) )        #   group 1: sheet name (greedy-ish)
        '?                              #   optional closing quote
        (?:                             #   optional :sheet range (3D ref)
            :                           #     literal colon
            '?                          #     optional opening quote
            ([A-Za-z_][\w\s]*)          #     group 2: end sheet name
            '?                          #     optional closing quote
        )?
        !                               #   bang
    )?
    (\$?[A-Z]{1,3}\$?\d{1,7})           # group 3: first cell (e.g. B2, $A$1)
    (?::                               # optional :cell range
        (\$?[A-Z]{1,3}\$?\d{1,7})       #   group 4: second cell
    )?
    """,
    re.IGNORECASE | re.VERBOSE,
)


def _normalize_coord(coord: str) -> str:
    """Strip absolute-reference markers from a coordinate: ``$A$1`` → ``A1``."""
    return coord.replace("$", "").upper()


def _expand_range(start: str, end: str) -> list[str]:
    """Expand an Excel range like ('A1','B3') into ['A1','A2','A3','B1','B2','B3']."""
    import re as _re
    start_m = _re.match(r"([A-Z]+)(\d+)", start.upper())
    end_m = _re.match(r"([A-Z]+)(\d+)", end.upper())
    if not start_m or not end_m:
        return [start]
    # Convert column letters to numbers for iteration.
    def col_to_num(c: str) -> int:
        n = 0
        for ch in c:
            n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
        return n
    def num_to_col(n: int) -> str:
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(ord("A") + r) + s
        return s
    c1, r1 = col_to_num(start_m.group(1)), int(start_m.group(2))
    c2, r2 = col_to_num(end_m.group(1)), int(end_m.group(2))
    if c1 > c2: c1, c2 = c2, c1
    if r1 > r2: r1, r2 = r2, r1
    # Cap expansion to avoid blowing up on huge ranges (e.g. A1:Z1000000).
    # A circular reference through a 1000-cell range is implausible; if
    # the range is bigger, just use the corners.
    if (c2 - c1 + 1) * (r2 - r1 + 1) > 1000:
        return [_normalize_coord(start), _normalize_coord(end)]
    cells = []
    for c in range(c1, c2 + 1):
        for r in range(r1, r2 + 1):
            cells.append(f"{num_to_col(c)}{r}")
    return cells


def _is_valid_coord(coord: str) -> bool:
    """Return True if ``coord`` is a real Excel cell coordinate (e.g. A1, XFD1048576).

    Used to filter out phantom matches from the regex — e.g. ``Table1[Revenue]``
    yields a ``BLE1`` match that isn't a valid coordinate. Validation:
    column letters 1-3 chars and at most XFD (16384); row 1-1048576.
    """
    import re as _re
    m = _re.match(r"^([A-Z]{1,3})(\d+)$", coord.upper())
    if not m:
        return False
    col, row = m.group(1), int(m.group(2))
    # Row must fit Excel's 1,048,576-row limit.
    if row < 1 or row > 1048576:
        return False
    # Column must fit Excel's 16,384-column limit (XFD).
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return 1 <= n <= 16384


def _strip_string_literals(formula: str) -> str:
    """Remove double-quoted string literals from a formula string.

    Excel string literals are delimited by ``"`` and use ``""`` for an
    escaped quote inside the literal. Cell-coordinate patterns (``A1``,
    ``B2``) appearing *inside* a string literal are NOT cell references —
    they're just text — so we strip them before regex-matching for refs.

    Without this, a formula like ``="see cell A1 for context"`` would
    produce a phantom ``A1`` reference, and a self-referential label
    like ``=A1&" totals"`` placed in A1 would falsely report a circular
    reference.

    We replace each literal with a single space (not the empty string)
    so adjacent tokens on either side of the literal don't accidentally
    fuse into a new coord-like token.
    """
    out: list[str] = []
    i = 0
    n = len(formula)
    while i < n:
        ch = formula[i]
        if ch == '"':
            # Skip until the closing unescaped quote. An escaped quote
            # inside the literal is "" (two consecutive double quotes).
            i += 1
            while i < n:
                if formula[i] == '"':
                    if i + 1 < n and formula[i + 1] == '"':
                        # Escaped quote — skip both chars.
                        i += 2
                        continue
                    # Unescaped closing quote.
                    i += 1
                    break
                i += 1
            out.append(" ")  # placeholder so tokens don't fuse
        else:
            out.append(ch)
            i += 1
    return "".join(out)


def extract_formula_references(
    formula: str,
    current_sheet: str,
    sheet_lookup: dict[str, str],
) -> set[str]:
    """Extract the set of ``"Sheet!Cell"`` references a formula depends on.

    Args:
        formula: The formula string (with or without leading ``=``).
        current_sheet: The sheet name where this formula lives — used to
            resolve unqualified references like ``A1``.
        sheet_lookup: Map of uppercased sheet name → original casing,
            so we can normalize cross-sheet refs to the workbook's actual
            sheet names.

    Returns:
        Set of normalized ``"Sheet!Cell"`` location keys. References to
        sheets that don't exist in ``sheet_lookup`` are dropped (they'll
        be caught as #REF! by the recalc engine if real).
    """
    if not formula:
        return set()

    # Strip string literals first so coord-like text inside them (e.g.
    # "see A1" or a VLOOKUP key "A1") isn't mistaken for a cell ref.
    cleaned = _strip_string_literals(formula)

    refs: set[str] = set()
    for match in _REF_TOKENS_RE.finditer(cleaned):
        sheet_raw, sheet_end_raw, cell_start, cell_end = (
            match.group(1),
            match.group(2),
            match.group(3),
            match.group(4),
        )

        if sheet_raw:
            # Resolve the (start) sheet. For a 3D ref like Sheet1:Sheet3!A1,
            # expand to every sheet in the workbook between the two names.
            start_sheet = sheet_lookup.get(sheet_raw.strip().upper())
            if start_sheet is None:
                # Unknown sheet — skip; recalc engine will flag #REF! if real.
                continue

            if sheet_end_raw:
                # 3D reference (Sheet1:Sheet3!A1). Expand across every sheet
                # in workbook order between start and end, inclusive. We use
                # workbook order from sheet_lookup's values rather than
                # alphabetical, matching Excel's behaviour.
                end_sheet = sheet_lookup.get(sheet_end_raw.strip().upper())
                if end_sheet is None:
                    continue
                # Preserve the workbook's sheet ordering for expansion.
                ordered_sheets = list(sheet_lookup.values())
                try:
                    i_start = ordered_sheets.index(start_sheet)
                    i_end = ordered_sheets.index(end_sheet)
                    if i_start > i_end:
                        i_start, i_end = i_end, i_start
                    target_sheets = ordered_sheets[i_start : i_end + 1]
                except ValueError:
                    target_sheets = [start_sheet]
            else:
                target_sheets = [start_sheet]
        else:
            # No sheet prefix — resolve to the formula's own sheet.
            target_sheets = [current_sheet]

        for sheet in target_sheets:
            if cell_end:
                # Cell range: expand to individual cells.
                start_norm = _normalize_coord(cell_start)
                end_norm = _normalize_coord(cell_end)
                # Drop the match entirely if either corner isn't a valid
                # Excel coordinate (e.g. phantom BLE1 from "Table1[Rev]").
                if not _is_valid_coord(start_norm) or not _is_valid_coord(end_norm):
                    continue
                for coord in _expand_range(start_norm, end_norm):
                    refs.add(_format_location(sheet, coord))
            else:
                coord = _normalize_coord(cell_start)
                if not _is_valid_coord(coord):
                    continue
                refs.add(_format_location(sheet, coord))

    return refs


def detect_circular_references(
    xlsx_bytes: bytes,
    sheet_names: list[str],
) -> list[CellError]:
    """Detect circular references in a workbook by graph analysis.

    Builds a directed dependency graph (formula cell → cells it references)
    and runs DFS to find cycles. Every cell that participates in a cycle
    is reported as a :class:`CellError` with ``error_type="#CIRC!"``.

    This is independent of the `formulas` library (which silently misses
    circular refs) — it works purely from the formula strings and the
    workbook structure.

    Returns an empty list if no cycles are found or if the workbook
    can't be parsed (best-effort, never raises).
    """
    try:
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(xlsx_bytes))
    except Exception as e:
        logger.debug("Circular-ref detection skipped (workbook parse failed): %s", e)
        return []

    sheet_lookup = _build_sheet_lookup(sheet_names)

    # Build the dependency graph: node "Sheet!Cell" → set of dependency nodes.
    graph: dict[str, set[str]] = {}
    # Track which nodes are formula cells (only these can be cycle members
    # that we report — a literal cell referenced by a formula isn't itself
    # circular even if it's in the graph).
    formula_nodes: set[str] = set()

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                node = _format_location(sheet_name, cell.coordinate)
                deps = extract_formula_references(
                    cell.value, sheet_name, sheet_lookup
                )
                graph[node] = deps
                formula_nodes.add(node)

    if not graph:
        return []

    # DFS cycle detection. Standard 3-color algorithm: WHITE (unvisited),
    # GRAY (on current stack), BLACK (fully processed). A back-edge to a
    # GRAY node means a cycle.
    WHITE, GRAY, BLACK = 0, 1, 2
    color: dict[str, int] = {n: WHITE for n in graph}
    # Collect one representative cell per cycle (the entry point where we
    # detected the back-edge) — reporting every cell in every cycle would
    # be noisy for large interconnected cycles.
    cyclic_nodes: set[str] = set()

    def dfs(node: str, stack: list[str]) -> None:
        color[node] = GRAY
        stack.append(node)
        for dep in graph.get(node, ()):
            if dep not in color:
                # Dependency isn't itself a formula node (it's a literal
                # cell or external). Skip — it can't participate in a cycle.
                continue
            if color[dep] == GRAY:
                # Back-edge: cycle detected. Mark every cell on the cycle
                # path from `dep` to `node` (inclusive) as circular.
                cycle_start = stack.index(dep)
                for cyclic in stack[cycle_start:]:
                    cyclic_nodes.add(cyclic)
            elif color[dep] == WHITE:
                dfs(dep, stack)
        stack.pop()
        color[node] = BLACK

    for node in list(graph.keys()):
        if color[node] == WHITE:
            dfs(node, [])

    # Convert cyclic nodes to CellError objects.
    errors: list[CellError] = []
    for location in sorted(cyclic_nodes):
        sheet, coord = _parse_location(location)
        errors.append(
            CellError(
                sheet=sheet,
                coordinate=coord,
                error_type=CIRCULAR_ERROR_TYPE,
            )
        )

    if errors:
        logger.warning(
            "Circular-reference detection found %d cell(s) in cycles", len(errors)
        )
    return errors


def _parse_location(location_key: str) -> tuple[str, str]:
    """Parse ``"Sheet!Cell"`` or ``"'Sheet Name'!Cell"`` into (sheet, cell).

    Inverse of :func:`_format_location`.
    """
    if location_key.startswith("'"):
        close = location_key.find("'", 1)
        if close != -1 and location_key[close + 1:close + 2] == "!":
            return location_key[1:close], location_key[close + 2:]
    parts = location_key.split("!", 1)
    if len(parts) == 1:
        return location_key, ""
    return parts[0], parts[1]
