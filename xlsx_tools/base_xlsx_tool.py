import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import SheetTitleException

from config import get_config
from upload_tools import upload_file
from .helpers import add_table_to_sheet
from .parser import (
    walk_markdown_lines,
    collect_table_positions,
    SheetEvent,
    HeaderEvent,
    TableEvent,
    DEFAULT_SHEET_NAME,
    _sanitize_sheet_name,
)

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
# Header font styles by level
HEADER_FONTS = {
    1: Font(size=16, bold=True, color="2F5597"),
    2: Font(size=14, bold=True, color="4472C4"),
}
HEADER_FONT_DEFAULT = Font(size=12, bold=True)


def markdown_to_excel(
    markdown_content: str,
    file_name: str | None = None,
    auto_filter: bool = False,
    default_font: str | None = None,
    financial_modeling: bool = False,
    recalc: bool | None = None,
) -> str:
    """Convert Markdown to Excel workbook (focused on tables and headers).

    Always starts from an empty Workbook (no templates).
    Supports multiple sheets via '## Sheet: Name' headings.
    Supports cross-sheet references via ``SheetName!T1.B[0]`` syntax.

    Args:
        markdown_content: Markdown string with tables.
        file_name: Optional custom filename (without extension).
        auto_filter: If True, apply Excel auto-filter to each table.
        default_font: Optional font family applied to every cell
            (e.g. 'Arial'). When None, falls back to the server's
            ``XLSX_DEFAULT_FONT`` config value, then openpyxl's default.
        financial_modeling: When True, apply CFA-standard color coding
            (blue inputs, black formulas, green cross-sheet refs, yellow
            sourced cells) and treat 4-digit-year data cells as text.
        recalc: When True, evaluate every formula in-process via the
            pure-Python ``formulas`` library and write the computed
            values back as cached <v> tags so the file previews
            correctly without Excel. When None (default), falls back to
            the server's ``XLSX_RECALC_ENABLED`` config value.
            If the caller explicitly passes recalc=True and recalculation
            detects formula errors (#REF!, #DIV/0!, etc.), the call
            fails with a descriptive RuntimeError so the model can fix
            the formulas and retry. When recalc runs as a defaulted
            behaviour, errors are logged but the file is still delivered.

    Raises:
        RuntimeError: If the markdown contains no tables, conversion
            fails, or (only when ``recalc`` is explicitly True) formula
            evaluation detects errors like #REF! or #DIV/0!.
    """
    logger.info("Starting markdown_to_excel conversion")

    # Track whether the caller explicitly requested recalculation. When
    # explicit, formula errors (#REF!, #DIV/0!, ...) cause the tool call
    # to fail with a descriptive message so the model can fix the formulas
    # and retry — matching the "zero formula errors" delivery standard.
    # When recalc runs as a defaulted behaviour (the caller didn't pass
    # the parameter), errors are logged but the file is still delivered,
    # so that a misconfigured environment can never break document delivery.
    recalc_explicitly_requested = recalc is not None

    # Resolve config-driven defaults for optional behaviour.
    try:
        cfg = get_config()
    except Exception:
        cfg = None
    if default_font is None and cfg is not None:
        default_font = cfg.xlsx_default_font
    if recalc is None and cfg is not None:
        recalc = cfg.xlsx_recalc_enabled
    elif recalc is None:
        recalc = True

    # Recalc timeout (seconds). Falls back to config, then 30s.
    recalc_timeout_seconds = 30
    if cfg is not None:
        recalc_timeout_seconds = cfg.xlsx_recalc_timeout_seconds

    # ── Input validation ──
    if not markdown_content or not markdown_content.strip():
        raise RuntimeError("Cannot create Excel workbook: markdown content is empty")

    # Split content into lines and parse into events (single shared state machine)
    lines: list[str] = markdown_content.split('\n')
    events = walk_markdown_lines(lines)

    # Build table position map from events (used for cross-sheet formula resolution)
    all_sheet_table_positions = collect_table_positions(events)
    logger.debug("Table positions (all sheets): %s", all_sheet_table_positions)

    # ── Build the actual workbook from events ──
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(DEFAULT_SHEET_NAME)

    # Per-sheet state for formula resolution
    table_positions: dict[str, int] = {}

    # Counters for summary
    headers_count = 0
    tables_count = 0

    try:
        # Track the actual worksheet titles as openpyxl stores them. When a
        # `## Sheet:` header sanitizes to a name that already exists (e.g. two
        # long names that collide after the 31-char truncation, or a duplicate
        # name), openpyxl silently appends a suffix to the new sheet. The
        # cross-sheet position map is keyed by the *requested* name, so the
        # divergence silently routes formulas to the wrong sheet. We warn so
        # the model knows to rename.
        seen_sheet_titles: set[str] = set()
        # Seed with the default sheet's title (set above).
        seen_sheet_titles.add(ws.title)

        for event in events:
            if isinstance(event, SheetEvent):
                if event.is_rename:
                    try:
                        ws.title = event.sheet_name
                        seen_sheet_titles = {event.sheet_name}
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Could not rename worksheet to '%s': %s — using default",
                            event.sheet_name, exc,
                        )
                else:
                    requested = event.sheet_name
                    if requested in seen_sheet_titles:
                        logger.warning(
                            "Sheet name '%s' collides with an existing sheet after "
                            "sanitization; openpyxl will auto-rename it, which may "
                            "break cross-sheet references using the original name. "
                            "Use a distinct sheet name.",
                            requested,
                        )
                    try:
                        ws = wb.create_sheet(title=requested)
                        seen_sheet_titles.add(ws.title)
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Invalid sheet name '%s': %s — using fallback",
                            event.sheet_name, exc,
                        )
                        ws = wb.create_sheet()
                        seen_sheet_titles.add(ws.title)
                    table_positions = {}

            elif isinstance(event, HeaderEvent):
                cell = ws.cell(row=event.row, column=1)
                cell.value = event.text
                base_font = HEADER_FONTS.get(event.level, HEADER_FONT_DEFAULT)
                # Honour default_font by keeping the size/bold/color from the
                # level-specific font but swapping the family.
                if default_font:
                    cell.font = Font(
                        name=default_font,
                        size=base_font.size,
                        bold=base_font.bold,
                        color=base_font.color,
                    )
                else:
                    cell.font = base_font
                headers_count += 1
                logger.debug("Header (level %d) at row %d: %s", event.level, event.row, event.text)

            elif isinstance(event, TableEvent):
                # Record this table's position for local formula resolution
                table_positions[event.table_key] = event.start_row

                # Write table to worksheet
                add_table_to_sheet(
                    event.table_data, ws, event.start_row, table_positions,
                    all_sheet_table_positions=all_sheet_table_positions,
                    auto_filter=auto_filter,
                    table_index=tables_count,
                    directives=event.directives,
                    default_font=default_font,
                    financial_modeling=financial_modeling,
                )

                # Handle freeze directive — freeze below header row of this table
                if 'freeze' in event.directives:
                    ws.freeze_panes = f"A{event.start_row + 1}"

                tables_count += 1
                logger.debug(
                    "Added table #%d (%s) with %d data rows on sheet '%s'",
                    tables_count, event.table_key, len(event.table_data) - 1, event.sheet_name,
                )

    except Exception as e:
        logger.error("Error generating Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error generating Excel workbook: {e}") from e

    # ── Validation: ensure at least one table was created ──
    if tables_count == 0:
        raise RuntimeError(
            "Cannot create Excel workbook: no valid markdown tables found in the input. "
            "Tables must use pipe syntax (| col1 | col2 |) with a separator row (|---|---|)."
        )

    # Save workbook to BytesIO. If recalc is enabled, we run a second pass
    # to evaluate formulas and inject cached values before upload.
    file_object = io.BytesIO()
    try:
        logger.info(
            "Saving Excel workbook to memory buffer (headers=%d, tables=%d, recalc=%s, financial=%s)",
            headers_count, tables_count, recalc, financial_modeling,
        )
        wb.save(file_object)
        original_bytes = file_object.getvalue()

        final_bytes = original_bytes
        formula_error_summary: str | None = None

        if recalc:
            final_bytes, formula_error_summary = _recalc_and_inject(
                original_bytes, wb.sheetnames, recalc_timeout_seconds
            )
        else:
            # Circular-reference detection runs unconditionally (see below) even
            # when recalc is disabled, but when the engine is also skipped we
            # need to populate formula_error_summary from the standalone
            # circular-ref scan so the explicit-recalc error policy can act on
            # it. Formula-level errors (#REF!, #DIV/0!, ...) are NOT detectable
            # without the recalc engine, so they're simply not reported here.
            circ_summary = _detect_circular_only(original_bytes, wb.sheetnames)
            if circ_summary:
                formula_error_summary = circ_summary

        # Circular-reference detection is intentionally run UNCONDITIONALLY,
        # even when recalc is disabled or unavailable. The DFS-based graph
        # analysis is pure string work (no recalc engine) and catches a bug
        # class (cycles) that the engine itself silently misses. The tool
        # description promises this guarantee, so it must hold regardless of
        # the caller's recalc choice.
        #
        # Note: _recalc_and_inject already runs detection internally when
        # recalc is enabled, so we only need the standalone pass for the
        # recalc-disabled branch above. We do NOT re-run it here to avoid
        # double-counting circular cells in the summary.

        # If formula errors surfaced AND the caller explicitly opted INTO
        # recalculation, fail the tool call with a clear message. This
        # enforces the "zero formula errors" delivery standard. When recalc
        # ran as a defaulted behaviour (parameter not passed), we still
        # deliver the file (errors are logged) so misconfiguration can't
        # break delivery.
        #
        # Note: when the caller explicitly passes recalc=False, circular
        # references are still detected and logged (the tool promises
        # independent cycle analysis), but they are NOT fatal — the user
        # explicitly opted out of validation, and surprising them with a
        # hard failure would be inconsistent with their choice.
        recalc_explicitly_enabled = recalc_explicitly_requested and bool(recalc)
        if formula_error_summary and recalc_explicitly_enabled:
            raise RuntimeError(
                f"Excel workbook contains formula errors. "
                f"Fix the formulas and retry. {formula_error_summary}"
            )

        # Hand the (possibly recalc'd) bytes to the upload layer.
        upload_buffer = io.BytesIO(final_bytes)
        result = upload_file(upload_buffer, "xlsx", filename=file_name)
        logger.info("Excel upload completed successfully")
        return result
    except Exception as e:
        logger.error("Error saving/uploading Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error saving/uploading Excel workbook: {e}") from e
    finally:
        file_object.close()


def _detect_circular_only(
    xlsx_bytes: bytes,
    sheet_names: list[str],
) -> str | None:
    """Run only circular-reference detection and return a grouped summary.

    Used when recalc is disabled but the caller still wants cycle detection
    (the tool description promises it runs even without the recalc engine).
    Returns the human-readable summary string if cycles are found, else None.
    Never raises — failures here are logged and treated as "no cycles".
    """
    try:
        from .formula_engine import detect_circular_references
    except ImportError as e:
        logger.debug("Circular-ref detection unavailable: %s", e)
        return None

    try:
        errors = detect_circular_references(xlsx_bytes, sheet_names)
    except Exception as e:
        logger.debug("Circular-ref detection failed: %s", e)
        return None

    if not errors:
        return None

    summary = _format_grouped_errors(errors)
    logger.warning("Excel circular references: %s", summary)
    return summary


def _recalc_and_inject(
    xlsx_bytes: bytes,
    sheet_names: list[str],
    timeout_seconds: int = 30,
) -> tuple[bytes, str | None]:
    """Evaluate formulas and inject cached values; return (bytes, error_summary).

    Best-effort: if the engine is unavailable, raises, or exceeds the
    timeout, the original bytes are returned unchanged with a short note
    in ``error_summary`` (or None for engine-skip cases). Formula-level
    errors (#REF!, #DIV/0!, ...) AND circular references (#CIRC!) are
    collected into a human-readable summary string grouped by error type.
    Returns None for the summary when there are no errors.

    The recalculation runs in a worker thread bounded by
    ``timeout_seconds`` so a pathological workbook can't hang the tool.
    On timeout we fall back to delivering the file without cached values
    (Excel will recalc on open via the ``fullCalcOnLoad`` flag openpyxl
    writes).
    """
    # Lazy imports so the recalc code path is only paid for when used.
    try:
        from .formula_engine import (
            recalculate_workbook,
            detect_circular_references,
            is_available,
        )
        from .xml_cache import inject_cached_values
    except ImportError as e:
        logger.warning("Recalc modules unavailable: %s — skipping", e)
        return xlsx_bytes, None

    # Always run circular-reference detection, even if the recalc engine
    # is unavailable — it's pure graph analysis on the formula strings
    # and catches a bug class (cycles) that the engine silently misses.
    all_errors: list = []
    try:
        all_errors.extend(detect_circular_references(xlsx_bytes, sheet_names))
    except Exception as e:
        logger.debug("Circular-ref detection failed: %s", e)

    values_map: dict = {}

    if is_available():
        # Run the (potentially slow) engine in a worker thread with a hard
        # timeout. concurrent.futures is stdlib and avoids the asyncio dance.
        import concurrent.futures

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                future = pool.submit(recalculate_workbook, xlsx_bytes, sheet_names)
                result = future.result(timeout=timeout_seconds)
        except concurrent.futures.TimeoutError:
            logger.warning(
                "Formula recalc skipped: exceeded %ds timeout — delivering file "
                "without cached values (Excel will recalc on open)",
                timeout_seconds,
            )
            # Still report any circular refs we already found.
            summary = _format_grouped_errors(all_errors) if all_errors else None
            return xlsx_bytes, summary
        except Exception as e:
            logger.warning("Formula recalc skipped: engine error: %s", e)
            summary = _format_grouped_errors(all_errors) if all_errors else None
            return xlsx_bytes, summary

        if result.recalc_performed:
            values_map = result.values_map
            all_errors.extend(result.errors)
    else:
        logger.info("Formula recalc skipped: `formulas` library not installed")

    injected = inject_cached_values(xlsx_bytes, values_map)

    # Collect a human-readable note when the engine couldn't run at all
    # (unsupported function, parse failure, etc.). The result object carries
    # this as `skip_reason`; we surface it through the summary so an explicit
    # recalc request doesn't silently ship a file with no cached values and
    # no explanation — the model needs to know it should try rewriting the
    # offending formula (e.g. XLOOKUP → INDEX/MATCH).
    skip_note: str | None = None
    if is_available() and not result.recalc_performed and result.skip_reason:
        skip_note = (
            f"Formula recalculation could not be completed ({result.skip_reason}). "
            "The file was delivered without cached values; consider simplifying "
            "formulas (e.g. replacing newer functions like XLOOKUP/TEXTBEFORE "
            "with INDEX/MATCH or LEFT/FIND) so values can be cached for preview."
        )
        logger.warning("Recalc skip: %s", result.skip_reason)

    if all_errors:
        # total_formulas gives the model a sense of scope: "2/15 formulas
        # had errors" is more actionable than a bare count. Falls back to
        # counting <f> elements if the engine didn't populate it.
        total = result.total_formulas if (
            is_available() and result.recalc_performed
        ) else None
        if total is None:
            from .xml_cache import count_formulas
            total = count_formulas(injected)
        summary = _format_grouped_errors(all_errors, total_formulas=total)
        if skip_note:
            summary = f"{summary} — {skip_note}"
        logger.warning("Excel formula errors: %s", summary)
        return injected, summary

    if skip_note:
        return injected, skip_note

    return injected, None


def _format_grouped_errors(errors: list, total_formulas: int = 0) -> str:
    """Format a list of CellError objects into a type-grouped summary string.

    Produces output like:
        "2/15 formula error(s): #DIV/0! (2): Sheet!B2, Sheet!B5; #REF! (1): Sheet!C10"
    or with circular refs:
        "2/15 formula error(s): #CIRC! (2): Sheet!A1, Sheet!A2 — circular references detected"

    The ``total_formulas`` count (when known) is included as ``"N/total"`` so
    the model can tell whether it's fixing 2 errors out of 5 formulas or 2
    out of 500 — very different scopes. Omitted from the prefix when 0/unknown.

    Grouping by type makes it easier for the model to fix all errors of
    a given kind at once, rather than parsing a flat list.
    """
    from .formula_engine import CIRCULAR_ERROR_TYPE
    from collections import OrderedDict

    # Group by error_type, preserving first-seen order for stable output.
    grouped: "OrderedDict[str, list[str]]" = OrderedDict()
    for err in errors:
        location = err.location if hasattr(err, "location") else str(err)
        grouped.setdefault(err.error_type, []).append(location)

    parts: list[str] = []
    for error_type, locations in grouped.items():
        # Show up to 5 locations per type to keep the message readable.
        shown = locations[:5]
        loc_str = ", ".join(shown)
        if len(locations) > 5:
            loc_str += f" (and {len(locations) - 5} more)"
        parts.append(f"{error_type} ({len(locations)}): {loc_str}")

    # Prefix with N/total when the total formula count is known; this
    # helps the model gauge scope ("2/15" vs "2/500").
    if total_formulas and total_formulas > 0:
        prefix = f"{len(errors)}/{total_formulas} formula error(s)"
    else:
        prefix = f"{len(errors)} formula error(s)"
    summary = prefix + ": " + "; ".join(parts)

    # Annotate circular references explicitly since #CIRC! is our own
    # sentinel, not an Excel error the model will recognise.
    if CIRCULAR_ERROR_TYPE in grouped:
        summary += (
            " — circular references detected (a formula depends on itself, "
            "directly or indirectly; fix by breaking the cycle)"
        )

    return summary
