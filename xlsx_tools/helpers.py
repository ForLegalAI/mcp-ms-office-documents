import re
import logging
from dataclasses import dataclass
from datetime import datetime

from dateutil import parser as dateutil_parser
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# ── Layout Constants ──────────────────────────────────────────────────────────
TABLE_BOTTOM_SPACING = 2
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 25
COLUMN_WIDTH_PADDING = 2

# ── Financial Modeling Color Conventions ─────────────────────────────────────
# Industry-standard (CFA / Wall Street) color coding for financial models:
#   blue    — hardcoded inputs (numbers the user will change for scenarios)
#   black   — formulas / calculations within the same sheet
#   green   — formulas pulling from a different worksheet in the same workbook
#   red     — formulas pulling from a different workbook file (external link)
#   yellow  — key assumptions / cells flagged for review
# These constants are used only when the caller opts in via financial_modeling=True.
FINANCIAL_INPUT_COLOR = "0000FF"      # blue
FINANCIAL_FORMULA_COLOR = "000000"    # black
FINANCIAL_CROSS_SHEET_COLOR = "008000"  # green
FINANCIAL_EXTERNAL_COLOR = "FF0000"   # red — external workbook references
FINANCIAL_ASSUMPTION_FILL = "FFFF00"  # yellow background

# ── Number-Format Variants for the `types:` directive ────────────────────────
# These Excel number-format strings implement standard financial-modeling
# conventions: zeros render as "-", negatives in parentheses, etc.
#   positive ; negative ; zero
# is the canonical three-section Excel format pattern.
NUMBER_FORMAT_VARIANTS = {
    "dash": "#,##0;(#,##0);-",
    "parens": "#,##0;(#,##0)",
    "comma_dash": "#,##0;(#,##0);-",
    "default": "#,##0",
}

PERCENT_FORMAT_VARIANTS = {
    "dash": "0.0%;(0.0%);-",
    "parens": "0.0%;(0.0%)",
    "default": "0.0%",
    "legacy": "0%",  # original behaviour when no variant is given
}

# Valuation multiples (EV/EBITDA, P/E, etc.) are typically rendered as
# "12.5x" — the value with a trailing lowercase "x" suffix. One decimal
# place is the CFA convention. Variants follow the same dash/parens logic.
MULTIPLES_FORMAT_VARIANTS = {
    "default": '0.0"x"',
    "dash": '0.0"x";(0.0"x");-',
    "parens": '0.0"x";(0.0"x")',
}

# Date formats to try before falling back to dateutil auto-detection.
# Order matters — more specific/common formats first.
# Each entry: (strptime_format, excel_number_format)
DATE_FORMATS: list[tuple[str, str]] = [
    # ISO
    ("%Y-%m-%d", "YYYY-MM-DD"),
    ("%Y-%m-%dT%H:%M:%S", "YYYY-MM-DD HH:MM:SS"),
    ("%Y-%m-%dT%H:%M", "YYYY-MM-DD HH:MM"),
    # European (day first)
    ("%d.%m.%Y", "DD.MM.YYYY"),
    ("%d/%m/%Y", "DD/MM/YYYY"),
    ("%d-%m-%Y", "DD-MM-YYYY"),
    ("%d. %m. %Y", "DD. MM. YYYY"),
    # US (month first)
    ("%m/%d/%Y", "MM/DD/YYYY"),
    # With time
    ("%d.%m.%Y %H:%M", "DD.MM.YYYY HH:MM"),
    ("%d.%m.%Y %H:%M:%S", "DD.MM.YYYY HH:MM:SS"),
    ("%m/%d/%Y %H:%M", "MM/DD/YYYY HH:MM"),
    # Short year
    ("%d.%m.%y", "DD.MM.YY"),
    ("%d/%m/%y", "DD/MM/YY"),
    ("%m/%d/%y", "MM/DD/YY"),
    # Named months
    ("%d %b %Y", "DD MMM YYYY"),
    ("%d %B %Y", "DD MMMM YYYY"),
    ("%b %d, %Y", "MMM DD, YYYY"),
    ("%B %d, %Y", "MMMM DD, YYYY"),
]

# Minimum length to even attempt date parsing (avoids matching plain numbers)
_MIN_DATE_LENGTH = 6
# Regex to quickly reject values that clearly can't be dates
_DATE_CANDIDATE_RE = re.compile(r'^\d{1,4}[\.\-/]|^\d{1,2}\s+\w|^\w+\s+\d')


def _try_parse_date(value: str) -> tuple[datetime, str] | None:
    """Attempt to parse a string as a date/datetime.

    Tries explicit formats first (fast, unambiguous), then falls back to
    dateutil for natural language dates.

    Returns (datetime_obj, excel_number_format) or None.
    """
    if len(value) < _MIN_DATE_LENGTH:
        return None
    if not _DATE_CANDIDATE_RE.match(value):
        return None

    # Try explicit formats first (deterministic, no ambiguity)
    for fmt, xl_fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(value, fmt)
            return dt, xl_fmt
        except ValueError:
            continue

    # Fallback to dateutil (handles many international/natural formats)
    try:
        dt = dateutil_parser.parse(value, dayfirst=True, fuzzy=False)
        # Only accept if the string is sufficiently "date-like" —
        # dateutil can parse things like "1" or "March" alone which we don't want
        if dt and len(value) >= 8:
            # Determine appropriate format based on whether time is present
            if dt.hour or dt.minute or dt.second:
                return dt, "YYYY-MM-DD HH:MM:SS"
            return dt, "YYYY-MM-DD"
    except (ValueError, TypeError, OverflowError):
        pass

    return None


def _is_separator_row(line: str) -> bool:
    """Check if a table line is a markdown separator row (e.g. |---|:---:|---:|).

    Only returns True if ALL cells in the row match the separator pattern,
    preventing false positives from data cells that happen to contain '---'.
    """
    cells = [c.strip() for c in line.split('|')[1:-1]]
    if not cells:
        return False
    return all(re.match(r'^:?-{3,}:?$', c) for c in cells)


def _parse_column_alignments(separator_line: str) -> list[str | None]:
    """Extract column alignments from a markdown separator row.

    Returns a list of alignment strings ('left', 'center', 'right') or None per column.
    This is the same logic used by docx_tools but returns generic strings
    instead of Word-specific enums.
    """
    cells = [c.strip() for c in separator_line.split('|')[1:-1]]
    alignments: list[str | None] = []
    for cell in cells:
        cell = cell.strip()
        if cell.startswith(':') and cell.endswith(':'):
            alignments.append('center')
        elif cell.endswith(':'):
            alignments.append('right')
        elif cell.startswith(':'):
            alignments.append('left')
        else:
            alignments.append(None)  # auto — will use heuristic
    return alignments


def parse_table(lines: list[str], start_idx: int) -> tuple[list[list[str]] | None, int]:
    """Parse markdown table and return (table_data, next_index).

    Also extracts column alignments from the separator row and attaches them
    as the 'col_alignments' attribute on the returned TableData instance.
    """
    table_lines: list[str] = []
    i = start_idx

    # Find all consecutive table lines (allow missing trailing pipe)
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith('|'):
            # Normalize: ensure trailing pipe for consistent splitting
            if not line.endswith('|'):
                line = line + '|'
            table_lines.append(line)
            i += 1
        else:
            break

    if len(table_lines) < 2:  # Need at least header and separator
        return None, i if i > start_idx else start_idx + 1

    # Parse table data, extracting alignment from separator row
    table_data: list[list[str]] = []
    col_alignments: list[str | None] = []
    for line in table_lines:
        if _is_separator_row(line):
            col_alignments = _parse_column_alignments(line)
            continue
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        table_data.append(cells)

    # Attach alignment info to the table_data list
    table_data_with_align = TableData(table_data, col_alignments)
    return table_data_with_align, i


class TableData(list):
    """A list subclass that carries column alignment metadata."""

    def __init__(self, data: list[list[str]], col_alignments: list[str | None] | None = None):
        super().__init__(data)
        self.col_alignments: list[str | None] = col_alignments or []


# ── Cell Resolution ───────────────────────────────────────────────────────────

@dataclass
class CellResult:
    """Resolved cell metadata — all information needed to write a cell to Excel."""
    value: str | int | float | datetime  # The cleaned value to write
    is_formula: bool = False
    is_percent: bool = False
    is_date: bool = False
    date_format: str = ""  # Excel number format for dates (e.g. "YYYY-MM-DD")
    bold: bool = False
    italic: bool = False
    monospace: bool = False

    @property
    def formatting_info(self) -> dict[str, bool]:
        """Legacy-compatible formatting dict for apply_cell_formatting()."""
        return {'bold': self.bold, 'italic': self.italic, 'monospace': self.monospace}




def _strip_markdown_formatting(raw_text: str) -> tuple[str, dict[str, bool]]:
    """Strip inline markdown formatting markers from a cell value.

    Returns (clean_text, formatting_dict) where formatting_dict has
    'bold', 'italic', 'monospace' keys.
    """
    clean_text = raw_text.strip()
    formatting = {'bold': False, 'italic': False, 'monospace': False}

    if clean_text.startswith('**') and clean_text.endswith('**') and len(clean_text) > 4:
        clean_text = clean_text[2:-2]
        formatting['bold'] = True
    elif clean_text.startswith('*') and clean_text.endswith('*') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        formatting['italic'] = True
    elif clean_text.startswith('`') and clean_text.endswith('`') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        formatting['monospace'] = True

    return clean_text, formatting


def resolve_cell(raw_text: str) -> CellResult:
    """Parse a raw markdown cell string into a fully resolved CellResult.

    Combines formatting detection, formula detection, and type conversion
    in a single pass — the unified replacement for the former three-function pipeline
    of parse_cell_formatting → detect_formula_pattern → format_cell_value.
    """
    # Step 1: Strip markdown formatting markers
    clean_text, formatting = _strip_markdown_formatting(raw_text)
    bold = formatting['bold']
    italic = formatting['italic']
    monospace = formatting['monospace']

    # Step 2: Check if it's an explicit formula (= prefix)
    if clean_text.startswith('='):
        return CellResult(
            value=clean_text, is_formula=True,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 3: Detect percent and convert to number
    is_percent = clean_text.endswith('%')
    if is_percent:
        try:
            numeric_val = float(clean_text[:-1]) / 100
            return CellResult(
                value=numeric_val, is_percent=True,
                bold=bold, italic=italic, monospace=monospace,
            )
        except ValueError:
            pass  # Not a valid percent number — fall through

    # Step 5: Try numeric conversion
    try:
        numeric_val = float(_strip_thousands_separators(clean_text))
        return CellResult(
            value=numeric_val,
            bold=bold, italic=italic, monospace=monospace,
        )
    except ValueError:
        pass

    # Step 6: Try date detection (after numeric, so "2024" isn't parsed as a date)
    date_result = _try_parse_date(clean_text)
    if date_result:
        dt, xl_fmt = date_result
        return CellResult(
            value=dt, is_date=True, date_format=xl_fmt,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 7: Plain text
    return CellResult(
        value=clean_text,
        bold=bold, italic=italic, monospace=monospace,
    )


def apply_cell_formatting(cell, formatting_info: dict[str, bool]) -> None:
    """Apply formatting information to an Excel cell."""
    current_font = cell.font
    if formatting_info['bold']:
        cell.font = Font(bold=True, color=current_font.color, size=current_font.size)
    elif formatting_info['italic']:
        cell.font = Font(italic=True, color=current_font.color, size=current_font.size)
    elif formatting_info['monospace']:
        cell.font = Font(name='Courier New', color=current_font.color, size=current_font.size)


def apply_default_font(cell, font_name: str | None) -> None:
    """Apply a default font family to a cell unless it has a more specific font.

    Used to honour the tool-level ``default_font`` parameter. Skips cells
    whose current font is already a non-default one (e.g. Courier New
    from inline ``code`` formatting), so monospace cells are preserved.
    """
    if not font_name:
        return
    current = cell.font
    # Don't override monospace or explicitly-named fonts.
    if current.name and current.name != 'Calibri':
        return
    cell.font = Font(
        name=font_name,
        size=current.size,
        bold=current.bold,
        italic=current.italic,
        color=current.color,
    )


def apply_financial_styling(
    cell,
    cell_value_for_check: str | None,
    source_cells: set[str] | None = None,
) -> None:
    """Apply CFA-standard financial-modeling color coding to a cell.

    Industry conventions:
    - Hardcoded literal values  → blue font (inputs the user changes)
    - Local formulas            → black font (calculations on this sheet)
    - Cross-sheet formulas      → green font (links to other worksheets
                                  in the same workbook)
    - External-link formulas    → red font (links to other workbook files)
    - Source-cited cells        → yellow background (assumptions needing review)

    External-link detection: Excel references another workbook with square
    brackets around the filename, e.g. ``=[Forecast.xlsx]Sheet1!A1`` or
    ``='[Forecast.xlsx]Sheet1'!A1``. The presence of ``[`` after the
    leading ``=`` (ignoring quotes) is a reliable signal.

    Args:
        cell: The openpyxl cell to style.
        cell_value_for_check: The string form of the cell's resolved value
            (e.g. ``"=SUM(B2:B5)"`` or ``"1000"``). Used to detect
            formulas and cross-sheet references without re-parsing.
        source_cells: Optional set of ``"Sheet!Cell"`` location keys
            that should get the yellow assumption fill. If the cell's
            coordinate is in this set, the fill is applied.
    """
    current_font = cell.font

    # Default color for hardcoded values (inputs).
    color = FINANCIAL_INPUT_COLOR

    is_formula = isinstance(cell_value_for_check, str) and cell_value_for_check.startswith('=')
    if is_formula:
        # External workbook reference? Look for '[' after the '=' and any
        # leading quote. This must be checked BEFORE the cross-sheet '!'.
        stripped = cell_value_for_check.lstrip("='")
        if '[' in stripped:
            color = FINANCIAL_EXTERNAL_COLOR
        # Cross-sheet references contain a sheet-qualified cell ref like
        # ``SheetName!A1`` (with optional quotes if the name has spaces).
        elif '!' in cell_value_for_check:
            color = FINANCIAL_CROSS_SHEET_COLOR
        else:
            color = FINANCIAL_FORMULA_COLOR

    cell.font = Font(
        name=current_font.name,
        size=current_font.size,
        bold=current_font.bold,
        italic=current_font.italic,
        color=color,
    )

    # Yellow background for cells flagged as sourced assumptions.
    if source_cells:
        # Build the location key the way the parser stores it: just the
        # coordinate (e.g. "B5") since we don't know the sheet here.
        coord = cell.coordinate
        if coord in source_cells:
            cell.fill = PatternFill(
                start_color=FINANCIAL_ASSUMPTION_FILL,
                end_color=FINANCIAL_ASSUMPTION_FILL,
                fill_type="solid",
            )


def attach_source_comment(cell, source_text: str) -> None:
    """Attach a source-citation comment to a cell.

    The comment text typically follows the pattern
    ``"Source: <document>, <date>, <reference>, <URL>"`` but any string
    is accepted. The comment author is set to the server name so users
    can identify programmatically-attached comments.
    """
    if not source_text:
        return
    try:
        cell.comment = Comment(source_text, "MCP")
    except Exception as e:
        logger.debug("Failed to attach comment to %s: %s", cell.coordinate, e)


def parse_sources_directive(value: str) -> dict[str, str]:
    """Parse a ``<!-- sources: ... -->`` directive into a {coordinate: source} map.

    Supported syntaxes (all produce a mapping of cell coordinate → source text):

    1. Per-cell: ``B2=Source text here, B5=Another source``
       Commas separate entries; the first ``=`` in each entry divides
       coordinate from source text.

    2. Range form: ``B2:B5=Source applies to all these cells``
       Expands the Excel range to individual cells, each getting the
       same source text.

    When a single-cell entry and a range entry both cover the same cell,
    the single-cell entry wins (it is more specific). For example, in
    ``B2:B5=Range source, B3=Cell-specific source`` the cell B3 gets
    the cell-specific text while B2, B4, B5 get the range text. This
    also applies when the single-cell entry is listed first in the
    directive.

    The coordinate keys in the returned dict are bare coordinates
    (e.g. ``"B2"``) without a sheet prefix — the caller knows which
    sheet it's processing.
    """
    if not value:
        return {}
    # Two passes: ranges first (less specific), then singles (more
    # specific) so a single-cell entry always overrides a range that
    # happens to cover the same cell, regardless of declaration order.
    range_entries: list[tuple[str, str]] = []
    single_entries: list[tuple[str, str]] = []
    for entry in value.split(','):
        entry = entry.strip()
        if not entry or '=' not in entry:
            continue
        coord_part, source_text = entry.split('=', 1)
        coord_part = coord_part.strip()
        source_text = source_text.strip()
        if not coord_part or not source_text:
            continue
        if ':' in coord_part:
            range_entries.append((coord_part, source_text))
        else:
            single_entries.append((coord_part, source_text))
    result: dict[str, str] = {}
    for coord_part, source_text in range_entries:
        for coord in _expand_coord_range(coord_part):
            result[coord] = source_text
    for coord_part, source_text in single_entries:
        result[coord_part] = source_text
    return result


def _expand_coord_range(range_str: str) -> list[str]:
    """Expand an Excel range like 'B2:B5' into ['B2', 'B3', 'B4', 'B5']."""
    match = re.fullmatch(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str.strip().upper())
    if not match:
        return [range_str.strip().upper()]
    col_start, row_start, col_end, row_end = match.groups()
    if col_start != col_end:
        # Multi-column ranges are rare in source citations; just return
        # the endpoints rather than expanding the full cartesian product.
        return [f"{col_start}{row_start}", f"{col_end}{row_end}"]
    start_row, end_row = int(row_start), int(row_end)
    if start_row > end_row:
        start_row, end_row = end_row, start_row
    return [f"{col_start}{r}" for r in range(start_row, end_row + 1)]


# ── Formula Reference Resolution ─────────────────────────────────────────────

def _quote_sheet_name(name: str) -> str:
    """Return the sheet name quoted for Excel if it contains spaces or special chars."""
    if re.search(r"[^A-Za-z0-9_]", name):
        return f"'{name}'"
    return name


def _resolve_row(positions: dict[str, int], table_num: int, offset: int, fallback_row: int) -> int:
    """Resolve a table-relative row reference to an absolute Excel row number.

    Args:
        positions: Table positions dict ({"T1": start_row, ...}) for the target sheet.
        table_num: Table number (1-based).
        offset: Row offset within the table (0 = first data row).
        fallback_row: Row to use if the table isn't found in positions.

    Returns:
        The absolute Excel row number.

    A missing table key (e.g. ``T9`` when only 3 tables exist) is logged at
    WARNING level — the formula still resolves (using the fallback row) so the
    file ships, but the model gets a signal that a reference likely points at
    the wrong cell, which is the most common source of silently-wrong formulas.
    """
    key = f"T{table_num}"
    base = positions.get(key)
    if base is not None:
        return base + 1 + offset  # +1 to skip header row
    logger.warning(
        "Formula references %s but no such table exists in the target sheet "
        "(known tables: %s); falling back to current row. This likely produces "
        "a wrong cell reference — check the table numbering.",
        key, ", ".join(sorted(positions.keys())) or "none",
    )
    return fallback_row + offset


def _warn_unknown_sheet(sheet: str, all_sheet_table_positions: dict[str, dict[str, int]]) -> None:
    """Log a warning when a cross-sheet reference names a sheet that doesn't exist.

    The formula still resolves (the regex emits a syntactically valid cross-sheet
    ref), but Excel will show ``#REF!`` on open — better to surface the typo
    (e.g. ``Revenue!T1.B[0]`` when the sheet is actually ``Revenue Model``)
    during generation than let it fail silently in the client.
    """
    if sheet not in all_sheet_table_positions:
        known = ", ".join(sorted(all_sheet_table_positions.keys())) or "none"
        logger.warning(
            "Formula references sheet '%s' which does not exist in the workbook "
            "(known sheets: %s). The generated reference will likely resolve to "
            "#REF! in Excel.",
            sheet, known,
        )


def _make_cell_ref(column: str, row: int, sheet: str | None = None) -> str:
    """Build a cell reference string, optionally with a quoted sheet prefix."""
    if sheet:
        return f"{_quote_sheet_name(sheet)}!{column}{row}"
    return f"{column}{row}"


def adjust_formula_references(
    formula: str,
    current_excel_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
) -> str:
    """Convert row-relative references [offset] and table references T1.B[1] to actual Excel row numbers.

    Also resolves cross-sheet references like ``SheetName!T1.B[0]`` → ``'SheetName'!B2``.
    """
    if not formula.startswith('='):
        return formula

    if table_positions is None:
        table_positions = {}
    if all_sheet_table_positions is None:
        all_sheet_table_positions = {}

    logger.debug("Resolving formula: %s (current_row=%d)", formula, current_excel_row)

    try:
        # ── Cross-sheet references (must be resolved BEFORE local patterns) ──

        # Cross-sheet function: SheetName!T1.SUM(B[0]:E[0])
        cs_func_pattern = r"([\w\s.]+)!T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)"

        def _replace_cs_func(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            func_name = match.group(3)
            start_col = match.group(4)
            start_offset = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            _warn_unknown_sheet(sheet, all_sheet_table_positions)
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, table_num, start_offset, current_excel_row)
            er = _resolve_row(pos, table_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            # Excel range syntax allows the sheet prefix only ONCE, on the
            # first endpoint: =SUM(Data!B2:B4). Putting it on both endpoints
            # (=SUM(Data!B2:Data!B4)) is invalid and yields #VALUE!.
            result = f"{func_name}({qs}!{start_col}{sr}:{end_col}{er})"
            logger.debug("  Cross-sheet func: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_func_pattern, _replace_cs_func, formula)

        # Cross-sheet range: SheetName!T1.B[0]:T1.E[0]
        cs_range_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_range(match):
            sheet = match.group(1).strip()
            st_num = int(match.group(2))
            start_col = match.group(3)
            start_offset = int(match.group(4))
            et_num = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            _warn_unknown_sheet(sheet, all_sheet_table_positions)
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, st_num, start_offset, current_excel_row)
            er = _resolve_row(pos, et_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            result = f"{qs}!{start_col}{sr}:{end_col}{er}"
            logger.debug("  Cross-sheet range: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_range_pattern, _replace_cs_range, formula)

        # Cross-sheet single cell: SheetName!T1.B[0]
        cs_cell_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_cell(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            column = match.group(3)
            offset = int(match.group(4))
            _warn_unknown_sheet(sheet, all_sheet_table_positions)
            pos = all_sheet_table_positions.get(sheet, {})
            actual_row = _resolve_row(pos, table_num, offset, current_excel_row)
            result = _make_cell_ref(column, actual_row, sheet)
            logger.debug("  Cross-sheet cell: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_cell_pattern, _replace_cs_cell, formula)

        # ── Local (same-sheet) references ──
        # NOTE: Range and function patterns must be processed BEFORE single-cell
        # to prevent the single-cell regex from consuming parts of range expressions.

        # Table range references e.g. T1.B[0]:T1.E[0]
        table_range_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_range(match):
            start_table_num = int(match.group(1))
            start_col = match.group(2)
            start_offset = int(match.group(3))
            end_table_num = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, start_table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, end_table_num, end_offset, current_excel_row)
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(table_range_pattern, replace_table_range, formula)

        # Simplified function over table range e.g. T1.SUM(B[0]:E[0])
        table_func_pattern = r'T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)'

        def replace_table_function(match):
            table_num = int(match.group(1))
            func_name = match.group(2)
            start_col = match.group(3)
            start_offset = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, table_num, end_offset, current_excel_row)
            return f"{func_name}({start_col}{start_row}:{end_col}{end_row})"

        adjusted = re.sub(table_func_pattern, replace_table_function, adjusted)

        # Table cell references e.g. T1.B[1] (AFTER range patterns)
        table_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_reference(match):
            table_num = int(match.group(1))
            column = match.group(2)
            offset = int(match.group(3))
            actual_row = _resolve_row(table_positions, table_num, offset, current_excel_row)
            result = f"{column}{actual_row}"
            logger.debug("  Local table ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(table_pattern, replace_table_reference, adjusted)

        # Local row-relative references (e.g. ``B[0]``, ``A[-1]``) — the form
        # documented as "current row references" in the tool description. These
        # resolve relative to the row the formula lives in: ``B[0]`` → current
        # row's column B, ``B[-1]`` → the row above, ``B[1]`` → the row below.
        #
        # This is deliberately distinct from the table-relative ``T1.B[n]``
        # form (resolved above), which is offset from the table's FIRST data
        # row regardless of which row the formula is in. Conflating the two
        # (the previous implementation used ``current_table_start + 1 + offset``
        # here) silently broke every row past the first data row — e.g. ``A[0]``
        # on row 4 resolved to the first data cell instead of A4, corrupting
        # running totals, period-over-period growth, and cumulative sums.

        # Row-relative range e.g. B[0]:E[0] (BEFORE single-cell relative)
        range_pattern = r'([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]'

        def replace_range(match):
            start_col = match.group(1)
            start_offset = int(match.group(2))
            end_col = match.group(3)
            end_offset = int(match.group(4))
            start_row = current_excel_row + start_offset
            end_row = current_excel_row + end_offset
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(range_pattern, replace_range, adjusted)

        # Handle row-relative references e.g. B[0] (AFTER range pattern)
        rel_pattern = r'([A-Z]+)\[([+-]?\d+)\]'

        def replace_rel(match):
            column = match.group(1)
            offset = int(match.group(2))
            actual_row = current_excel_row + offset
            result = f"{column}{actual_row}"
            logger.debug("  Relative ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(rel_pattern, replace_rel, adjusted)

        logger.debug("  Resolved formula: %s → %s", formula, adjusted)
        return adjusted

    except Exception as e:
        logger.warning("Failed to adjust formula references for '%s': %s", formula, e)
        return formula


# ── Directive Helpers ──────────────────────────────────────────────────────────

# Currency symbols → Excel format string
_CURRENCY_FORMATS = {
    '$': '$#,##0.00',
    '€': '#,##0.00 €',
    '£': '£#,##0.00',
    '¥': '¥#,##0',
    'Kč': '#,##0.00 "Kč"',
    'zł': '#,##0.00 "zł"',
    'kr': '#,##0.00 "kr"',
    'CHF': '"CHF" #,##0.00',
    'R$': '"R$" #,##0.00',
    '₹': '₹#,##0.00',
}

# Zero-decimal variants for currency (`currency:<symbol>:integer`).
# Used for whole-unit financial figures ($mm, $bn) where decimals are noise.
_CURRENCY_ZERO_DECIMAL_FORMATS = {
    '$': '$#,##0',
    '€': '#,##0 €',
    '£': '£#,##0',
    '¥': '¥#,##0',
    'Kč': '#,##0 "Kč"',
    'zł': '#,##0 "zł"',
    'kr': '#,##0 "kr"',
    'CHF': '"CHF" #,##0',
    'R$': '"R$" #,##0',
    '₹': '₹#,##0',
}


def _parse_types_directive(value: str) -> list[str | None]:
    """Parse a types directive value like 'text, currency:$, date, bool, number'.

    Returns a list of type specs (or None for unspecified columns).

    Commas separate columns, but Excel number formats themselves contain
    commas (``#,##0``), so a naive ``split(',')`` shreds a literal format
    like ``number:#,##0.00`` into ``number:#`` + ``##0.00`` and shifts every
    later column by one (silent data corruption). We split on commas but
    re-join any fragment that does NOT start a new column spec back onto the
    previous one. A new column spec is either empty (unspecified column) or
    begins with one of the known type keywords; anything else is a
    continuation of the preceding fragment's format string.
    """
    if not value:
        return []
    fragments = value.split(',')
    specs: list[str | None] = []
    for frag in fragments:
        stripped = frag.strip()
        token = stripped.split(':', 1)[0].strip().lower()
        is_new_spec = (stripped == "") or (token in _KNOWN_TYPE_KEYWORDS)
        if is_new_spec or not specs:
            specs.append(stripped or None)
        else:
            # Continuation of a literal format that contained a comma —
            # re-join with the comma that split() consumed.
            prev = specs[-1] or ""
            specs[-1] = f"{prev},{frag}".strip() or None
    return specs


# Type keywords that legitimately start a new column spec. Used by
# _parse_types_directive to tell a real new column from a comma that lives
# inside a number format (e.g. the ',' in number:#,##0).
_KNOWN_TYPE_KEYWORDS = frozenset(
    {"text", "bool", "currency", "number", "date", "percent", "multiple"}
)


def _apply_column_type(cell, raw_text: str, type_spec: str | None) -> bool:
    """Apply column type coercion to a cell based on directive.

    Returns True if type was applied (caller should skip default processing),
    False if default processing should continue.
    """
    if not type_spec:
        return False

    clean = raw_text.strip()
    type_lower = type_spec.lower()

    # Alias: `number:multiple` / `number:multiples` → `multiple`. Users
    # naturally write `number:multiple` expecting "number formatted as a
    # multiple", but `multiple` is its own type, not a `number:` variant.
    # Without this rewrite the parser would treat `multiple` as a literal
    # number-format string (which Excel doesn't understand), leaving the
    # value as raw text '12.5x' and breaking formulas that reference it.
    if type_lower in ('number:multiple', 'number:multiples'):
        type_lower = 'multiple'
        type_spec = 'multiple'

    # text — force string, no conversion
    if type_lower == 'text':
        cell.value = clean
        return True

    # bool — map common boolean strings to Excel boolean
    if type_lower == 'bool':
        lower_val = clean.lower()
        if lower_val in ('true', 'yes', '1', 'on'):
            cell.value = True
        elif lower_val in ('false', 'no', '0', 'off'):
            cell.value = False
        else:
            cell.value = clean  # Unrecognized → keep as text
        return True

    # currency:<symbol> or currency:<symbol>:<variant> — strip symbol and thousands separators
    if type_lower.startswith('currency'):
        parts = type_spec.split(':')
        symbol = parts[1].strip() if len(parts) > 1 and parts[1].strip() else '$'
        variant = parts[2].strip().lower() if len(parts) > 2 else None
        # Strip the currency symbol and common thousand separators
        numeric_str = clean.replace(symbol, '').replace(' ', '').strip()
        # Detect accounting-style negative notation: (1234) → -1234.
        # Strip the parens and remember to negate the final value.
        is_negative = False
        if numeric_str.startswith('(') and numeric_str.endswith(')'):
            numeric_str = numeric_str[1:-1]
            is_negative = True
        # Handle both comma-as-thousands (1,234.56) and dot-as-thousands (1.234,56)
        if ',' in numeric_str and '.' in numeric_str:
            # Determine which is the decimal separator (last one wins)
            last_comma = numeric_str.rfind(',')
            last_dot = numeric_str.rfind('.')
            if last_comma > last_dot:
                # European: 1.234,56
                numeric_str = numeric_str.replace('.', '').replace(',', '.')
            else:
                # English: 1,234.56
                numeric_str = numeric_str.replace(',', '')
        elif ',' in numeric_str and '.' not in numeric_str:
            # Could be thousands (1,234) or decimal (1,5) — assume thousands if >3 digits after comma
            parts_n = numeric_str.split(',')
            if len(parts_n[-1]) == 3:
                numeric_str = numeric_str.replace(',', '')
            else:
                numeric_str = numeric_str.replace(',', '.')
        try:
            value = float(numeric_str)
            if is_negative:
                value = -abs(value)
            cell.value = value
            base_format = _currency_base_format(symbol, variant)
            cell.number_format = _apply_format_variant(base_format, variant)
        except ValueError:
            cell.value = clean  # Can't parse → keep as text
        return True

    # number, number:<format>, or number:<variant> — parse as number with optional format
    if type_lower.startswith('number'):
        parts = type_spec.split(':', 1)
        fmt_or_variant = parts[1].strip() if len(parts) > 1 else None
        numeric_str = _strip_thousands_separators(clean)
        try:
            cell.value = float(numeric_str)
            if fmt_or_variant:
                # If it's a known variant keyword, apply the variant format;
                # otherwise treat as a literal format string.
                if fmt_or_variant.lower() in NUMBER_FORMAT_VARIANTS:
                    cell.number_format = NUMBER_FORMAT_VARIANTS[fmt_or_variant.lower()]
                else:
                    cell.number_format = fmt_or_variant
            else:
                # No explicit format: pick by magnitude, preserving decimals
                # (was a bug: everything >= 1000 was force-rounded to #,##0).
                cell.number_format = DEFAULT_NUMBER_FORMAT if cell.value.is_integer() else DEFAULT_NUMBER_FORMAT_DECIMALS
        except ValueError:
            cell.value = clean
        return True

    # date or date:<format> — parse with dateutil, apply format
    if type_lower.startswith('date'):
        fmt = type_spec.split(':', 1)[1].strip() if ':' in type_spec else None
        result = _try_parse_date(clean)
        if result:
            dt, default_fmt = result
            cell.value = dt
            cell.number_format = fmt or default_fmt
        else:
            cell.value = clean
        return True

    # percent or percent:<variant> — parse as percent
    if type_lower.startswith('percent'):
        parts = type_spec.split(':', 1)
        variant = parts[1].strip().lower() if len(parts) > 1 else None
        numeric_str = clean.rstrip('%').strip()
        try:
            cell.value = float(numeric_str) / 100
            cell.number_format = _apply_percent_format_variant(variant)
        except ValueError:
            cell.value = clean
        return True

    # multiple, multiple:<variant>, or multiple:<decimals> — valuation
    # multiples (EV/EBITDA, P/E, etc.) rendered as "12.5x". The stored
    # value is the raw multiple (NOT divided); the "x" is purely a
    # display suffix via the number format.
    if type_lower.startswith('multiple'):
        parts = type_spec.split(':', 1)
        variant = parts[1].strip().lower() if len(parts) > 1 else None
        # Strip a trailing "x" if the user wrote "12.5x" in the data.
        numeric_str = clean.rstrip('x').rstrip('X').strip()
        try:
            cell.value = float(numeric_str)
            cell.number_format = _apply_multiples_format_variant(variant)
        except ValueError:
            cell.value = clean
        return True

    return False


def _number_format_for_type(type_spec: str | None) -> str | None:
    """Return the Excel number format a column ``types`` spec implies, or None.

    Used to format a *formula* cell that sits in a typed column. Formula
    cells bypass ``_apply_column_type`` (their value is a formula string,
    not a literal to coerce), so they'd otherwise lose the column's
    intended number format. This mirrors the format-selection logic in
    ``_apply_column_type`` without touching the cell value. Returns None
    for types that have no numeric format (text/bool) or unknown specs.
    """
    if not type_spec:
        return None
    type_lower = type_spec.lower()
    if type_lower in ('number:multiple', 'number:multiples'):
        type_lower = 'multiple'
        type_spec = 'multiple'

    if type_lower in ('text', 'bool'):
        return None

    if type_lower.startswith('currency'):
        parts = type_spec.split(':')
        symbol = parts[1].strip() if len(parts) > 1 and parts[1].strip() else '$'
        variant = parts[2].strip().lower() if len(parts) > 2 else None
        return _apply_format_variant(_currency_base_format(symbol, variant), variant)

    if type_lower.startswith('number'):
        parts = type_spec.split(':', 1)
        fmt_or_variant = parts[1].strip() if len(parts) > 1 else None
        if fmt_or_variant:
            if fmt_or_variant.lower() in NUMBER_FORMAT_VARIANTS:
                return NUMBER_FORMAT_VARIANTS[fmt_or_variant.lower()]
            return fmt_or_variant
        return None  # let the magnitude-based default apply

    if type_lower.startswith('date'):
        return type_spec.split(':', 1)[1].strip() if ':' in type_spec else None

    if type_lower.startswith('percent'):
        parts = type_spec.split(':', 1)
        variant = parts[1].strip().lower() if len(parts) > 1 else None
        return _apply_percent_format_variant(variant)

    if type_lower.startswith('multiple'):
        parts = type_spec.split(':', 1)
        variant = parts[1].strip().lower() if len(parts) > 1 else None
        return _apply_multiples_format_variant(variant)

    return None


def _strip_thousands_separators(s: str) -> str:
    """Normalize a numeric string that may contain thousands separators.

    Handles English (``1,234.56``), European (``1.234,56``), and bare
    thousands (``1,234``). Used so plain numeric cells without a `types`
    directive still parse when the value includes commas — previously
    ``float("1,234")`` raised and the value stayed a string.
    """
    s = s.strip()
    if ',' not in s and '.' not in s:
        return s
    if ',' in s and '.' in s:
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_comma > last_dot:
            # European: 1.234,56
            return s.replace('.', '').replace(',', '.')
        # English: 1,234.56
        return s.replace(',', '')
    if ',' in s:
        # Ambiguous: assume thousands if exactly 3 digits follow the comma.
        parts = s.split(',')
        if len(parts[-1]) == 3:
            return s.replace(',', '')
        return s.replace(',', '.')
    return s


def _apply_format_variant(base_format: str, variant: str | None) -> str:
    """Apply a financial-modeling variant (dash, parens) to a base number format.

    For currency/number formats, the variants control how zeros and
    negatives render. We keep the positive section of ``base_format`` and
    rewrite the negative/zero sections according to the variant.
    """
    if not variant or variant == "default":
        return base_format
    # Split on ';' to isolate sections. Excel formats can have up to 4
    # sections (positive;negative;zero;text). Most base formats here have
    # just one section, so we synthesise the rest.
    sections = base_format.split(';')
    positive = sections[0]
    if variant == "dash":
        return f"{positive};({positive});-"
    if variant == "parens":
        return f"{positive};({positive})"
    if variant == "comma_dash":
        return f"{positive};({positive});-"
    return base_format


# Variant keywords that select the zero-decimal currency format. Financial
# models frequently show whole-dollar amounts (revenue/profit in $mm, $bn)
# where the `.00` is noise. `currency:$:integer` (or `:int`/`:whole`) maps
# these to the symbol's zero-decimal base format.
_INTEGER_VARIANT_KEYWORDS = {"integer", "int", "whole"}


def _currency_base_format(symbol: str, variant: str | None) -> str:
    """Return the base currency number format for a symbol and variant.

    The variant selects precision: ``integer``/``int``/``whole`` (or any
    variant combined with ``integer``) yields the zero-decimal form
    (``$#,##0``), otherwise the default two-decimal form (``$#,##0.00``).
    Unknown symbols fall back to ``#,##0.00 "<symbol>"`` / ``#,##0 "<symbol>"``.
    """
    zero_decimal = bool(variant) and any(
        kw in variant for kw in _INTEGER_VARIANT_KEYWORDS
    )
    if zero_decimal:
        return _CURRENCY_ZERO_DECIMAL_FORMATS.get(
            symbol, f'#,##0 "{symbol}"'
        )
    return _CURRENCY_FORMATS.get(symbol, f'#,##0.00 "{symbol}"')


def _apply_percent_format_variant(variant: str | None) -> str:
    """Return the Excel percent number format for the given variant.

    Defaults to ``0.0%`` (one decimal) per the CFA/financial-modeling
    convention — bare ``50.5%`` should display as ``50.5%``, not ``51%``
    (which is what ``0%`` would produce, silently losing the decimal).
    Use the explicit ``percent:integer`` variant for the old no-decimal
    ``0%`` behaviour when backward compatibility with a pre-existing
    template matters.
    """
    if not variant:
        return "0.0%"  # CFA convention: one decimal
    if variant == "integer":
        return "0%"  # opt-out for users who want no decimals
    if variant == "default":
        return PERCENT_FORMAT_VARIANTS["default"]
    return PERCENT_FORMAT_VARIANTS.get(variant, "0.0%")


def _apply_multiples_format_variant(variant: str | None) -> str:
    """Return the Excel valuation-multiple format for the given variant.

    Defaults to ``0.0"x"`` (e.g. ``12.5x``) per CFA convention. The
    ``dash`` and ``parens`` variants follow the same negative/zero
    rendering rules as the other format families.
    """
    if not variant or variant == "default":
        return MULTIPLES_FORMAT_VARIANTS["default"]
    return MULTIPLES_FORMAT_VARIANTS.get(variant, MULTIPLES_FORMAT_VARIANTS["default"])


# ── Default number formats for plain numeric cells (no `types` directive) ─────
#
# When a column has no explicit `types` directive, `resolve_cell` returns a
# bare numeric value with no number_format. The rendering block in
# `add_table_to_sheet` then assigns a default based on magnitude. These
# formats implement two conventions:
#
#   - Whole numbers → `#,##0` (no trailing `.00` noise, no precision lost).
#   - Non-whole numbers → `#,##0.00` (preserves decimals; the old code
#     unconditionally used `#,##0` for values >= 1000, silently rounding
#     `1500.75` to `1,501`).
#
# When financial_modeling is active, the same defaults are promoted to the
# dash variant so zeros render as `-` and negatives in parentheses — the
# standard financial-modeling convention.
DEFAULT_NUMBER_FORMAT = "#,##0"
DEFAULT_NUMBER_FORMAT_DECIMALS = "#,##0.00"

# Financial dash/parens variants applied as the default number format when
# `financial_modeling=True` and the cell has no explicit format from a
# `types` directive. These are the same strings used by the corresponding
# variants in NUMBER_FORMAT_VARIANTS / PERCENT_FORMAT_VARIANTS, inlined
# here for the "no directive" default path.
FINANCIAL_DEFAULT_NUMBER = "#,##0;(#,##0);-"
FINANCIAL_DEFAULT_NUMBER_DECIMALS = "#,##0.00;(#,##0.00);-"
FINANCIAL_DEFAULT_PERCENT = "0.0%;(0.0%);-"


def _default_number_format_for(value: float, financial_modeling: bool) -> str:
    """Pick the default number format for a plain numeric cell.

    Whole numbers use the integer format; non-whole numbers preserve two
    decimals (was a bug: anything >= 1000 was force-rounded). When
    ``financial_modeling`` is true, the dash/parens variant is returned so
    zeros render as ``-`` and negatives in parentheses per CFA convention.
    """
    is_whole = float(value).is_integer()
    if financial_modeling:
        return FINANCIAL_DEFAULT_NUMBER if is_whole else FINANCIAL_DEFAULT_NUMBER_DECIMALS
    return DEFAULT_NUMBER_FORMAT if is_whole else DEFAULT_NUMBER_FORMAT_DECIMALS


# Plausible year window for the financial-modeling "years as text" convention.
# Excel's own date system starts at 1900, and forward-looking models rarely
# project past 2100. Restricting to this window keeps the convention (so a
# column of period headers like 2024, 2025, 2026E stays as text labels and
# doesn't get summed/formatted as the number 2,024) while NOT coercing
# legitimate 4-digit magnitudes — revenue of $1,500mm, a count of 5000, an
# ID of 1234 — to text, which would silently break SUM in real Excel and
# drop their number formatting.
_YEAR_MIN = 1900
_YEAR_MAX = 2100


def _is_year_string(value: str) -> bool:
    """Return True if value is a 4-digit year like '2024' within a plausible range.

    Restricted to ``_YEAR_MIN``–``_YEAR_MAX`` so that non-year 4-digit numbers
    (revenue, counts, IDs) are not mistakenly forced to text. See the
    ``_YEAR_MIN`` / ``_YEAR_MAX`` constants for the rationale.
    """
    stripped = value.strip()
    if not re.fullmatch(r"\d{4}", stripped):
        return False
    n = int(stripped)
    return _YEAR_MIN <= n <= _YEAR_MAX


# ── Table Rendering ───────────────────────────────────────────────────────────

def add_table_to_sheet(
    table_data: list[list[str]],
    worksheet,
    start_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
    auto_filter: bool = False,
    table_index: int = 0,
    directives: dict[str, str] | None = None,
    default_font: str | None = None,
    financial_modeling: bool = False,
) -> int:
    """Add table data to Excel worksheet with proper formatting and formula support."""
    if not table_data:
        return start_row

    directives = directives or {}

    # Parse column type hints from <!-- types: text, currency:$, date, bool --> directive
    col_types: list[str | None] = _parse_types_directive(directives.get('types', ''))

    # Parse source-citation directive: <!-- sources: B2=Source text, B5=Another -->
    sources_map: dict[str, str] = parse_sources_directive(directives.get('sources', ''))

    # Extract column alignments if available (from TableData subclass)
    col_alignments: list[str | None] = []
    if hasattr(table_data, 'col_alignments'):
        col_alignments = table_data.col_alignments

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    formula_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # When financial modeling is active, the formula_fill (light-blue) is
    # replaced by per-cell color coding via apply_financial_styling().

    # Fill cells
    for row_idx, row_data in enumerate(table_data):
        current_excel_row = start_row + row_idx
        for col_idx, cell_text in enumerate(row_data):
            try:
                cell = worksheet.cell(row=current_excel_row, column=col_idx + 1)
                # Apply the user-selected default font family first. Subsequent
                # font operations (header styling, financial color coding,
                # inline formatting) inherit the family unless they override it.
                if default_font and row_idx > 0:
                    apply_default_font(cell, default_font)

                # If column type directive applies (data rows only), use it
                col_type = col_types[col_idx] if col_idx < len(col_types) else None
                # A formula cell (leading '=' after stripping markdown) must
                # NOT go through type coercion — its references still need
                # resolving via adjust_formula_references, and float() on
                # '=SUM(B[-1])' raises and leaves the unresolved literal in
                # place (which is #NAME? in Excel and aborts the recalc
                # engine). Fall through to the normal formula path; the
                # column's number format is re-applied to the result below.
                clean_for_formula_check, _ = _strip_markdown_formatting(cell_text)
                is_formula_cell = clean_for_formula_check.startswith('=')
                if row_idx > 0 and col_type and not is_formula_cell:
                    # Strip markdown formatting before type coercion
                    clean_text, fmt_info = _strip_markdown_formatting(cell_text)
                    if _apply_column_type(cell, clean_text, col_type):
                        # Type directive handled the cell value — apply formatting, border, alignment
                        apply_cell_formatting(cell, fmt_info)
                        cell.border = border
                        explicit_align = col_alignments[col_idx] if col_idx < len(col_alignments) else None
                        if explicit_align:
                            cell.alignment = Alignment(horizontal=explicit_align)
                        elif isinstance(cell.value, bool):
                            cell.alignment = Alignment(horizontal='center')
                        elif isinstance(cell.value, (int, float, datetime)):
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')
                        # Financial color coding for typed cells (treat as input).
                        if financial_modeling:
                            apply_financial_styling(cell, None, set(sources_map.keys()))
                        # Source citation comment.
                        coord = cell.coordinate
                        if coord in sources_map:
                            attach_source_comment(cell, sources_map[coord])
                        continue

                resolved = resolve_cell(cell_text)

                if resolved.is_formula:
                    adjusted_formula = adjust_formula_references(
                        resolved.value, current_excel_row, table_positions, all_sheet_table_positions
                    )
                    cell.value = adjusted_formula
                    if not financial_modeling:
                        cell.fill = formula_fill
                    # A formula in a typed column gets the column's intended
                    # number format applied to its (numeric) result — e.g. a
                    # =SUM(...) in a `currency:$` column displays as currency.
                    # The format is harmless on non-numeric results (Excel
                    # ignores number formats on strings/errors).
                    if row_idx > 0 and col_type:
                        type_fmt = _number_format_for_type(col_type)
                        if type_fmt:
                            cell.number_format = type_fmt
                else:
                    # Header row must remain as strings — Excel Tables require
                    # string headers; numeric-looking headers (e.g. "2024") must
                    # not be converted to numbers.
                    if row_idx == 0:
                        # Use the original stripped text for headers to avoid
                        # artifacts like "2024.0" from float conversion
                        clean_header, _ = _strip_markdown_formatting(cell_text)
                        cell.value = clean_header
                    else:
                        # Financial-modeling convention: 4-digit year strings
                        # (e.g. "2024") in data rows are treated as text labels
                        # so they don't get auto-converted to numbers and lose
                        # their formatting in charts/pivots.
                        stripped = cell_text.strip()
                        if financial_modeling and _is_year_string(stripped):
                            clean_year, _ = _strip_markdown_formatting(cell_text)
                            cell.value = str(clean_year).strip()
                        else:
                            cell.value = resolved.value

                # Apply inline formatting (bold/italic/monospace) — skip for header row
                # since header styling will override it immediately below
                if row_idx > 0:
                    apply_cell_formatting(cell, resolved.formatting_info)
                cell.border = border

                # Alignment — use explicit column alignment from separator if available,
                # otherwise fall back to heuristic
                explicit_align = col_alignments[col_idx] if col_idx < len(col_alignments) else None
                if row_idx == 0:
                    cell.alignment = Alignment(horizontal='center')
                elif explicit_align:
                    cell.alignment = Alignment(horizontal=explicit_align)
                elif isinstance(cell.value, (int, float, datetime)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Header row styling (overrides inline formatting)
                if row_idx == 0:
                    if default_font:
                        cell.font = Font(name=default_font, bold=True, color="FFFFFF")
                    else:
                        cell.font = header_font
                    cell.fill = header_fill
                elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    # Default number format for plain numeric cells (no `types`
                    # directive on this column). Whole numbers use the integer
                    # format; non-whole numbers preserve two decimals (the old
                    # code force-rounded everything >= 1000 to `#,##0`,
                    # silently turning 1500.75 into 1,501). When financial
                    # modeling is active, the dash/parens variant is applied so
                    # zeros render as `-` and negatives in parentheses per CFA
                    # convention.
                    cell.number_format = _default_number_format_for(
                        cell.value, financial_modeling
                    )

                # Apply percentage number format. Per CFA convention the default
                # is one decimal (`0.0%`); in financial mode the dash variant is
                # used so zero renders as `-`.
                if resolved.is_percent and isinstance(cell.value, (int, float)):
                    cell.number_format = (
                        FINANCIAL_DEFAULT_PERCENT
                        if financial_modeling
                        else "0.0%"
                    )

                # Apply date number format
                if resolved.is_date and resolved.date_format:
                    cell.number_format = resolved.date_format

                # Financial-modeling color coding (data rows only, after all
                # other styling so the font color wins).
                if financial_modeling and row_idx > 0:
                    value_for_check = cell.value if isinstance(cell.value, str) else None
                    apply_financial_styling(cell, value_for_check, set(sources_map.keys()))
                    # Financial models also expect dash/parens number formatting
                    # on formula results (zero → "-", negative → "(...)"). The
                    # result value isn't known at render time, so we apply the
                    # format optimistically — Excel ignores number formats on
                    # non-numeric (string/error) results, so this is safe.
                    if (
                        isinstance(cell.value, str)
                        and cell.value.startswith('=')
                        and cell.number_format == 'General'
                    ):
                        cell.number_format = FINANCIAL_DEFAULT_NUMBER

                # Source citation comment (data rows only).
                if row_idx > 0:
                    coord = cell.coordinate
                    if coord in sources_map:
                        attach_source_comment(cell, sources_map[coord])
            except Exception as e:
                logger.warning("Error processing cell [row=%d, col=%d]: %s", current_excel_row, col_idx + 1, e)

    # Column widths — based on clean text length (not raw markdown with formatting markers)
    # When type directives are active, estimate display width from the type spec.
    FORMULA_WIDTH_CAP = 12  # Formulas display as numbers, cap their width contribution
    for col_idx in range(len(table_data[0]) if table_data else 0):
        column_letter = get_column_letter(col_idx + 1)
        col_type = col_types[col_idx] if col_idx < len(col_types) else None
        max_length = 0
        for row_idx, row in enumerate(table_data):
            if col_idx < len(row):
                # For data rows with a type directive, estimate from the directive
                if row_idx > 0 and col_type:
                    type_lower = col_type.lower()
                    if type_lower == 'bool':
                        length = 5  # "FALSE" is longest
                    elif type_lower.startswith('currency'):
                        # Symbol + number — use raw text length as decent estimate
                        length = len(row[col_idx].strip())
                    elif type_lower.startswith('date'):
                        fmt = col_type.split(':', 1)[1].strip() if ':' in col_type else "YYYY-MM-DD"
                        length = len(fmt)
                    elif type_lower == 'percent':
                        length = 6  # e.g. "85.0%"
                    else:
                        length = len(row[col_idx].strip())
                else:
                    resolved = resolve_cell(row[col_idx])
                    if resolved.is_formula:
                        length = FORMULA_WIDTH_CAP
                    elif resolved.is_date:
                        length = len(resolved.date_format)
                    else:
                        length = len(str(resolved.value))
                max_length = max(max_length, length)
        adjusted_width = min(max(max_length + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Auto-filter: create a proper Excel Table object (supports multiple per sheet)
    if auto_filter:
        num_cols = len(table_data[0]) if table_data else 0
        if num_cols > 0:
            last_col_letter = get_column_letter(num_cols)
            last_data_row = start_row + len(table_data) - 1
            table_ref = f"A{start_row}:{last_col_letter}{last_data_row}"
            # Excel table names must be unique across the workbook
            table_name = f"Table_{worksheet.title.replace(' ', '_')}_{table_index + 1}"
            # Sanitize: Excel table names allow only letters, digits, underscores
            table_name = re.sub(r'[^A-Za-z0-9_]', '', table_name)
            excel_table = Table(displayName=table_name, ref=table_ref)
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            worksheet.add_table(excel_table)

    return start_row + len(table_data) + TABLE_BOTTOM_SPACING
